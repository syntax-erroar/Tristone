#!/usr/bin/env python3
"""
Universal HTML Table Extractor and Excel Reconstructor

This tool can extract tables from ANY HTML file and perfectly reconstruct them in Excel,
preserving the original structure, formatting, and layout.

Features:
- Universal HTML table detection and extraction
- Advanced table structure analysis (colspan, rowspan, nested tables)
- Perfect Excel reconstruction with original formatting
- Support for complex SEC filings and any HTML table format
- Multi-parser approach for maximum compatibility
"""

import os
import re
import json
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass
from pathlib import Path

# Core libraries
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# HTML parsing libraries
from bs4 import BeautifulSoup
from lxml import etree, html
try:
    import html5lib
except ImportError:
    html5lib = None

# Additional utilities
import numpy as np
from collections import defaultdict
import logging
import requests
import time
from urllib.parse import urljoin

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class TableCell:
    """Represents a single table cell with all its properties"""
    content: str
    row: int
    col: int
    colspan: int = 1
    rowspan: int = 1
    is_header: bool = False
    styles: Dict[str, Any] = None
    width: Optional[float] = None
    height: Optional[float] = None
    alignment: Optional[str] = None
    font_weight: Optional[str] = None
    background_color: Optional[str] = None
    border_style: Optional[str] = None

@dataclass
class TableStructure:
    """Represents the complete structure of a table"""
    cells: List[TableCell]
    rows: int
    cols: int
    title: Optional[str] = None
    caption: Optional[str] = None
    table_style: Dict[str, Any] = None
    has_headers: bool = False
    is_financial: bool = False

class UniversalHTMLTableExtractor:
    """Universal HTML table extractor that works with any HTML file"""
    
    def __init__(self):
        self.parsers = {
            'lxml': self._parse_with_lxml,
            'beautifulsoup': self._parse_with_beautifulsoup
        }
        
        # Add html5lib parser if available
        if html5lib is not None:
            self.parsers['html5lib'] = self._parse_with_html5lib
        
        self.extracted_tables = []
        
    def extract_tables_from_html(self, html_file_path: str) -> List[TableStructure]:
        """
        Extract all tables from an HTML file using the best available parser.
        
        Args:
            html_file_path (str): Path to the HTML file
            
        Returns:
            List[TableStructure]: List of extracted table structures
        """
        logger.info(f"Extracting tables from: {html_file_path}")
        
        with open(html_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            html_content = f.read()
        
        # Try different parsers in order of preference
        for parser_name, parser_func in self.parsers.items():
            try:
                logger.info(f"Trying parser: {parser_name}")
                tables = parser_func(html_content)
                if tables:
                    logger.info(f"Successfully extracted {len(tables)} tables using {parser_name}")
                    self.extracted_tables.extend(tables)
                    return tables
            except Exception as e:
                logger.warning(f"Parser {parser_name} failed: {e}")
                continue
        
        logger.error("All parsers failed to extract tables")
        return []
    
    def extract_tables_from_html_content(self, html_content: str) -> List[TableStructure]:
        """Extract tables from HTML content string"""
        # Create temporary file for processing
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            f.write(html_content)
            temp_file = f.name
        
        try:
            tables = self.extract_tables_from_html(temp_file)
            return tables
        finally:
            # Clean up temporary file
            import os
            os.unlink(temp_file)
    
    def _parse_with_lxml(self, html_content: str) -> List[TableStructure]:
        """Parse HTML using lxml (best for complex/malformed HTML)"""
        try:
            # Parse with lxml
            doc = html.fromstring(html_content)
            
            # Find all table elements
            tables = doc.xpath('//table')
            extracted_tables = []
            
            for i, table in enumerate(tables):
                table_structure = self._extract_table_structure_lxml(table, i)
                if table_structure:
                    extracted_tables.append(table_structure)
            
            return extracted_tables
            
        except Exception as e:
            logger.error(f"lxml parsing failed: {e}")
            return []
    
    def _parse_with_html5lib(self, html_content: str) -> List[TableStructure]:
        """Parse HTML using html5lib (best for modern HTML5)"""
        try:
            soup = BeautifulSoup(html_content, 'html5lib')
            return self._extract_tables_from_soup(soup)
        except Exception as e:
            logger.error(f"html5lib parsing failed: {e}")
            return []
    
    def _parse_with_beautifulsoup(self, html_content: str) -> List[TableStructure]:
        """Parse HTML using BeautifulSoup (fallback)"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            return self._extract_tables_from_soup(soup)
        except Exception as e:
            logger.error(f"BeautifulSoup parsing failed: {e}")
            return []
    
    def _extract_tables_from_soup(self, soup: BeautifulSoup) -> List[TableStructure]:
        """Extract tables from BeautifulSoup object"""
        tables = soup.find_all('table')
        extracted_tables = []
        
        for i, table in enumerate(tables):
            table_structure = self._extract_table_structure_soup(table, i)
            if table_structure:
                extracted_tables.append(table_structure)
        
        return extracted_tables
    
    def _extract_table_structure_lxml(self, table_element, table_index: int) -> Optional[TableStructure]:
        """Extract table structure using lxml"""
        try:
            # Get table title/caption
            title = self._get_table_title_lxml(table_element)
            
            # Find all rows
            rows = table_element.xpath('.//tr')
            if not rows:
                return None
            
            cells = []
            max_cols = 0
            current_row = 0
            
            for row_idx, row in enumerate(rows):
                # Get all cells in this row (td and th)
                row_cells = row.xpath('.//td | .//th')
                
                current_col = 0
                for cell in row_cells:
                    # Extract cell content
                    content = self._extract_cell_content_lxml(cell)
                    
                    # Get cell attributes
                    colspan = int(cell.get('colspan', 1))
                    rowspan = int(cell.get('rowspan', 1))
                    is_header = cell.tag == 'th'
                    
                    # Extract styles
                    styles = self._extract_cell_styles_lxml(cell)
                    
                    # Create cell object
                    cell_obj = TableCell(
                        content=content,
                        row=current_row,
                        col=current_col,
                        colspan=colspan,
                        rowspan=rowspan,
                        is_header=is_header,
                        styles=styles
                    )
                    
                    cells.append(cell_obj)
                    current_col += colspan
                
                max_cols = max(max_cols, current_col)
                current_row += 1
            
            # Determine if this is a financial table
            is_financial = self._is_financial_table(cells)
            
            return TableStructure(
                cells=cells,
                rows=current_row,
                cols=max_cols,
                title=title,
                has_headers=any(cell.is_header for cell in cells),
                is_financial=is_financial
            )
            
        except Exception as e:
            logger.error(f"Failed to extract table structure: {e}")
            return None
    
    def _extract_table_structure_soup(self, table_element, table_index: int) -> Optional[TableStructure]:
        """Extract table structure using BeautifulSoup"""
        try:
            # Get table title/caption
            title = self._get_table_title_soup(table_element)
            
            # Find all rows
            rows = table_element.find_all('tr')
            if not rows:
                return None
            
            cells = []
            max_cols = 0
            current_row = 0
            
            for row_idx, row in enumerate(rows):
                # Get all cells in this row (td and th)
                row_cells = row.find_all(['td', 'th'])
                
                current_col = 0
                for cell in row_cells:
                    # Extract cell content
                    content = self._extract_cell_content_soup(cell)
                    
                    # Get cell attributes
                    colspan = int(cell.get('colspan', 1))
                    rowspan = int(cell.get('rowspan', 1))
                    is_header = cell.name == 'th'
                    
                    # Extract styles
                    styles = self._extract_cell_styles_soup(cell)
                    
                    # Create cell object
                    cell_obj = TableCell(
                        content=content,
                        row=current_row,
                        col=current_col,
                        colspan=colspan,
                        rowspan=rowspan,
                        is_header=is_header,
                        styles=styles
                    )
                    
                    cells.append(cell_obj)
                    current_col += colspan
                
                max_cols = max(max_cols, current_col)
                current_row += 1
            
            # Determine if this is a financial table
            is_financial = self._is_financial_table(cells)
            
            return TableStructure(
                cells=cells,
                rows=current_row,
                cols=max_cols,
                title=title,
                has_headers=any(cell.is_header for cell in cells),
                is_financial=is_financial
            )
            
        except Exception as e:
            logger.error(f"Failed to extract table structure: {e}")
            return None
    
    def _get_table_title_lxml(self, table_element) -> Optional[str]:
        """Get table title using lxml"""
        # Look for caption
        caption = table_element.xpath('.//caption')
        if caption:
            return caption[0].text_content().strip()
        
        # Look for preceding heading
        prev_sibling = table_element.getprevious()
        if prev_sibling is not None:
            if prev_sibling.tag in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                return prev_sibling.text_content().strip()
        
        return None
    
    def _get_table_title_soup(self, table_element) -> Optional[str]:
        """Get table title using BeautifulSoup"""
        # Look for caption
        caption = table_element.find('caption')
        if caption:
            return caption.get_text().strip()
        
        # Look for preceding heading
        prev_sibling = table_element.find_previous_sibling(['h1', 'h2', 'h3', 'h4', 'h5', 'h6'])
        if prev_sibling:
            return prev_sibling.get_text().strip()
        
        return None
    
    def _extract_cell_content_lxml(self, cell) -> str:
        """Extract cell content using lxml"""
        # Get text content, preserving structure
        content = cell.text_content()
        
        # Clean up whitespace
        content = re.sub(r'\s+', ' ', content).strip()
        
        # Handle special characters
        content = content.replace('&#160;', ' ').replace('&nbsp;', ' ')
        
        # Fix currency formatting - ensure dollar signs are properly attached to numbers
        content = self._fix_currency_formatting(content)
        
        return content
    
    def _extract_cell_content_soup(self, cell) -> str:
        """Extract cell content using BeautifulSoup"""
        # Get text content, preserving structure
        content = cell.get_text(separator=' ', strip=True)
        
        # Clean up whitespace
        content = re.sub(r'\s+', ' ', content).strip()
        
        # Handle special characters
        content = content.replace('&#160;', ' ').replace('&nbsp;', ' ')
        
        # Fix currency formatting - ensure dollar signs are properly attached to numbers
        content = self._fix_currency_formatting(content)
        
        return content
    
    def _fix_currency_formatting(self, content: str) -> str:
        """Fix currency formatting issues where dollar signs are separated from numbers"""
        if not content:
            return content
        
        # Pattern to match dollar signs followed by whitespace and then numbers
        # This handles cases like "$ 71,074" or "$71,074" 
        content = re.sub(r'\$\s+(\d)', r'$\1', content)
        
        # Pattern to match numbers followed by whitespace and then dollar signs
        # This handles cases like "71,074 $" - need to be more specific about what constitutes a number
        content = re.sub(r'([\d,]+\.?\d*)\s+\$', r'\1$', content)
        
        # Pattern to match standalone dollar signs that should be attached to the next number
        # This handles cases where $ is in a separate cell from the number
        content = re.sub(r'\$\s*$', '', content)  # Remove trailing dollar signs
        
        # Clean up any remaining extra whitespace
        content = re.sub(r'\s+', ' ', content).strip()
        
        return content
    
    def _extract_cell_styles_lxml(self, cell) -> Dict[str, Any]:
        """Extract cell styles using lxml"""
        styles = {}
        
        # Get style attribute
        style_attr = cell.get('style', '')
        if style_attr:
            styles.update(self._parse_style_attribute(style_attr))
        
        # Get class attribute
        class_attr = cell.get('class', '')
        if class_attr:
            styles['class'] = class_attr
        
        return styles
    
    def _extract_cell_styles_soup(self, cell) -> Dict[str, Any]:
        """Extract cell styles using BeautifulSoup"""
        styles = {}
        
        # Get style attribute
        style_attr = cell.get('style', '')
        if style_attr:
            styles.update(self._parse_style_attribute(style_attr))
        
        # Get class attribute
        class_attr = cell.get('class', [])
        if class_attr:
            styles['class'] = ' '.join(class_attr)
        
        return styles
    
    def _parse_style_attribute(self, style_str: str) -> Dict[str, str]:
        """Parse CSS style attribute into dictionary"""
        styles = {}
        if not style_str:
            return styles
        
        # Split by semicolon and parse each property
        for prop in style_str.split(';'):
            if ':' in prop:
                key, value = prop.split(':', 1)
                styles[key.strip()] = value.strip()
        
        return styles
    
    def _is_financial_table(self, cells: List[TableCell]) -> bool:
        """Determine if this is a financial table based on content"""
        financial_keywords = [
            'revenue', 'income', 'profit', 'loss', 'assets', 'liabilities',
            'equity', 'cash', 'sales', 'expenses', 'earnings', 'million',
            'billion', 'thousand', 'dollar', 'balance', 'statement'
        ]
        
        # Check cell contents for financial keywords
        content_text = ' '.join([cell.content.lower() for cell in cells])
        
        keyword_count = sum(1 for keyword in financial_keywords if keyword in content_text)
        
        # Also check for numerical patterns
        has_numbers = any(re.search(r'\d+[.,]\d+', cell.content) for cell in cells)
        
        return keyword_count >= 3 or (keyword_count >= 1 and has_numbers)

class ExcelTableReconstructor:
    """Reconstructs extracted tables perfectly in Excel with original formatting"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_excel_from_tables(self, tables: List[TableStructure], output_path: str, 
                                company_name: str = "Company") -> str:
        """
        Create Excel file from extracted table structures.
        
        Args:
            tables: List of extracted table structures
            output_path: Path for the output Excel file
            company_name: Name of the company
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Creating Excel file with {len(tables)} tables")
        
        # Create workbook
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        # Create sheets for each table
        for i, table in enumerate(tables):
            sheet_name = self._get_sheet_name(table, i)
            self.worksheet = self.workbook.create_sheet(title=sheet_name)
            
            # Reconstruct table in Excel
            self._reconstruct_table_in_excel(table)
        
        # Save workbook
        self.workbook.save(output_path)
        logger.info(f"Excel file saved to: {output_path}")
        
        return output_path
    
    def _get_sheet_name(self, table: TableStructure, index: int) -> str:
        """Generate appropriate sheet name for table"""
        if table.title:
            # Clean title for sheet name
            name = re.sub(r'[^\w\s-]', '', table.title)
            name = name[:31]  # Excel sheet name limit
            return name
        
        if table.is_financial:
            return f"Financial_Table_{index + 1}"
        
        return f"Table_{index + 1}"
    
    def _reconstruct_table_in_excel(self, table: TableStructure):
        """Reconstruct a single table in Excel with perfect formatting"""
        # Create a matrix to represent the table
        table_matrix = self._create_table_matrix(table)
        
        # Fill the matrix with cell data
        for cell in table.cells:
            self._place_cell_in_matrix(table_matrix, cell)
        
        # Write matrix to Excel
        self._write_matrix_to_excel(table_matrix, table)
        
        # Apply formatting
        self._apply_table_formatting(table)
    
    def _create_table_matrix(self, table: TableStructure) -> List[List[Optional[TableCell]]]:
        """Create a matrix representation of the table"""
        matrix = [[None for _ in range(table.cols)] for _ in range(table.rows)]
        return matrix
    
    def _place_cell_in_matrix(self, matrix: List[List[Optional[TableCell]]], cell: TableCell):
        """Place a cell in the matrix, handling colspan and rowspan"""
        for row_offset in range(cell.rowspan):
            for col_offset in range(cell.colspan):
                target_row = cell.row + row_offset
                target_col = cell.col + col_offset
                
                if (target_row < len(matrix) and 
                    target_col < len(matrix[target_row])):
                    matrix[target_row][target_col] = cell
    
    def _write_matrix_to_excel(self, matrix: List[List[Optional[TableCell]]], table: TableStructure):
        """Write the matrix to Excel worksheet"""
        for row_idx, row in enumerate(matrix, 1):
            for col_idx, cell in enumerate(row, 1):
                if cell:
                    excel_cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    excel_cell.value = cell.content
                    
                    # Apply cell formatting
                    self._apply_cell_formatting(excel_cell, cell)
    
    def _apply_cell_formatting(self, excel_cell, cell: TableCell):
        """Apply formatting to individual cell"""
        if not cell.styles:
            return
        
        # Font formatting
        font_kwargs = {}
        if 'font-weight' in cell.styles:
            font_kwargs['bold'] = cell.styles['font-weight'] in ['bold', '700', 'bolder']
        if 'font-size' in cell.styles:
            font_kwargs['size'] = self._parse_font_size(cell.styles['font-size'])
        if 'font-family' in cell.styles:
            font_kwargs['name'] = cell.styles['font-family'].split(',')[0].strip('"\'')
        
        if font_kwargs:
            excel_cell.font = Font(**font_kwargs)
        
        # Alignment
        alignment_kwargs = {}
        if 'text-align' in cell.styles:
            align_map = {
                'left': 'left',
                'center': 'center',
                'right': 'right',
                'justify': 'justify'
            }
            alignment_kwargs['horizontal'] = align_map.get(cell.styles['text-align'], 'left')
        
        if 'vertical-align' in cell.styles:
            valign_map = {
                'top': 'top',
                'middle': 'center',
                'bottom': 'bottom'
            }
            alignment_kwargs['vertical'] = valign_map.get(cell.styles['vertical-align'], 'center')
        
        if alignment_kwargs:
            excel_cell.alignment = Alignment(**alignment_kwargs)
        
        # Apply currency formatting for financial numbers
        self._apply_currency_formatting(excel_cell, cell)
        
        # Background color
        if 'background-color' in cell.styles:
            color = self._parse_color(cell.styles['background-color'])
            if color:
                excel_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        
        # Border
        if 'border' in cell.styles or 'border-style' in cell.styles:
            border_side = Side(style='thin')
            excel_cell.border = Border(
                left=border_side,
                right=border_side,
                top=border_side,
                bottom=border_side
            )
    
    def _apply_currency_formatting(self, excel_cell, cell: TableCell):
        """Apply currency formatting to Excel cells containing financial data"""
        if not cell.content:
            return
        
        content = str(cell.content).strip()
        
        # Skip year headers and other non-currency numbers
        if self._is_year_header(content) or self._is_non_currency_number(content):
            return
        
        # Only apply currency formatting to specific financial statement line items
        # Check if this is a financial number that should have currency formatting
        is_currency = self._should_format_as_currency(content, excel_cell)
        
        if is_currency:
            # Try to extract the numeric value
            numeric_value = self._extract_numeric_value(content)
            if numeric_value is not None:
                # Set the numeric value and apply currency formatting
                excel_cell.value = numeric_value
                excel_cell.number_format = '$#,##0'
                
                # Right-align currency values
                if not excel_cell.alignment:
                    excel_cell.alignment = Alignment(horizontal='right')
                else:
                    excel_cell.alignment.horizontal = 'right'
    
    def _extract_numeric_value(self, content: str) -> Optional[float]:
        """Extract numeric value from a string that may contain currency symbols"""
        if not content:
            return None
        
        # Remove currency symbols and clean up
        cleaned = re.sub(r'[$,]', '', content).strip()
        
        # Handle negative numbers in parentheses
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = '-' + cleaned[1:-1]
        
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    def _is_year_header(self, content: str) -> bool:
        """Check if content is a year header (e.g., 2021, 2022, 2023)"""
        if not content:
            return False
        
        # Check if it's a 4-digit year
        if content.isdigit() and len(content) == 4:
            year = int(content)
            # Check if it's a reasonable year (1900-2100)
            if 1900 <= year <= 2100:
                return True
        
        return False
    
    def _is_non_currency_number(self, content: str) -> bool:
        """Check if content is a number that shouldn't be formatted as currency"""
        if not content:
            return False
        
        # Check for common non-currency numbers
        non_currency_patterns = [
            r'^\d{4}$',  # 4-digit years
            r'^\d{1,2}$',  # Single or double digit numbers (likely row numbers, percentages, etc.)
            r'^\d+\.\d{1,2}$',  # Decimal numbers with 1-2 decimal places (likely percentages, ratios)
            r'^\d+%$',  # Numbers with percentage sign
            r'^\d+\.\d+$',  # Decimal numbers (could be ratios, not currency)
        ]
        
        for pattern in non_currency_patterns:
            if re.match(pattern, content.strip()):
                return True
        
        return False
    
    def _should_format_as_currency(self, content: str, excel_cell) -> bool:
        """Determine if a cell should be formatted as currency based on context"""
        if not content:
            return False
        
        # Check if this looks like a financial number
        is_financial_number = (
            '$' in content or 
            (content.replace(',', '').replace('.', '').isdigit() and len(content.replace(',', '').replace('.', '')) >= 3)
        )
        
        if not is_financial_number:
            return False
        
        # Get the row and column to check context
        row = excel_cell.row
        col = excel_cell.column
        
        # Try to get the row label (first column of the same row)
        try:
            row_label_cell = excel_cell.parent.cell(row=row, column=1)
            row_label = str(row_label_cell.value or "").strip().lower()
        except:
            row_label = ""
        
        # Only apply currency formatting to specific line items
        currency_line_items = [
            "total revenue",
            "revenue",
            "total cost of revenue", 
            "cost of revenue",
            "gross margin",
            "operating income",
            "income before income taxes",
            "net income",
            "total assets",
            "total liabilities",
            "stockholders' equity",
            "total stockholders' equity"
        ]
        
        # Check if the row label matches any of the currency line items
        for item in currency_line_items:
            if item in row_label:
                return True
        
        # For now, don't apply currency formatting to other rows
        return False
    
    def _apply_table_formatting(self, table: TableStructure):
        """Apply overall table formatting"""
        # Auto-adjust column widths
        for col in range(1, table.cols + 1):
            max_length = 0
            for row in range(1, table.rows + 1):
                cell = self.worksheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width with some padding
            self.worksheet.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    def _parse_font_size(self, font_size: str) -> int:
        """Parse font size from CSS"""
        try:
            # Remove units and convert to int
            size = re.sub(r'[^\d.]', '', font_size)
            return int(float(size))
        except:
            return 11  # Default font size
    
    def _parse_color(self, color: str) -> Optional[str]:
        """Parse color from CSS"""
        # Remove whitespace
        color = color.strip()
        
        # Handle hex colors
        if color.startswith('#'):
            return color[1:]  # Remove # for openpyxl
        
        # Handle rgb() colors
        if color.startswith('rgb('):
            rgb_match = re.search(r'rgb\((\d+),\s*(\d+),\s*(\d+)\)', color)
            if rgb_match:
                r, g, b = rgb_match.groups()
                return f"{int(r):02x}{int(g):02x}{int(b):02x}"
        
        # Handle named colors
        color_map = {
            'red': 'FF0000',
            'green': '00FF00',
            'blue': '0000FF',
            'yellow': 'FFFF00',
            'white': 'FFFFFF',
            'black': '000000',
            'gray': '808080',
            'grey': '808080'
        }
        
        return color_map.get(color.lower())

class SECAPIClient:
    """SEC API client for querying EDGAR filings and extracting sections"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.sec-api.io"
        self.extractor_url = "https://api.sec-api.io/extractor"
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': api_key,
            'Content-Type': 'application/json'
        })
    
    def search_filings(self, query: str, from_pos: int = 0, size: int = 50, 
                      sort_field: str = "filedAt", sort_order: str = "desc") -> Dict[str, Any]:
        """
        Search EDGAR filings using the Query API
        
        Args:
            query: Search query in Lucene syntax
            from_pos: Starting position for pagination
            size: Number of filings to return (max 50)
            sort_field: Field to sort by
            sort_order: Sort order ('asc' or 'desc')
            
        Returns:
            Dictionary containing search results
        """
        payload = {
            "query": query,
            "from": str(from_pos),
            "size": str(size),
            "sort": [{sort_field: {"order": sort_order}}]
        }
        
        try:
            response = self.session.post(self.base_url, json=payload)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"SEC API search failed: {e}")
            return {"total": {"value": 0}, "filings": []}
    
    def get_recent_10k_filings(self, ticker: str = None, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent 10-K filings for a specific ticker or all companies"""
        if ticker:
            query = f'formType:"10-K" AND ticker:{ticker.upper()}'
        else:
            query = 'formType:"10-K"'
        
        result = self.search_filings(query, size=limit)
        return result.get('filings', [])
    
    def get_recent_10q_filings(self, ticker: str = None, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent 10-Q filings for a specific ticker or all companies"""
        if ticker:
            query = f'formType:"10-Q" AND ticker:{ticker.upper()}'
        else:
            query = 'formType:"10-Q"'
        
        result = self.search_filings(query, size=limit)
        return result.get('filings', [])
    
    def get_8k_filings(self, ticker: str = None, item_code: str = None, limit: int = 10) -> List[Dict[str, Any]]:
        """Get 8-K filings with optional item code filter"""
        if ticker and item_code:
            query = f'formType:"8-K" AND ticker:{ticker.upper()} AND items:"{item_code}"'
        elif ticker:
            query = f'formType:"8-K" AND ticker:{ticker.upper()}'
        elif item_code:
            query = f'formType:"8-K" AND items:"{item_code}"'
        else:
            query = 'formType:"8-K"'
        
        result = self.search_filings(query, size=limit)
        return result.get('filings', [])
    
    def extract_section(self, filing_url: str, item: str, return_type: str = "text") -> str:
        """
        Extract a specific section from a 10-K, 10-Q, or 8-K filing
        
        Args:
            filing_url: URL of the filing
            item: Item code to extract (e.g., "1A", "8", "part2item1a")
            return_type: "text" or "html"
            
        Returns:
            Extracted section content
        """
        params = {
            'url': filing_url,
            'item': item,
            'type': return_type,
            'token': self.api_key
        }
        
        try:
            response = requests.get(self.extractor_url, params=params)
            response.raise_for_status()
            return response.text
        except requests.exceptions.RequestException as e:
            logger.error(f"Section extraction failed: {e}")
            return ""
    
    def extract_financial_statements(self, filing_url: str) -> str:
        """Extract Item 8 (Financial Statements) from 10-K or Part 1 Item 1 from 10-Q"""
        return self.extract_section(filing_url, "8", "html")
    
    def extract_risk_factors(self, filing_url: str, form_type: str = "10-K") -> str:
        """Extract Risk Factors section based on form type"""
        if form_type.upper() == "10-Q":
            return self.extract_section(filing_url, "part2item1a", "text")
        else:
            return self.extract_section(filing_url, "1A", "text")
    
    def extract_mda(self, filing_url: str, form_type: str = "10-K") -> str:
        """Extract Management's Discussion and Analysis section"""
        if form_type.upper() == "10-Q":
            return self.extract_section(filing_url, "part1item2", "text")
        else:
            return self.extract_section(filing_url, "7", "text")

class SECFilingAnalyzer:
    """Analyzes SEC filings using both Query and Extractor APIs"""
    
    def __init__(self, api_key: str):
        self.api_client = SECAPIClient(api_key)
        self.table_extractor = UniversalHTMLTableExtractor()
        self.excel_reconstructor = ExcelTableReconstructor()
    
    def analyze_company_filings(self, ticker: str, form_types: List[str] = None, 
                               years: List[int] = None) -> Dict[str, Any]:
        """
        Comprehensive analysis of a company's SEC filings
        
        Args:
            ticker: Company ticker symbol
            form_types: List of form types to analyze (e.g., ["10-K", "10-Q"])
            years: List of years to analyze
            
        Returns:
            Dictionary containing analysis results
        """
        if form_types is None:
            form_types = ["10-K", "10-Q"]
        
        results = {
            "ticker": ticker,
            "filings": {},
            "analysis": {}
        }
        
        for form_type in form_types:
            logger.info(f"Analyzing {form_type} filings for {ticker}")
            
            # Get recent filings
            if form_type == "10-K":
                filings = self.api_client.get_recent_10k_filings(ticker, limit=20)
            elif form_type == "10-Q":
                filings = self.api_client.get_recent_10q_filings(ticker, limit=20)
            else:
                filings = self.api_client.get_8k_filings(ticker, limit=20)
            
            results["filings"][form_type] = filings
            
            # Extract key sections from recent filings
            for filing in filings[:5]:  # Analyze top 5 most recent
                filing_url = filing.get('linkToFilingDetails', '')
                if not filing_url:
                    continue
                
                # Extract financial statements (HTML for table extraction)
                financial_html = self.api_client.extract_financial_statements(filing_url)
                if financial_html:
                    # Extract tables from HTML
                    tables = self.table_extractor.extract_tables_from_html_content(financial_html)
                    if tables:
                        results["analysis"][f"{form_type}_{filing['accessionNo']}"] = {
                            "financial_tables": len(tables),
                            "filing_date": filing.get('filedAt', ''),
                            "period": filing.get('periodOfReport', '')
                        }
        
        return results
    
    def create_financial_analysis_excel(self, ticker: str, output_path: str = None) -> str:
        """
        Create comprehensive Excel analysis of company's financial filings
        
        Args:
            ticker: Company ticker symbol
            output_path: Output Excel file path
            
        Returns:
            Path to created Excel file
        """
        if output_path is None:
            output_path = f"{ticker}_financial_analysis.xlsx"
        
        logger.info(f"Creating financial analysis Excel for {ticker}")
        
        # Get recent 10-K and 10-Q filings
        k10_filings = self.api_client.get_recent_10k_filings(ticker, limit=5)
        q10_filings = self.api_client.get_recent_10q_filings(ticker, limit=5)
        
        all_tables = []
        
        # Process 10-K filings
        for filing in k10_filings:
            filing_url = filing.get('linkToFilingDetails', '')
            if filing_url:
                financial_html = self.api_client.extract_financial_statements(filing_url)
                if financial_html:
                    tables = self.table_extractor.extract_tables_from_html_content(financial_html)
                    for table in tables:
                        table.title = f"10-K {filing.get('periodOfReport', '')} - {table.title or 'Financial Statement'}"
                    all_tables.extend(tables)
        
        # Process 10-Q filings
        for filing in q10_filings:
            filing_url = filing.get('linkToFilingDetails', '')
            if filing_url:
                financial_html = self.api_client.extract_financial_statements(filing_url)
                if financial_html:
                    tables = self.table_extractor.extract_tables_from_html_content(financial_html)
                    for table in tables:
                        table.title = f"10-Q {filing.get('periodOfReport', '')} - {table.title or 'Financial Statement'}"
                    all_tables.extend(tables)
        
        if all_tables:
            # Create Excel file
            self.excel_reconstructor.create_excel_from_tables(all_tables, output_path, ticker)
            logger.info(f"Financial analysis Excel created: {output_path}")
            return output_path
        else:
            logger.warning(f"No financial tables found for {ticker}")
            return ""

    def extract_tables_from_html_content(self, html_content: str) -> List[TableStructure]:
        """Extract tables from HTML content string"""
        # Create temporary file for processing
        import tempfile
        with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False) as f:
            f.write(html_content)
            temp_file = f.name
        
        try:
            tables = self.table_extractor.extract_tables_from_html(temp_file)
            return tables
        finally:
            # Clean up temporary file
            import os
            os.unlink(temp_file)

def _create_multi_sheet_excel(tables_by_period, output_file, ticker, form_type):
    """Create Excel file with separate sheets for each filing period"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # Create workbook
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Create sheets for each period (sorted by date, newest first)
    for period in sorted(tables_by_period.keys(), reverse=True):
        tables = tables_by_period[period]
        
        # Clean period for sheet name (Excel sheet names have restrictions)
        sheet_name = period.replace('-', '_')[:31]  # Excel sheet name limit
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Add period as header
        worksheet.cell(row=1, column=1, value=f"{ticker} {form_type} - {period}")
        worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        
        # Process tables for this period
        current_row = 3  # Start after header
        
        for table in tables:
            # Add table title
            table_title = table.title or "Financial Table"
            worksheet.cell(row=current_row, column=1, value=table_title)
            worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
            current_row += 2
            
            # Create table matrix
            table_matrix = [[None for _ in range(table.cols)] for _ in range(table.rows)]
            
            # Place cells in matrix
            for cell in table.cells:
                for row_offset in range(cell.rowspan):
                    for col_offset in range(cell.colspan):
                        target_row = cell.row + row_offset
                        target_col = cell.col + col_offset
                        
                        if (target_row < len(table_matrix) and 
                            target_col < len(table_matrix[target_row])):
                            table_matrix[target_row][target_col] = cell
            
            # Write matrix to Excel
            for row_idx, row in enumerate(table_matrix):
                for col_idx, cell in enumerate(row):
                    if cell:
                        excel_cell = worksheet.cell(row=current_row + row_idx, column=col_idx + 1)
                        excel_cell.value = cell.content
                        
                        # Apply basic formatting
                        if cell.is_header:
                            excel_cell.font = openpyxl.styles.Font(bold=True)
                        if cell.styles and 'text-align' in cell.styles:
                            if cell.styles['text-align'] == 'right':
                                excel_cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                            elif cell.styles['text-align'] == 'center':
                                excel_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            current_row += table.rows + 2  # Add space between tables
        
        # Auto-adjust column widths
        for col in range(1, 20):  # Adjust first 20 columns
            max_length = 0
            for row in range(1, current_row):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            if max_length > 0:
                worksheet.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    # Save workbook
    workbook.save(output_file)

def main():
    """Interactive SEC filing search and table extraction"""
    # SEC API Key
    API_KEY = "62ff63ea351833fb6ad40b2f4becbf5539a91740ce09544e96b42600de5853c5"
    
    print("=== SEC Filing Search and Table Extraction ===\n")
    
    # Initialize components
    extractor = UniversalHTMLTableExtractor()
    reconstructor = ExcelTableReconstructor()
    sec_client = SECAPIClient(API_KEY)
    analyzer = SECFilingAnalyzer(API_KEY)
    
    while True:
        print("\n" + "="*60)
        print("SEC Filing Search and Extraction")
        print("="*60)
        
        try:
            # Get search parameters from user
            ticker = input("Enter company ticker symbol (e.g., MSFT, AAPL): ").strip().upper()
            if not ticker:
                print("Ticker symbol is required.")
                continue
            
            form_type = input("Enter form type (10-K, 10-Q, 8-K) or press Enter for 10-K: ").strip()
            if not form_type:
                form_type = "10-K"
            
            limit = input("Enter number of filings to retrieve (1-50, default 5): ").strip()
            try:
                limit = int(limit) if limit else 5
                limit = max(1, min(50, limit))
            except ValueError:
                limit = 5
            
            # Search for filings
            print(f"\nSearching for {form_type} filings for {ticker}...")
            
            if form_type == "10-K":
                filings = sec_client.get_recent_10k_filings(ticker, limit=limit)
            elif form_type == "10-Q":
                filings = sec_client.get_recent_10q_filings(ticker, limit=limit)
            elif form_type == "8-K":
                filings = sec_client.get_8k_filings(ticker, limit=limit)
            else:
                print(f"Unsupported form type: {form_type}")
                continue
            
            if not filings:
                print(f"No {form_type} filings found for {ticker}")
                continue
            
            print(f"Found {len(filings)} {form_type} filings:")
            for i, filing in enumerate(filings, 1):
                print(f"  {i}. {filing.get('periodOfReport', 'N/A')} - Filed: {filing.get('filedAt', 'N/A')}")
                print(f"     Accession: {filing.get('accessionNo', 'N/A')}")
            
            # Ask user if they want to extract from all filings or just one
            print(f"\nExtraction options:")
            print("1. Extract from ALL filings (recommended for multi-year analysis)")
            print("2. Extract from a single filing")
            
            extraction_choice = input("Choose extraction method (1 or 2, default 1): ").strip()
            if not extraction_choice:
                extraction_choice = "1"
            
            if extraction_choice == "1":
                # Extract from all filings
                print(f"\nExtracting from ALL {len(filings)} {form_type} filings...")
                
                # Determine section to extract
                print("\nAvailable sections to extract:")
                if form_type in ["10-K", "10-Q"]:
                    print("  - Financial Statements (Item 8 for 10-K, Part 1 Item 1 for 10-Q)")
                    print("  - Risk Factors (Item 1A for 10-K, Part 2 Item 1A for 10-Q)")
                    print("  - Management Discussion & Analysis (Item 7 for 10-K, Part 1 Item 2 for 10-Q)")
                    print("  - Custom section (enter item code)")
                
                section_choice = input("\nEnter section to extract (or press Enter for Financial Statements): ").strip().lower()
                
                if not section_choice or section_choice in ["financial", "financial statements", "8", "1"]:
                    if form_type == "10-Q":
                        item_code = "part1item1"
                    else:
                        item_code = "8"
                    section_name = "Financial Statements"
                elif section_choice in ["risk", "risk factors", "1a"]:
                    if form_type == "10-Q":
                        item_code = "part2item1a"
                    else:
                        item_code = "1A"
                    section_name = "Risk Factors"
                elif section_choice in ["mda", "management", "7"]:
                    if form_type == "10-Q":
                        item_code = "part1item2"
                    else:
                        item_code = "7"
                    section_name = "Management Discussion & Analysis"
                else:
                    item_code = section_choice
                    section_name = f"Item {section_choice}"
                
                # Process all filings - create separate sheets for each filing
                tables_by_period = {}
                successful_extractions = 0
                
                for i, filing in enumerate(filings, 1):
                    filing_url = filing.get('linkToFilingDetails', '')
                    if not filing_url:
                        print(f"  {i}. Skipping - No URL available")
                        continue
                    
                    period = filing.get('periodOfReport', 'Unknown')
                    print(f"  {i}. Processing {period} filing...")
                    
                    if section_choice in ["financial", "financial statements", "8"] or not section_choice:
                        # Extract as HTML for table extraction
                        content = sec_client.extract_section(filing_url, item_code, "html")
                    else:
                        # Extract as text for other sections
                        content = sec_client.extract_section(filing_url, item_code, "text")
                    
                    if content:
                        print(f"     ✓ Extracted {len(content)} characters")
                        
                        if section_choice in ["financial", "financial statements", "8"] or not section_choice:
                            # Extract tables from this filing
                            tables = extractor.extract_tables_from_html_content(content)
                            if tables:
                                print(f"     ✓ Found {len(tables)} tables")
                                # Store tables by period
                                tables_by_period[period] = tables
                                successful_extractions += 1
                            else:
                                print(f"     ⚠ No tables found")
                        else:
                            # Save individual text files for each filing
                            output_file = f"{ticker}_{form_type}_{period}_{section_name.replace(' ', '_')}.txt"
                            with open(output_file, 'w', encoding='utf-8') as f:
                                f.write(content)
                            print(f"     ✓ Saved to {output_file}")
                            successful_extractions += 1
                    else:
                        print(f"     ✗ Failed to extract section")
                
                # Create Excel file with separate sheets for each period
                if section_choice in ["financial", "financial statements", "8"] or not section_choice:
                    if tables_by_period:
                        print(f"\n✓ Successfully extracted tables from {successful_extractions} filings")
                        
                        # Ask for output file name
                        output_file = input("Enter output Excel file name (or press Enter for default): ").strip()
                        if not output_file:
                            output_file = f"{ticker}_{form_type}_Multi_Year_Sheets.xlsx"
                        elif not output_file.endswith('.xlsx'):
                            output_file += '.xlsx'
                        
                        # Create multi-sheet Excel file
                        _create_multi_sheet_excel(tables_by_period, output_file, ticker, form_type)
                        print(f"✓ Excel file created: {output_file}")
                        print(f"  Contains separate sheets for {len(tables_by_period)} filing periods")
                        for period in sorted(tables_by_period.keys(), reverse=True):
                            print(f"    - {period}: {len(tables_by_period[period])} tables")
                    else:
                        print("✗ No tables found in any of the filings")
                else:
                    print(f"✓ Successfully processed {successful_extractions} filings")
            
            else:
                # Extract from single filing (original behavior)
                if len(filings) == 1:
                    selected_filing = filings[0]
                    print(f"\nExtracting from the only available filing...")
                else:
                    while True:
                        try:
                            choice_num = int(input(f"\nSelect filing number (1-{len(filings)}): "))
                            if 1 <= choice_num <= len(filings):
                                selected_filing = filings[choice_num - 1]
                                break
                            else:
                                print(f"Please enter a number between 1 and {len(filings)}")
                        except ValueError:
                            print("Please enter a valid number")
                
                # Extract section
                print("\nAvailable sections to extract:")
                if form_type in ["10-K", "10-Q"]:
                    print("  - Financial Statements (Item 8 for 10-K, Part 1 Item 1 for 10-Q)")
                    print("  - Risk Factors (Item 1A for 10-K, Part 2 Item 1A for 10-Q)")
                    print("  - Management Discussion & Analysis (Item 7 for 10-K, Part 1 Item 2 for 10-Q)")
                    print("  - Custom section (enter item code)")
                
                section_choice = input("\nEnter section to extract (or press Enter for Financial Statements): ").strip().lower()
                
                if not section_choice or section_choice in ["financial", "financial statements", "8", "1"]:
                    if form_type == "10-Q":
                        item_code = "part1item1"
                    else:
                        item_code = "8"
                    section_name = "Financial Statements"
                elif section_choice in ["risk", "risk factors", "1a"]:
                    if form_type == "10-Q":
                        item_code = "part2item1a"
                    else:
                        item_code = "1A"
                    section_name = "Risk Factors"
                elif section_choice in ["mda", "management", "7"]:
                    if form_type == "10-Q":
                        item_code = "part1item2"
                    else:
                        item_code = "7"
                    section_name = "Management Discussion & Analysis"
                else:
                    item_code = section_choice
                    section_name = f"Item {section_choice}"
                
                # Extract the section
                filing_url = selected_filing.get('linkToFilingDetails', '')
                if not filing_url:
                    print("No filing URL available")
                    continue
                
                print(f"\nExtracting {section_name} from {ticker} {form_type}...")
                print(f"URL: {filing_url}")
                
                if section_choice in ["financial", "financial statements", "8"] or not section_choice:
                    # Extract as HTML for table extraction
                    content = sec_client.extract_section(filing_url, item_code, "html")
                else:
                    # Extract as text for other sections
                    content = sec_client.extract_section(filing_url, item_code, "text")
                
                if content:
                    print(f"Section extracted successfully! Content length: {len(content)} characters")
                    
                    # Extract tables if it's financial statements
                    if section_choice in ["financial", "financial statements", "8"] or not section_choice:
                        tables = extractor.extract_tables_from_html_content(content)
                        if tables:
                            print(f"Found {len(tables)} tables in the financial statements")
                            
                            # Ask for output file name
                            output_file = input("Enter output Excel file name (or press Enter for default): ").strip()
                            if not output_file:
                                output_file = f"{ticker}_{form_type}_{selected_filing.get('periodOfReport', 'latest')}_tables.xlsx"
                            elif not output_file.endswith('.xlsx'):
                                output_file += '.xlsx'
                            
                            reconstructor.create_excel_from_tables(tables, output_file, ticker)
                            print(f"Excel file created: {output_file}")
                        else:
                            print("No tables found in the financial statements")
                    else:
                        # Save text content to file
                        output_file = input("Enter output text file name (or press Enter for default): ").strip()
                        if not output_file:
                            output_file = f"{ticker}_{form_type}_{section_name.replace(' ', '_')}.txt"
                        elif not output_file.endswith('.txt'):
                            output_file += '.txt'
                        
                        with open(output_file, 'w', encoding='utf-8') as f:
                            f.write(content)
                        print(f"Text file created: {output_file}")
                else:
                    print("Failed to extract section. The section might not exist in this filing.")
        
        except Exception as e:
            print(f"Error: {e}")
            print("Make sure you have a valid API key and internet connection")
        
        # Ask if user wants to continue
        continue_choice = input("\nDo you want to search for another filing? (y/n): ").strip().lower()
        if continue_choice not in ['y', 'yes']:
            break
    
    print("\nThank you for using the SEC Filing Search and Table Extraction tool!")

if __name__ == "__main__":
    main()

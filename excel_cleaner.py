#!/usr/bin/env python3
"""
Excel Cleaning and Structure Enhancement Library
Combines pandas and openpyxl for professional Excel output
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import re

class ExcelCleaner:
    """Professional Excel cleaning and structuring"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
    
    def clean_duplicates(self, sheet_name=None):
        """Remove duplicate values in adjacent cells"""
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        for row in sheet.iter_rows():
            for i in range(len(row) - 1):
                if (row[i].value and row[i+1].value and 
                    row[i].value == row[i+1].value):
                    row[i].value = None
    
    def apply_financial_formatting(self, sheet_name=None):
        """Apply professional financial statement formatting"""
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Currency formatting
        currency_font = Font(name="Arial", size=10)
        currency_alignment = Alignment(horizontal="right")
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Format headers
                    if self._is_header(cell.value):
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Format currency values
                    elif self._is_currency(cell.value):
                        cell.font = currency_font
                        cell.alignment = currency_alignment
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0'
    
    def _is_header(self, value):
        """Check if value is a header"""
        if not value:
            return False
        value_str = str(value).strip()
        
        # Year headers
        if value_str.isdigit() and len(value_str) == 4:
            year = int(value_str)
            if 1900 <= year <= 2100:
                return True
        
        # Text headers
        header_keywords = [
            "revenue", "income", "expense", "assets", "liabilities", 
            "equity", "cash", "debt", "year ended", "quarter ended"
        ]
        return any(keyword in value_str.lower() for keyword in header_keywords)
    
    def _is_currency(self, value):
        """Check if value is a currency amount"""
        if not value:
            return False
        
        if isinstance(value, (int, float)):
            return abs(value) > 100  # Assume amounts > 100 are currency
        
        value_str = str(value).strip()
        # Check for currency patterns
        currency_patterns = [
            r'^\$[\d,]+\.?\d*$',  # $1,234.56
            r'^[\d,]+\.?\d*$',    # 1,234.56
            r'^\([\d,]+\.?\d*\)$' # (1,234.56)
        ]
        return any(re.match(pattern, value_str) for pattern in currency_patterns)
    
    def add_borders(self, sheet_name=None):
        """Add professional borders to tables"""
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    def auto_adjust_columns(self, sheet_name=None):
        """Auto-adjust column widths"""
        if sheet_name:
            sheet = self.workbook[sheet_name]
        else:
            sheet = self.workbook.active
        
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, summary_data):
        """Create a summary sheet with key metrics"""
        summary_sheet = self.workbook.create_sheet("Summary", 0)
        
        # Add summary data
        for i, (metric, value) in enumerate(summary_data.items(), 1):
            summary_sheet.cell(row=i, column=1, value=metric)
            summary_sheet.cell(row=i, column=2, value=value)
        
        # Format summary sheet
        self.apply_financial_formatting("Summary")
        self.auto_adjust_columns("Summary")
    
    def clean_all_sheets(self):
        """Apply cleaning to all sheets"""
        for sheet_name in self.workbook.sheetnames:
            print(f"Cleaning sheet: {sheet_name}")
            self.clean_duplicates(sheet_name)
            self.apply_financial_formatting(sheet_name)
            self.add_borders(sheet_name)
            self.auto_adjust_columns(sheet_name)
    
    def save(self, output_path=None):
        """Save the cleaned workbook"""
        if output_path:
            self.workbook.save(output_path)
        else:
            self.workbook.save(self.file_path)

def clean_excel_file(input_path, output_path=None):
    """One-stop function to clean an Excel file"""
    if not output_path:
        output_path = input_path.replace('.xlsx', '_cleaned.xlsx')
    
    cleaner = ExcelCleaner(input_path)
    cleaner.clean_all_sheets()
    
    # Create summary
    summary_data = {
        "Total Sheets": len(cleaner.workbook.sheetnames),
        "File Created": "Financial Statement Extractor",
        "Status": "Cleaned and Formatted"
    }
    cleaner.create_summary_sheet(summary_data)
    
    cleaner.save(output_path)
    print(f"✅ Cleaned Excel file saved: {output_path}")
    return output_path

if __name__ == "__main__":
    # Example usage
    input_file = "MSFT_10-K_Multi_Year_Cleaned.xlsx"
    output_file = "MSFT_10-K_Multi_Year_Professional.xlsx"
    
    clean_excel_file(input_file, output_file)

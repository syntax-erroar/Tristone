#!/usr/bin/env python3
"""
Test script to verify that duplicate year headers have been cleaned up
"""

import openpyxl
import os

def test_duplicate_cleanup():
    """Test the Excel file to check if duplicate year headers have been removed"""
    
    excel_file = "MSFT_10-K_Multi_Year_Cleaned.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file {excel_file} not found")
        return False
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        print(f"✅ Successfully loaded {excel_file}")
        print(f"📊 Available sheets: {workbook.sheetnames}")
        
        # Test the 2025 sheet (newest data)
        if "2025_06_30" in workbook.sheetnames:
            sheet = workbook["2025_06_30"]
            print(f"\n🔍 Examining 2025_06_30 sheet for duplicate year headers...")
            
            # Look for year headers in the first few rows
            year_headers_found = []
            duplicate_years = []
            
            for row in range(1, 20):  # Check first 20 rows
                row_years = []
                for col in range(1, 20):  # Check first 20 columns
                    cell = sheet.cell(row=row, column=col)
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        
                        # Check if it's a year header
                        if cell_value.isdigit() and len(cell_value) == 4 and 1900 <= int(cell_value) <= 2100:
                            row_years.append((cell_value, col))
                            year_headers_found.append(f"Row {row}, Col {col}: {cell_value}")
                
                # Check for duplicates in this row
                if len(row_years) > 1:
                    years_in_row = [year for year, col in row_years]
                    unique_years = set(years_in_row)
                    if len(unique_years) < len(years_in_row):
                        duplicate_years.append(f"Row {row}: {years_in_row} (duplicates found)")
            
            print(f"\n📈 Results:")
            print(f"✅ Year headers found: {len(year_headers_found)}")
            print(f"❌ Rows with duplicate years: {len(duplicate_years)}")
            
            if year_headers_found:
                print(f"\n📅 Year headers found:")
                for header in year_headers_found[:10]:
                    print(f"   - {header}")
                if len(year_headers_found) > 10:
                    print(f"   ... and {len(year_headers_found) - 10} more")
            
            if duplicate_years:
                print(f"\n🚨 Duplicate years found:")
                for duplicate in duplicate_years:
                    print(f"   - {duplicate}")
            else:
                print(f"\n🎉 No duplicate year headers found!")
            
            # Success criteria: No duplicate years in any row
            success = len(duplicate_years) == 0
            
            if success:
                print(f"\n✅ SUCCESS: Duplicate year headers have been cleaned up!")
            else:
                print(f"\n❌ FAILED: Duplicate year headers still exist")
            
            return success
            
        else:
            print(f"❌ 2025_06_30 sheet not found in workbook")
            return False
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

if __name__ == "__main__":
    print("🧪 Testing Duplicate Year Header Cleanup")
    print("=" * 60)
    
    success = test_duplicate_cleanup()
    
    if success:
        print(f"\n🎉 SUCCESS: Duplicate year headers have been successfully removed!")
        print(f"   - Each year (2025, 2024, 2023) now appears only once per row")
        print(f"   - Excel output is clean and properly formatted")
    else:
        print(f"\n❌ FAILED: Duplicate year headers still need to be addressed")

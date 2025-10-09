#!/usr/bin/env python3
"""
Advanced Automated SEC Excel Downloader
Uses multiple techniques to bypass SEC's anti-bot protection
"""

import os
import requests
import json
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
import time
import random
import subprocess
import webbrowser
from urllib.parse import urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import sys
import re
import numpy as np
try:
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:
    cosine_similarity = None

class AdvancedSECDownloader:
    """Advanced downloader with multiple bypass techniques"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.sec-api.io"
        self.sec_base_url = "https://www.sec.gov"
        self.downloaded_files = []
        self.last_used_excel_url: Optional[str] = None
        self._finlang_model = None
        
        # Multiple user agents to rotate
        self.user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15'
        ]
        
        # Lazy-loaded semantic model for metric similarity
        self._finlang_model = None

    # ===================== Deduplication Helpers =====================
    def _load_finlang_model(self):
        """Lazy load FinLang embeddings model; fallback to None on failure."""
        if self._finlang_model is not None:
            return self._finlang_model
        try:
            from sentence_transformers import SentenceTransformer
            self._finlang_model = SentenceTransformer("FinLang/finance-embeddings-investopedia")
        except Exception:
            self._finlang_model = None
        return self._finlang_model

    def _normalize_metric_name(self, text):
        if pd.isna(text) or text == '':
            return ''
        s = str(text).strip().lower()
        s = re.sub(r'\s+', ' ', s)
        s = re.sub(r'[^\w\s]', '', s)
        return s

    def _normalize_value(self, val):
        """Normalize for comparison and alignment.
        - Strip formatting
        - If value has restatement syntax like "100{105}" or "100{105,110}",
          compare using the LAST restated value (e.g., 105 or 110).
        """
        if pd.isna(val):
            return ''
        val_str = str(val).strip()
        # Extract last restated value if present
        if '{' in val_str and '}' in val_str:
            try:
                matches = re.findall(r'\{([^}]+)\}', val_str)
                if matches:
                    last = matches[-1]
                    last_split = [p.strip() for p in last.split(',') if p.strip() != '']
                    if last_split:
                        val_str = last_split[-1]
            except Exception:
                pass
        # Remove formatting chars
        val_str = val_str.replace(',', '').replace('$', '').replace('%', '')
        return val_str.lower()

    def _are_metrics_similar(self, metric1, metric2, threshold=0.78):
        if not metric1 or not metric2:
            return False
        norm1 = self._normalize_metric_name(metric1)
        norm2 = self._normalize_metric_name(metric2)
        if norm1 == norm2 and norm1 != '':
            return True
        # Quick alias normalization for common revenue/sales synonyms
        alias_map = {
            'revenues': 'revenue', 'total revenues': 'revenue', 'total revenue': 'revenue',
            'net sales': 'revenue', 'sales': 'revenue', 'sales revenue': 'revenue',
            'operating revenues': 'revenue',
        }
        alias1 = alias_map.get(norm1, norm1)
        alias2 = alias_map.get(norm2, norm2)
        if alias1 == alias2 and alias1 != '':
            return True
        # Token overlap fallback (Jaccard-like)
        t1 = set([t for t in re.split(r'\s+', alias1) if t])
        t2 = set([t for t in re.split(r'\s+', alias2) if t])
        if t1 and t2:
            inter = len(t1 & t2); uni = len(t1 | t2)
            if uni > 0 and (inter / uni) >= 0.6:
                return True
        model = self._load_finlang_model()
        if model is None or cosine_similarity is None:
            return False
        try:
            emb1 = model.encode([str(metric1)])
            emb2 = model.encode([str(metric2)])
            sim = float(cosine_similarity(emb1, emb2)[0][0])
            return sim >= threshold
        except Exception:
            return False

    def _is_numeric_value(self, text):
        if pd.isna(text) or str(text).strip() == '':
            return False
        s = str(text).strip()
        return bool(re.match(r'^-?\d+\.?\d*$|^-?\d{1,3}(,\d{3})*(\.\d+)?$', s))

    def _detect_metrics_in_row(self, row):
        metrics_found = []
        i = 0
        n = len(row)
        while i < n:
            cell = row[i]
            if pd.isna(cell) or str(cell).strip() == '':
                i += 1
                continue
            cell_str = str(cell).strip()
            if re.search(r'[a-zA-Z]', cell_str) and not self._is_numeric_value(cell_str):
                values = []
                j = i + 1
                while j < n:
                    next_cell = row[j]
                    # Skip pure empties/spacers and continue scanning
                    if pd.isna(next_cell) or str(next_cell).strip() == '':
                        j += 1
                        continue
                    next_str = str(next_cell).strip()
                    if self._is_numeric_value(next_str):
                        values.append(next_str)
                        j += 1
                        continue
                    # Stop only when a new metric name starts (alpha present)
                    if re.search(r'[a-zA-Z]', next_str):
                        break
                    # Otherwise treat as spacer and continue
                    j += 1
                # Accept blocks with >=1 numeric to capture trailing single values
                if len(values) >= 1:
                    metrics_found.append((cell_str, values, i))
                    i = j
                else:
                    i += 1
            else:
                i += 1
        return metrics_found

    def _find_value_overlap(self, base_values, new_values):
        if not base_values or not new_values:
            return 0
        max_possible_overlap = min(len(base_values), len(new_values))
        for overlap_size in range(1, max_possible_overlap + 1):
            base_tail = [self._normalize_value(v) for v in base_values[-overlap_size:]]
            new_head = [self._normalize_value(v) for v in new_values[:overlap_size]]
            if base_tail == new_head:
                return len(base_values) - overlap_size
        best_offset = len(base_values)
        max_consecutive = 0
        for offset in range(len(base_values)):
            consecutive = 0
            check_len = min(len(new_values), len(base_values) - offset)
            for i in range(check_len):
                if self._normalize_value(base_values[offset + i]) == self._normalize_value(new_values[i]):
                    consecutive += 1
                else:
                    if consecutive > max_consecutive:
                        max_consecutive = consecutive
                        best_offset = offset
                    break
            else:
                if consecutive > max_consecutive:
                    max_consecutive = consecutive
                    best_offset = offset
        if max_consecutive > 0 and max_consecutive < len(new_values):
            return best_offset
        return len(base_values)

    def _deduplicate_row(self, row, verbose=False):
        detected_metrics = self._detect_metrics_in_row(row)
        if not detected_metrics:
            return None
        metric_groups = []
        for metric_name, values, position in detected_metrics:
            matched_group = None
            for group in metric_groups:
                base_name, base_values, _ = group['base_metric']
                # Only group when semantically the same (aliases included)
                if self._are_metrics_similar(metric_name, base_name):
                    matched_group = group
                    break
            if matched_group:
                matched_group['duplicates'].append((metric_name, values, position))
            else:
                metric_groups.append({'base_metric': (metric_name, values, position), 'duplicates': []})
        output_rows = []
        for group in metric_groups:
            base_name, base_values, _ = group['base_metric']
            merged_values = list(base_values)
            for dup_name, dup_values, _ in group['duplicates']:
                offset = self._find_value_overlap(merged_values, dup_values)
                required_length = offset + len(dup_values)
                if required_length > len(merged_values):
                    merged_values.extend([None] * (required_length - len(merged_values)))
                for i, dup_val in enumerate(dup_values):
                    target_idx = offset + i
                    if merged_values[target_idx] is None:
                        merged_values[target_idx] = dup_val
                    else:
                        existing_val = merged_values[target_idx]
                        if self._normalize_value(existing_val) != self._normalize_value(dup_val):
                            if '{' not in str(existing_val):
                                merged_values[target_idx] = f"{existing_val}{{{dup_val}}}"
                            else:
                                # Append additional restatement(s) preserving previous ones
                                merged_values[target_idx] = str(merged_values[target_idx])[:-1] + f",{dup_val}}}"
            output_row = [base_name] + merged_values
            output_rows.append(output_row)
        return output_rows

    def deduplicate_excel_file(self, input_path: str, output_path: Optional[str] = None, verbose: bool = False) -> Optional[str]:
        """Process an Excel file and deduplicate metrics row by row for every sheet."""
        try:
            xls = pd.ExcelFile(input_path)
        except Exception as e:
            print(f"  ❌ Error reading Excel for deduplication: {e}")
            return None
        if output_path is None:
            base, ext = os.path.splitext(input_path)
            output_path = f"{base}_deduplicated{ext}"
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        results = {}
        for sheet_name in xls.sheet_names:
            try:
                df = pd.read_excel(input_path, sheet_name=sheet_name, header=None)
            except Exception as e:
                print(f"  ⚠️ Skipping sheet '{sheet_name}' (read error): {e}")
                continue
            processed_rows = []
            rows_with_patterns = 0
            for _, row in df.iterrows():
                row_data = row.tolist()
                deduped = self._deduplicate_row(row_data, verbose=verbose)
                if deduped:
                    rows_with_patterns += 1
                    for drow in deduped:
                        processed_rows.append(drow)
                else:
                    if any(not pd.isna(cell) and str(cell).strip() != '' for cell in row_data):
                        processed_rows.append(row_data)
            if processed_rows:
                max_cols = max(len(r) for r in processed_rows)
                padded = [r + [None] * (max_cols - len(r)) for r in processed_rows]
                col_names = ['Metric Name'] + [f'Value {i+1}' for i in range(max_cols - 1)]
                out_df = pd.DataFrame(padded, columns=col_names[:max_cols])
                out_df.to_excel(writer, sheet_name=sheet_name, index=False)
                results[sheet_name] = {
                    'total_rows': len(df),
                    'rows_with_patterns': rows_with_patterns,
                    'output_rows': len(out_df),
                    'max_values': max_cols - 1
                }
            else:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                results[sheet_name] = {'error': 'No data processed'}
        try:
            writer.close()
        except Exception:
            pass
        print(f"  ✅ Deduplicated file created: {os.path.basename(output_path)}")
        return output_path
    
    def _normalize_form_types(self, form_type_input: str) -> List[str]:
        """Normalize user-provided form types into a list like ["10-K", "10-Q"]. Supports
        comma-separated inputs and aliases like 10K/10Q, and the keyword 'all'."""
        if not form_type_input:
            return ["10-K"]
        raw = form_type_input.strip().lower()
        if raw == "all":
            return ["10-K", "10-Q"]
        parts = [p.strip() for p in raw.replace(";", ",").split(",") if p.strip()]
        normalized = []
        for p in parts:
            p_clean = p.replace(" ", "").replace("_", "-")
            if p_clean in ["10k", "10-k"]:
                normalized.append("10-K")
            elif p_clean in ["10q", "10-q"]:
                normalized.append("10-Q")
            # 8-K removed - no longer supported
            else:
                # Fallback: keep as-is uppercased, but filter out 8-K
                form = p.upper()
                if form not in ["8-K", "8K"]:
                    normalized.append(form)
        # De-duplicate, preserve order
        seen = set(); out = []
        for f in normalized:
            if f not in seen:
                out.append(f); seen.add(f)
        return out or ["10-K"]

    def _build_form_query(self, form_types: List[str]) -> str:
        """Build the formType query segment for SEC-API. Supports multiple OR form types."""
        if not form_types:
            return 'formType:"10-K"'
        if len(form_types) == 1:
            return f'formType:"{form_types[0]}"'
        ors = " OR ".join([f'formType:"{ft}"' for ft in form_types])
        return f"({ors})"

    def search_filings(self, ticker: str, form_type: str = "10-K", start_date: Optional[str] = None, end_date: Optional[str] = None) -> List[Dict]:
        """Search for SEC filings. Supports count or date range.
        - form_type: single or comma-separated (incl. 'all')
        - limit: number of most recent filings (ignored if date range is provided)
        - start_date/end_date: ISO dates YYYY-MM-DD to filter by filedAt
        """
        print(f"🔍 Searching for {form_type} filings for {ticker}...")
        form_types = self._normalize_form_types(form_type)
        query_str = f"ticker:{ticker} AND {self._build_form_query(form_types)}"
        if start_date or end_date:
            # Lucene-style range appended to query
            if start_date and end_date:
                query_str += f" AND filedAt:[{start_date} TO {end_date}]"
            elif start_date:
                query_str += f" AND filedAt:[{start_date} TO *]"
            elif end_date:
                query_str += f" AND filedAt:[* TO {end_date}]"
        query_payload = {
            "query": query_str,
            "from": "0",
            "size": "200",
            "sort": [{"filedAt": {"order": "desc"}}]
        }
        
        url = f"{self.base_url}?token={self.api_key}"
        
        try:
            response = requests.post(url, json=query_payload, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            filings = data.get("filings", [])
            print(f"📊 Found {len(filings)} filings for {ticker}")
            return filings
            
        except Exception as e:
            print(f"❌ Error searching filings: {e}")
            return []

    def validate_ticker(self, ticker: str) -> bool:
        """Validate if ticker exists by making a test API call"""
        try:
            print(f"🔍 Validating ticker: {ticker}...")
            # Make a simple test query to see if ticker exists
            test_query = f"ticker:{ticker}"
            query_payload = {
                "query": test_query,
                "from": "0",
                "size": "1"
            }
            
            url = f"{self.base_url}?token={self.api_key}"
            response = requests.post(url, json=query_payload, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            filings = data.get("filings", [])
            if filings:
                print(f"✅ Ticker {ticker} is valid")
                return True
            else:
                print(f"❌ Ticker {ticker} not found")
                return False
                
        except Exception as e:
            print(f"❌ Error validating ticker {ticker}: {e}")
            return False

    def _derive_filing_label(self, form_type: str, filing: Dict) -> str:
        """Return a user-friendly label with better naming for 10-K and 10-Q filings.
        Uses periodOfReport when available; falls back to filedAt.
        """
        try:
            ft = (form_type or "").upper()
            period = (filing or {}).get("periodOfReport") or (filing or {}).get("filedAt")
            if not period:
                return "Unknown"
            date_part = str(period)[:10]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            year = dt.year
            month = dt.month
            
            if ft.startswith("10-K"):
                # For 10-K, use fiscal year with better formatting
                return f"FY{year}"
            elif ft.startswith("10-Q"):
                # For 10-Q, use quarter with year and month info
                quarter = 1 if month <= 3 else 2 if month <= 6 else 3 if month <= 9 else 4
                month_name = dt.strftime("%b")  # Jan, Feb, Mar, etc.
                return f"Q{quarter}_{year}_{month_name}"
            elif ft.startswith("8-K"):
                # For 8-K, use date with better formatting
                return f"8K_{year}_{month:02d}_{dt.day:02d}"
            else:
                # Generic fallback: use date
                return f"{year}_{month:02d}_{dt.day:02d}"
        except Exception:
            return "Unknown"

    def _compute_simple_label(self, form_type: str, filing_or_date) -> str:
        """Return simple label like '2024-10K' or '2023-Q2' using periodOfReport/filedAt.
        Accepts either a filing dict with keys periodOfReport/filedAt/filed_at or a date string.
        """
        try:
            ft = (form_type or "").upper()
            period = None
            if isinstance(filing_or_date, dict):
                d = filing_or_date or {}
                period = d.get("periodOfReport") or d.get("filedAt") or d.get("filed_at") or d.get("period_of_report")
            else:
                period = filing_or_date
            if not period:
                return "Unknown"
            date_part = str(period)[:10]
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            year = dt.year
            month = dt.month
            if ft.startswith("10-K"):
                return f"{year}-10K"
            if ft.startswith("10-Q"):
                quarter = 1 if month <= 3 else 2 if month <= 6 else 3 if month <= 9 else 4
                return f"{year}-Q{quarter}"
            return f"{year}"
        except Exception:
            return "Unknown"
    
    def download_with_curl(self, url: str, output_file: str) -> bool:
        """Download using curl command (often bypasses restrictions)"""
        try:
            print(f"🔄 Trying curl download...")
            
            # Use curl with proper headers
            cmd = [
                'curl',
                '-L',  # Follow redirects
                '-H', 'User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                '-H', 'Accept: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,*/*',
                '-H', 'Accept-Language: en-US,en;q=0.9',
                '-H', 'Accept-Encoding: gzip, deflate, br',
                '-H', 'Connection: keep-alive',
                '-H', 'Upgrade-Insecure-Requests: 1',
                '--compressed',
                '--retry', '3',
                '--retry-delay', '2',
                '-o', output_file,
                url
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0 and os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                # Check if it's actually an Excel file or an error message
                with open(output_file, 'rb') as f:
                    content = f.read(100)  # Read first 100 bytes
                    if b'<Error>' in content or b'NoSuchKey' in content:
                        print(f"⚠️ File not available (SEC error): {output_file}")
                        os.remove(output_file)  # Remove the error file
                        return False
                    elif content.startswith(b'PK'):  # Excel files start with PK
                        print(f"✅ Curl download successful: {output_file}")
                        return True
                    else:
                        print(f"⚠️ Downloaded file may not be valid Excel: {output_file}")
                        return True
            else:
                print(f"❌ Curl download failed: {result.stderr}")
                return False
                
        except Exception as e:
            print(f"❌ Curl download error: {e}")
            return False
    
    def download_with_wget(self, url: str, output_file: str) -> bool:
        """Download using wget command"""
        try:
            print(f"🔄 Trying wget download...")
            
            cmd = [
                'wget',
                '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                '--header=Accept: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                '--header=Accept-Language: en-US,en;q=0.9',
                '--retry=3',
                '--wait=2',
                '--random-wait',
                '-O', output_file,
                url
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0 and os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                print(f"✅ Wget download successful: {output_file}")
                return True
            else:
                print(f"❌ Wget download failed: {result.stderr}")
                return False
                
        except Exception as e:
            print(f"❌ Wget download error: {e}")
            return False
    
    def download_with_requests_advanced(self, url: str, output_file: str) -> bool:
        """Advanced requests download with session management"""
        try:
            print(f"🔄 Trying advanced requests download...")
            
            session = requests.Session()
            
            # Rotate user agent
            user_agent = random.choice(self.user_agents)
            session.headers.update({
                'User-Agent': user_agent,
                'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel,application/octet-stream,*/*',
                'Accept-Language': 'en-US,en;q=0.9',
                'Accept-Encoding': 'gzip, deflate, br',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
                'Sec-Fetch-Dest': 'document',
                'Sec-Fetch-Mode': 'navigate',
                'Sec-Fetch-Site': 'none',
                'Cache-Control': 'max-age=0',
                'Referer': 'https://www.sec.gov/'
            })
            
            # Add random delay
            time.sleep(random.uniform(1, 3))
            
            response = session.get(url, timeout=30, allow_redirects=True)
            
            if response.status_code == 200:
                with open(output_file, 'wb') as f:
                    f.write(response.content)
                
                if os.path.getsize(output_file) > 0:
                    print(f"✅ Advanced requests download successful: {output_file}")
                    return True
                else:
                    print(f"❌ Downloaded file is empty")
                    return False
            else:
                print(f"❌ HTTP {response.status_code}: {response.reason}")
                return False
                
        except Exception as e:
            print(f"❌ Advanced requests error: {e}")
            return False
    
    def _discover_candidate_excel_files(self, cik: str, clean_accession: str) -> List[str]:
        """Discover possible Excel filenames within a filing by reading the index.json directory listing."""
        candidates = []
        index_url = f"{self.sec_base_url}/Archives/edgar/data/{cik}/{clean_accession}/index.json"
        try:
            resp = requests.get(index_url, timeout=20)
            if resp.status_code == 200:
                data = resp.json()
                items = (data.get("directory", {}) or {}).get("item", [])
                excel_names = [it.get("name") for it in items if isinstance(it, dict) and isinstance(it.get("name"), str) and (it.get("name").lower().endswith(".xlsx") or it.get("name").lower().endswith(".xls"))]
                # Prioritize Financial_Report first, then names containing 'financial', then others
                prio1 = [n for n in excel_names if "financial_report" in n.lower()]
                prio2 = [n for n in excel_names if n not in prio1 and ("financial" in n.lower() or "statements" in n.lower())]
                prio3 = [n for n in excel_names if n not in prio1 and n not in prio2]
                candidates = prio1 + prio2 + prio3
        except Exception:
            pass
        return candidates

    def download_excel_file(self, accession_no: str, cik: str, output_dir: str, filename: str) -> Optional[str]:
        """Try multiple download methods for a single filing. Attempts several candidate Excel files."""
        clean_accession = accession_no.replace("-", "")
        base_dir = f"{self.sec_base_url}/Archives/edgar/data/{cik}/{clean_accession}"
        # Build candidate URLs: default Financial_Report.xlsx plus discovered ones
        candidate_files = ["Financial_Report.xlsx"]
        discovered = self._discover_candidate_excel_files(cik, clean_accession)
        for name in discovered:
            if name not in candidate_files:
                candidate_files.append(name)
        output_file = os.path.join(output_dir, filename)
        
        print(f"\n📥 Downloading: {filename}")
        print(f"🔍 Available Excel files: {candidate_files}")
        
        for excel_name in candidate_files:
            excel_url = f"{base_dir}/{excel_name}"
            print(f"🔗 Trying URL: {excel_url}")
            # Try multiple download methods per candidate
            methods = [
                ("curl", lambda: self.download_with_curl(excel_url, output_file)),
                ("wget", lambda: self.download_with_wget(excel_url, output_file)),
                ("requests", lambda: self.download_with_requests_advanced(excel_url, output_file))
            ]
            for method_name, method_func in methods:
                try:
                    if method_func():
                        # Process the downloaded file to extract only the 3 main statements
                        processed_file = self._extract_main_statements(output_file, filename)
                        if processed_file:
                            self.last_used_excel_url = excel_url
                            self.downloaded_files.append(processed_file)
                            return processed_file
                        else:
                            self.last_used_excel_url = excel_url
                            self.downloaded_files.append(output_file)
                            return output_file
                except Exception as e:
                    print(f"⚠️ {method_name} method failed: {e}")
                    continue
            # If we reach here, this candidate failed; remove partial file if created
            try:
                if os.path.exists(output_file):
                    os.remove(output_file)
            except Exception:
                pass
        
        print(f"❌ All download methods failed for {filename}")
        print(f"💡 This filing may not have Excel files available, or they may be in a different format")
        print(f"🔍 Check the filing directly at: {base_dir}/")
        return None
    
    def _extract_main_statements(self, input_file: str, original_filename: str) -> Optional[str]:
        """Extract only the 3 main financial statements from the downloaded Excel file"""
        try:
            print(f"🔧 Processing {original_filename} to extract main statements...")
            
            # Read the original Excel file
            with pd.ExcelFile(input_file) as xl_file:
                # Find the main statements
                main_statements = {
                    'income_statement': None,
                    'balance_sheet': None,
                    'cash_flow': None
                }
                
                # First pass: Look for obvious main statements by name
                priority_sheets = []
                other_sheets = []
                
                for sheet_name in xl_file.sheet_names:
                    sheet_lower = sheet_name.lower()
                    if any(keyword in sheet_lower for keyword in [
                        'consolidated statements of operations', 'consolidated statement of operations',
                        'consolidated statements of earnings', 'consolidated statement of earnings',
                        'consolidated statements of income', 'consolidated statement of income',
                        'consolidated balance sheet', 'consolidated balance sheets',
                        'consolidated statements of cash', 'consolidated statement of cash flows',
                        'consolidated statements of cash flows'
                    ]):
                        priority_sheets.append(sheet_name)
                    else:
                        other_sheets.append(sheet_name)
                
                # Process priority sheets first
                for sheet_name in priority_sheets + other_sheets:
                    try:
                        # Universal: preserve visible first rows (including year headers) across companies
                        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
                        if df.empty or len(df.columns) < 2 or len(df) < 5:
                            continue
                        
                        # Analyze the sheet content to determine statement type
                        # Pass raw DataFrame (no headers) for robust detection across formats
                        statement_type = self._analyze_sheet_content(df, sheet_name)
                        
                        if statement_type and not main_statements[statement_type]:
                            main_statements[statement_type] = {
                                'name': self._get_clean_statement_name(statement_type),
                                'data': df,
                                'original_sheet': sheet_name
                            }
                            print(f"  ✅ Found {self._get_clean_statement_name(statement_type)}: {sheet_name}")
                            
                            # Stop if we found all three
                            if all(main_statements.values()):
                                break
                    
                    except Exception as e:
                        continue
            
            # Check if we found all three statements
            if not all(main_statements.values()):
                missing = [k for k, v in main_statements.items() if v is None]
                print(f"  ⚠️ Missing statements: {missing}")
                return None
            
            # Create new Excel file with all statements in one sheet
            processed_filename = original_filename.replace('.xlsx', '_Individual_Financials.xlsx')
            processed_file = os.path.join(os.path.dirname(input_file), processed_filename)
            
            # Create new workbook with single sheet
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Consolidated Financials"
            
            current_row = 1
            
            # Add each main statement to the same sheet
            for statement_type, statement_data in main_statements.items():
                if statement_data:
                    # Add statement title
                    title_cell = ws.cell(row=current_row, column=1)
                    title_cell.value = statement_data['name']
                    title_cell.font = Font(bold=True, size=14)
                    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                    
                    # Merge cells for title
                    ws.merge_cells(f'A{current_row}:{get_column_letter(len(statement_data["data"].columns))}{current_row}')
                    current_row += 1
                    
                    # Use DataFrame as-is to preserve top rows (year headers)
                    df = statement_data['data']
                    
                    # Write data to worksheet
                    # Write rows exactly as present; r_idx==1 corresponds to original first row
                    for r_idx, row in enumerate(pd.DataFrame(df).itertuples(index=False), 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=c_idx)
                            cell.value = value
                            
                            # Format header row
                            if r_idx == 1:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
                            
                            # Format numeric values in accounting format
                            elif isinstance(value, (int, float)) and not pd.isna(value):
                                if abs(value) >= 1000000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting millions (no decimals)
                                elif abs(value) >= 1000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting thousands (no decimals)
                                else:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Accounting small numbers (keep decimals)
                        
                        current_row += 1
                    
                    # Add spacing between statements
                    current_row += 2
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the processed file
            workbook.save(processed_file)
            
            # Small delay to ensure file handles are closed
            time.sleep(0.5)
            
            # Remove the original file
            try:
                os.remove(input_file)
            except PermissionError:
                print(f"  ⚠️ Could not remove original file (may be in use): {input_file}")
            
            print(f"  ✅ Created individual financials file: {processed_filename}")
            print(f"     - Income Statement: {main_statements['income_statement']['original_sheet']}")
            print(f"     - Balance Sheet: {main_statements['balance_sheet']['original_sheet']}")
            print(f"     - Cash Flow Statement: {main_statements['cash_flow']['original_sheet']}")
            print(f"     - All statements in single sheet")
            
            return processed_file
            
        except Exception as e:
            print(f"  ❌ Error processing file: {e}")
            return None
    
    def _analyze_sheet_content(self, df: pd.DataFrame, sheet_name: str) -> Optional[str]:
        """Universal content analysis to detect main financial statements across all companies"""
        # Convert all data to string for analysis
        content_str = ' '.join([str(val) for val in df.values.flatten() if pd.notna(val)]).lower()
        sheet_lower = sheet_name.lower()
        
        # Comprehensive Income Statement indicators
        income_indicators = [
            # Revenue terms
            'revenue', 'net sales', 'total revenue', 'gross revenue', 'sales revenue',
            'service revenue', 'product revenue', 'net revenue', 'total net sales',
            # Profit/Loss terms
            'gross profit', 'gross income', 'operating income', 'operating profit',
            'net income', 'net earnings', 'earnings', 'profit', 'loss', 'net loss',
            'income from operations', 'operating earnings', 'operating profit',
            # Cost terms
            'cost of revenue', 'cost of sales', 'cost of goods sold', 'cogs',
            'gross margin', 'operating expenses', 'operating costs',
            'selling general administrative', 'sga', 'sg&a',
            'research development', 'rd expenses', 'research and development',
            # Financial terms
            'interest expense', 'interest income', 'interest revenue',
            'income before taxes', 'pretax income', 'provision for taxes',
            'income tax expense', 'income tax benefit', 'tax expense',
            # EPS terms
            'basic earnings per share', 'diluted earnings per share', 'eps',
            'basic eps', 'diluted eps', 'earnings per share',
            # Other income statement terms
            'other income', 'other expense', 'non-operating income',
            'depreciation', 'amortization', 'impairment', 'restructuring'
        ]
        
        # Comprehensive Balance Sheet indicators
        balance_indicators = [
            # Asset terms
            'assets', 'total assets', 'current assets', 'non-current assets',
            'long-term assets', 'fixed assets', 'property plant equipment', 'ppe',
            'intangible assets', 'goodwill', 'other assets', 'other current assets',
            'cash and cash equivalents', 'cash equivalents', 'marketable securities',
            'accounts receivable', 'receivables', 'inventory', 'inventories',
            'prepaid expenses', 'deferred tax assets', 'investments',
            'short-term investments', 'long-term investments', 'available-for-sale',
            'held-to-maturity', 'trading securities', 'equity securities',
            'debt securities', 'derivative assets', 'other current assets',
            'property and equipment', 'land', 'buildings', 'machinery',
            'equipment', 'furniture', 'fixtures', 'leasehold improvements',
            'accumulated depreciation', 'net property plant equipment',
            # Liability terms
            'liabilities', 'total liabilities', 'current liabilities',
            'non-current liabilities', 'long-term liabilities', 'long term debt',
            'accounts payable', 'payables', 'accrued expenses', 'accrued liabilities',
            'deferred revenue', 'unearned revenue', 'deferred tax liabilities',
            'notes payable', 'bonds payable', 'lease liabilities',
            'short-term debt', 'current portion of long-term debt',
            'accrued compensation', 'accrued benefits', 'accrued interest',
            'accrued taxes', 'income taxes payable', 'other current liabilities',
            'long-term debt', 'senior notes', 'subordinated notes',
            'convertible debt', 'deferred tax liabilities', 'pension liabilities',
            'postretirement benefits', 'other long-term liabilities',
            # Equity terms
            'equity', 'stockholders equity', 'shareholders equity', 'shareholders\' equity',
            'total equity', 'total stockholders equity', 'total shareholders equity',
            'retained earnings', 'common stock', 'preferred stock', 'treasury stock',
            'additional paid-in capital', 'paid-in capital', 'capital surplus',
            'accumulated other comprehensive income', 'aoci', 'other comprehensive income',
            'treasury shares', 'outstanding shares', 'issued shares',
            'par value', 'stated value', 'paid-in capital in excess of par',
            'additional paid-in capital', 'capital in excess of par value',
            'retained deficit', 'accumulated deficit', 'other equity',
            # Balance sheet specific terms
            'as of', 'at december', 'at june', 'at september', 'at march',
            'fiscal year', 'fiscal period', 'reporting period',
            'consolidated balance sheet', 'consolidated balance sheets',
            'statement of financial position', 'statement of financial condition',
            'consolidated statement of financial position'
        ]
        
        # Comprehensive Cash Flow indicators
        cash_flow_indicators = [
            # Operating activities
            'cash flows from operating activities', 'operating activities',
            'net cash provided by operating activities', 'net cash used in operating activities',
            'cash from operations', 'operating cash flow', 'operating cash flows',
            'cash provided by operating activities', 'cash used in operating activities',
            # Investing activities
            'cash flows from investing activities', 'investing activities',
            'net cash provided by investing activities', 'net cash used in investing activities',
            'cash from investing', 'investing cash flow', 'investing cash flows',
            'cash provided by investing activities', 'cash used in investing activities',
            # Financing activities
            'cash flows from financing activities', 'financing activities',
            'net cash provided by financing activities', 'net cash used in financing activities',
            'cash from financing', 'financing cash flow', 'financing cash flows',
            'cash provided by financing activities', 'cash used in financing activities',
            # Cash position terms
            'net increase in cash', 'net decrease in cash', 'net change in cash',
            'cash and cash equivalents at beginning', 'cash and cash equivalents at end',
            'cash at beginning of period', 'cash at end of period',
            'beginning cash', 'ending cash', 'cash balance',
            'cash and cash equivalents at the beginning', 'cash and cash equivalents at the end',
            # Specific cash flow items
            'depreciation and amortization', 'stock-based compensation',
            'changes in working capital', 'working capital changes',
            'purchases of property plant equipment', 'capital expenditures',
            'proceeds from issuance of debt', 'repayments of debt',
            'dividends paid', 'share repurchases', 'stock repurchases',
            'purchases of common stock', 'proceeds from common stock',
            'proceeds from long-term debt', 'repayments of long-term debt',
            'acquisitions', 'disposals', 'purchases of investments',
            'proceeds from sales of investments', 'purchases of marketable securities',
            'proceeds from sales of marketable securities', 'deferred income taxes',
            'accounts receivable', 'inventory', 'accounts payable',
            'accrued expenses', 'other operating activities',
            'other investing activities', 'other financing activities'
        ]
        
        # Count matches for each statement type
        income_score = sum(1 for indicator in income_indicators if indicator in content_str)
        balance_score = sum(1 for indicator in balance_indicators if indicator in content_str)
        cash_flow_score = sum(1 for indicator in cash_flow_indicators if indicator in content_str)
        
        # Enhanced sheet name analysis
        name_income_indicators = [
            'income', 'operations', 'earnings', 'profit', 'inco', 'p&l', 'pnl',
            'statement of operations', 'statement of earnings', 'statement of income',
            'consolidated statements of operations', 'consolidated statement of operations',
            'consolidated statements of earnings', 'consolidated statement of earnings',
            'consolidated statements of income', 'consolidated statement of income'
        ]
        
        name_balance_indicators = [
            'balance', 'financial position', 'assets', 'liabilities', 'equity',
            'statement of financial position', 'consolidated balance sheet',
            'consolidated balance sheets', 'statement of financial condition',
            'consolidated statement of financial position'
        ]
        
        name_cash_flow_indicators = [
            'cash', 'flow', 'flows', 'statement of cash flows', 'statement of cash flow',
            'consolidated statements of cash flows', 'consolidated statement of cash flows',
            'consolidated statements of cash flow', 'consolidated statement of cash flow',
            'cash flows statement', 'cash flow statement'
        ]
        
        # Add name-based scores (highest priority for exact matches)
        for indicator in name_income_indicators:
            if indicator in sheet_lower:
                income_score += 10 if len(indicator) > 15 else 8  # Much higher score for specific terms
        
        for indicator in name_balance_indicators:
            if indicator in sheet_lower:
                balance_score += 10 if len(indicator) > 15 else 8
        
        for indicator in name_cash_flow_indicators:
            if indicator in sheet_lower:
                cash_flow_score += 10 if len(indicator) > 15 else 8
        
        # Check for year columns (financial statements typically have multiple years)
        has_year_columns = any(str(col).isdigit() and len(str(col)) == 4 for col in df.columns)
        if has_year_columns:
            income_score += 2
            balance_score += 2
            cash_flow_score += 2
        
        # Check for financial statement structure (multiple columns with numeric data)
        numeric_columns = sum(1 for col in df.columns if df[col].dtype in ['int64', 'float64'] or 
                             any(isinstance(val, (int, float)) and not pd.isna(val) for val in df[col].iloc[:5]))
        if numeric_columns >= 2:  # At least 2 numeric columns
            income_score += 1
            balance_score += 1
            cash_flow_score += 1
        
        # Check for typical financial statement row count (not too few, not too many)
        row_count = len(df)
        if 5 <= row_count <= 50:  # Typical range for main statements
            income_score += 1
            balance_score += 1
            cash_flow_score += 1
        
        # Determine the statement type with highest score
        scores = {
            'income_statement': income_score,
            'balance_sheet': balance_score,
            'cash_flow': cash_flow_score
        }
        
        # Debug information
        if max(scores.values()) >= 2:  # Lower threshold for better detection
            print(f"    📊 Analysis for '{sheet_name}': Income={income_score}, Balance={balance_score}, Cash Flow={cash_flow_score}")
            
        # Additional debugging for balance sheet detection
        if 'balance' in sheet_lower or balance_score > 0:
            print(f"    🔍 Balance sheet analysis for '{sheet_name}':")
            print(f"      - Sheet name contains 'balance': {'balance' in sheet_lower}")
            print(f"      - Balance score: {balance_score}")
            print(f"      - Content analysis: {balance_score} balance indicators found")
        
        # PRIORITY 1: Check for exact name matches first (highest priority)
        if 'consolidated statements of cash' in sheet_lower or 'consolidated statement of cash' in sheet_lower:
            print(f"    ✅ Cash flow detected by exact name match: {sheet_name}")
            return 'cash_flow'
        
        if 'consolidated balance sheet' in sheet_lower or 'consolidated balance sheets' in sheet_lower:
            return 'balance_sheet'
        
        if 'consolidated statements of operations' in sheet_lower or 'consolidated statement of operations' in sheet_lower:
            return 'income_statement'
        
        if 'consolidated statements of income' in sheet_lower or 'consolidated statement of income' in sheet_lower:
            return 'income_statement'
        
        if 'consolidated statements of earnings' in sheet_lower or 'consolidated statement of earnings' in sheet_lower:
            return 'income_statement'
        
        # PRIORITY 2: Check for partial name matches
        if 'cash' in sheet_lower and 'flow' in sheet_lower and scores['cash_flow'] >= 8:
            print(f"    ✅ Cash flow detected by partial name match: {sheet_name} (score: {scores['cash_flow']})")
            return 'cash_flow'
        
        if 'balance' in sheet_lower and scores['balance_sheet'] >= 6:  # Lowered threshold for balance sheet
            print(f"    ✅ Balance sheet detected by partial name match: {sheet_name} (score: {scores['balance_sheet']})")
            return 'balance_sheet'
        
        if any(keyword in sheet_lower for keyword in ['income', 'operations', 'earnings']) and scores['income_statement'] >= 8:
            return 'income_statement'
        
        # PRIORITY 3: Use content-based scoring as fallback
        max_score = max(scores.values())
        min_threshold = 3
        
        if max_score >= min_threshold:
            for statement_type, score in scores.items():
                if score == max_score:
                    # Additional validation for specific statement types
                    if statement_type == 'income_statement' and score < 6:
                        continue
                    if statement_type == 'balance_sheet' and score < 5:  # Lowered threshold for balance sheet
                        continue
                    if statement_type == 'cash_flow' and score < 6:
                        continue
                    
                    return statement_type
        
        return None
    
    def _get_clean_statement_name(self, statement_type: str) -> str:
        """Get clean, standardized statement name"""
        names = {
            'income_statement': 'Income Statement',
            'balance_sheet': 'Balance Sheet',
            'cash_flow': 'Cash Flow Statement'
        }
        return names.get(statement_type, statement_type)
    
    def _enhance_cash_flow_extraction(self, df: pd.DataFrame) -> pd.DataFrame:
        """Enhance cash flow statement extraction to capture more line items"""
        if df.empty:
            return df
        
        enhanced_rows = []
        
        for idx, row in df.iterrows():
            row_values = [str(val) if pd.notna(val) else "" for val in row.values]
            row_text = " ".join(row_values).lower()
            
            # Check if this row contains cash flow related content
            cash_flow_indicators = [
                'operating activities', 'investing activities', 'financing activities',
                'cash flows from operating', 'cash flows from investing', 'cash flows from financing',
                'net cash provided by operating', 'net cash used in operating',
                'net cash provided by investing', 'net cash used in investing',
                'net cash provided by financing', 'net cash used in financing',
                'depreciation and amortization', 'stock-based compensation',
                'changes in working capital', 'working capital changes',
                'purchases of property plant equipment', 'capital expenditures',
                'proceeds from issuance of debt', 'repayments of debt',
                'dividends paid', 'share repurchases', 'stock repurchases',
                'purchases of common stock', 'proceeds from common stock',
                'proceeds from long-term debt', 'repayments of long-term debt',
                'acquisitions', 'disposals', 'purchases of investments',
                'proceeds from sales of investments', 'purchases of marketable securities',
                'proceeds from sales of marketable securities', 'deferred income taxes',
                'accounts receivable', 'inventory', 'accounts payable',
                'accrued expenses', 'other operating activities',
                'other investing activities', 'other financing activities',
                'cash and cash equivalents at beginning', 'cash and cash equivalents at end',
                'net increase in cash', 'net decrease in cash', 'net change in cash'
            ]
            
            # Include rows that have cash flow indicators or contain numeric data
            has_cash_flow_content = any(indicator in row_text for indicator in cash_flow_indicators)
            has_numeric_data = any(isinstance(val, (int, float)) and not pd.isna(val) for val in row.values)
            
            if has_cash_flow_content or has_numeric_data:
                enhanced_rows.append(row)
        
        if enhanced_rows:
            return pd.DataFrame(enhanced_rows).reset_index(drop=True)
        else:
            return df
    
    def create_batch_download_script(self, excel_info: List[Dict], ticker: str, form_type: str):
        """Create an advanced batch download script"""
        script_content = f"""@echo off
REM Advanced SEC Excel Downloader
REM Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

echo Starting automated SEC Excel download for {ticker} {form_type}...
echo.

REM Create download directory
if not exist "SEC_Excel_Downloads_{ticker}_{form_type}" mkdir "SEC_Excel_Downloads_{ticker}_{form_type}"
cd "SEC_Excel_Downloads_{ticker}_{form_type}"

echo Download directory created: SEC_Excel_Downloads_{ticker}_{form_type}
echo.

"""
        
        for info in excel_info:
            label = info.get('label') or f"{info.get('filing_number', '')}"
            filename = f"{ticker}_{form_type}_{label}_{info['accession_no'].replace('-', '_')}.xlsx"
            excel_url = info.get('excel_url') or ""
            
            script_content += f"""REM Filing #{info['filing_number']}: {info['company_name']}
echo Downloading filing #{info['filing_number']}...
echo URL: {excel_url}

REM Try curl first
curl -L -H "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36" -H "Accept: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" --retry 3 --retry-delay 2 -o "{filename}" "{excel_url}"
if %errorlevel% equ 0 (
    echo SUCCESS: Downloaded {filename}
) else (
    echo ERROR: Curl failed for {filename}
)

REM Add delay between downloads
timeout /t 3 /nobreak >nul
echo.

"""
        
        script_content += """
echo.
echo Download complete! Check the SEC_Excel_Downloads folder.
pause
"""
        
        script_filename = f"download_{ticker}_{form_type}_advanced.bat"
        with open(script_filename, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        print(f"📝 Created advanced batch script: {script_filename}")
        return script_filename
    
    def create_master_consolidated_file(self, individual_files: List[str], ticker: str, form_type: str, output_dir: str):
        """Create master consolidated file with all years horizontally aligned"""
        try:
            print(f"\n🔄 Creating master consolidated file for {ticker}...")
            
            # Parse all individual files
            all_years_data = {}
            
            for file_path in individual_files:
                if not os.path.exists(file_path):
                    continue
                    
                # Extract year from filename
                filename = os.path.basename(file_path)
                year = self._extract_year_from_filename(filename)
                
                # Parse the individual file
                statements = self._parse_individual_financial_file(file_path)
                if statements:
                    all_years_data[year] = statements
                    print(f"  ✅ Parsed {year}: {len([s for s in statements.values() if s is not None])} statements")
            
            if not all_years_data:
                print("  ❌ No valid data found for consolidation")
                return None
            
            # Create master consolidated file
            master_file = os.path.join(output_dir, f"{ticker}_{form_type}_Master_Consolidated_Financials.xlsx")
            self._create_horizontal_consolidated_excel(all_years_data, master_file, ticker)
            
            print(f"  ✅ Created master consolidated file: {os.path.basename(master_file)}")
            return master_file
            
        except Exception as e:
            print(f"  ❌ Error creating master consolidated file: {e}")
            return None
    
    def create_master_consolidated_file_with_years(self, file_year_mapping: List[tuple], ticker: str, form_type: str, output_dir: str):
        """Create master consolidated file with known years"""
        try:
            print(f"\n🔄 Creating master consolidated file for {ticker}...")
            
            # Parse all individual files with their years
            all_years_data = {}
            
            for file_path, year in file_year_mapping:
                if not os.path.exists(file_path):
                    continue
                
                # Parse the individual file
                statements = self._parse_individual_financial_file(file_path)
                if statements:
                    all_years_data[year] = statements
                    found_statements = [k for k, v in statements.items() if v is not None]
                    print(f"  ✅ Parsed {year}: {len(found_statements)} statements - {found_statements}")
                    
                    # Debug: Show data structure for each year
                    for stmt_type, df in statements.items():
                        if df is not None and not df.empty:
                            print(f"    📊 {stmt_type}: {df.shape[0]} rows, {df.shape[1]} cols")
                            if stmt_type == 'income_statement' and year == '2023':
                                print(f"    🔍 2023 Income Statement sample:")
                                print(f"       Columns: {list(df.columns)[:5]}...")
                                print(f"       First few rows: {df.iloc[:3, :3].values.tolist()}")
                else:
                    print(f"  ⚠️ No statements found for {year}")
            
            if not all_years_data:
                print("  ❌ No valid data found for consolidation")
                return None
            
            # Create master consolidated file
            master_file = os.path.join(output_dir, f"{ticker}_{form_type}_Master_Consolidated_Financials.xlsx")
            self._create_horizontal_consolidated_excel(all_years_data, master_file, ticker)
            
            print(f"  ✅ Created master consolidated file: {os.path.basename(master_file)}")
            return master_file
            
        except Exception as e:
            print(f"  ❌ Error creating master consolidated file: {e}")
            return None
    
    def _extract_year_from_filename(self, filename: str) -> str:
        """Extract year from filename - universal approach"""
        import re
        
        # Method 1: Look for 4-digit year anywhere in filename (most reliable)
        year_match = re.search(r'20\d{2}', filename)
        if year_match:
            return year_match.group()
        
        # Method 2: Look for 2-digit year that could be 20XX
        two_digit_year = re.search(r'(?<![0-9])([2-9][0-9])(?![0-9])', filename)
        if two_digit_year:
            year_part = two_digit_year.group(1)
            if int(year_part) >= 20:  # Assume 20XX for years 20-99
                return f"20{year_part}"
        
        # Method 3: Extract from accession number pattern (SEC specific)
        # Pattern: 10-digit CIK_2-digit year_6-digit sequence
        # Note: SEC filing year is one year ahead of fiscal year
        accession_match = re.search(r'(\d{10})_(\d{2})_(\d{6})', filename)
        if accession_match:
            year_part = accession_match.group(2)
            if int(year_part) >= 20:  # Valid year range
                # Convert filing year to fiscal year (filing year - 1)
                fiscal_year = int(f"20{year_part}") - 1
                return str(fiscal_year)
        
        return "Unknown"
    
    def _parse_individual_financial_file(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """Parse individual financial file and extract the 3 statements - universal approach"""
        try:
            # Read the Excel file
            df = pd.read_excel(file_path, sheet_name=0, header=None)
            
            statements = {
                'income_statement': None,
                'balance_sheet': None,
                'cash_flow': None
            }
            
            # Find statement boundaries using multiple detection methods
            statement_boundaries = self._find_statement_boundaries(df)
            
            # Extract each statement with robust parsing
            for stmt_type, (start_row, end_row) in statement_boundaries.items():
                if start_row is not None and end_row is not None:
                    stmt_data = df.iloc[start_row:end_row+1].copy()
                    
                    # Detect years within this statement's data
                    years_in_data = self._detect_years_in_statement_data(stmt_data, stmt_type)
                    if years_in_data:
                        print(f"  📅 Years detected in {stmt_type}: {years_in_data}")
                    else:
                        print(f"  ⚠️ No years detected in {stmt_type} - may need manual review")
                    
                    # Clean and structure the data
                    cleaned_data = self._clean_statement_data(stmt_data, stmt_type)
                    if cleaned_data is not None and not cleaned_data.empty:
                        # For cash flow statements, use raw data without enhancement
                        if stmt_type == 'cash_flow':
                            print(f"  🔍 Cash flow data (raw, no enhancement): {cleaned_data.shape[0]} rows, {cleaned_data.shape[1]} columns")
                            statements[stmt_type] = cleaned_data
                            print(f"  ✅ Raw cash flow data preserved: {cleaned_data.shape[0]} rows")
                        else:
                            statements[stmt_type] = cleaned_data
            
            return statements
            
        except Exception as e:
            print(f"  ❌ Error parsing {file_path}: {e}")
            return {'income_statement': None, 'balance_sheet': None, 'cash_flow': None}
    
    def _detect_years_in_statement_data(self, df: pd.DataFrame, stmt_type: str) -> List[str]:
        """Detect years within individual statement data - handles missing headers"""
        years = []
        
        # Look for years in the first 15 rows (where headers typically are)
        for row_idx in range(min(15, len(df))):
            for col_idx in range(min(15, len(df.columns))):
                cell_value = str(df.iloc[row_idx, col_idx]).strip()
                
                # Match various year formats
                year_patterns = [
                    r'20\d{2}',  # 2023, 2024, etc.
                    r'Dec\.?\s*31,?\s*(20\d{2})',  # Dec. 31, 2023
                    r'December\s*31,?\s*(20\d{2})',  # December 31, 2023
                    r'Year\s+Ended.*?(20\d{2})',  # Year Ended December 31, 2023
                    r'As\s+of.*?(20\d{2})',  # As of December 31, 2023
                    r'Fiscal\s+Year\s+(20\d{2})',  # Fiscal Year 2023
                ]
                
                for pattern in year_patterns:
                    matches = re.findall(pattern, cell_value, re.IGNORECASE)
                    for match in matches:
                        if isinstance(match, tuple):
                            year = match[0] if match[0] else match[1]
                        else:
                            year = match
                        
                        if year and 2015 <= int(year) <= 2030 and year not in years:
                            # Additional validation for balance sheet - avoid false positives
                            if stmt_type == 'balance_sheet':
                                # Check if this looks like a real year (not a large number)
                                if len(year) == 4 and int(year) <= 2030 and int(year) >= 2015:
                                    # Additional check: make sure it's not part of a larger number
                                    if len(cell_value.strip()) <= 6:  # Year should be standalone or with short context
                                        years.append(year)
                                        print(f"    📅 Found year {year} in {stmt_type} at row {row_idx}, col {col_idx}: '{cell_value[:30]}...'")
                            else:
                                years.append(year)
                                print(f"    📅 Found year {year} in {stmt_type} at row {row_idx}, col {col_idx}: '{cell_value[:30]}...'")
        
        # If no years found in headers, try to infer from data structure
        if not years and stmt_type == 'balance_sheet':
            years = self._infer_years_from_balance_sheet_structure(df)
        
        # For balance sheets, if still no years found, try to use the same years as other statements
        if not years and stmt_type == 'balance_sheet':
            # Look for years in the entire file context (not just this statement)
            print(f"    🔍 No years found in balance sheet, trying broader context...")
            # This will be handled by the calling function with file context
        
        return sorted(years, reverse=True) if years else []

    def _detect_years_from_sheet(self, df: pd.DataFrame) -> List[int]:
        """Scan the top 20 rows/15 cols for 4-digit years and return unique ints."""
        years = set()
        for r in range(min(20, len(df))):
            for c in range(min(15, len(df.columns))):
                try:
                    cell = str(df.iloc[r, c])
                except Exception:
                    continue
                m = re.findall(r"\b(20\d{2})\b", cell)
                for y in m:
                    yi = int(y)
                    if 2010 <= yi <= 2035:
                        years.add(yi)
        return sorted(years)
    
    def _infer_years_from_balance_sheet_structure(self, df: pd.DataFrame) -> List[str]:
        """Infer years for balance sheet when headers are missing"""
        years = []
        
        # First, try to find years in the broader context (look at more rows)
        for row_idx in range(min(15, len(df))):
            for col_idx in range(min(10, len(df.columns))):
                cell_value = str(df.iloc[row_idx, col_idx]).strip()
                
                # Look for year patterns that might be in balance sheet context
                year_patterns = [
                    r'20\d{2}',  # 2023, 2024, etc.
                    r'Dec\.?\s*31,?\s*(20\d{2})',  # Dec. 31, 2023
                    r'December\s*31,?\s*(20\d{2})',  # December 31, 2023
                    r'As\s+of.*?(20\d{2})',  # As of December 31, 2023
                    r'Fiscal\s+Year\s+(20\d{2})',  # Fiscal Year 2023
                ]
                
                for pattern in year_patterns:
                    matches = re.findall(pattern, cell_value, re.IGNORECASE)
                    for match in matches:
                        if isinstance(match, tuple):
                            year = match[0] if match[0] else match[1]
                        else:
                            year = match
                        
                        if year and 2015 <= int(year) <= 2030 and year not in years:
                            years.append(year)
                            print(f"    📅 Found year {year} in balance sheet context at row {row_idx}, col {col_idx}: '{cell_value[:30]}...'")
        
        # If still no years found, try to infer from data structure
        if not years:
            numeric_columns = []
            
            for col_idx in range(1, min(len(df.columns), 6)):  # Check first 5 data columns
                numeric_count = 0
                for row_idx in range(min(20, len(df))):  # Check first 20 rows
                    value = df.iloc[row_idx, col_idx]
                    if pd.notna(value) and isinstance(value, (int, float)) and value != 0:
                        numeric_count += 1
                
                if numeric_count >= 3:  # At least 3 numeric values suggests a year column
                    numeric_columns.append(col_idx)
            
            # If we found 2-3 numeric columns, assume they represent years
            if 2 <= len(numeric_columns) <= 3:
                # Use more realistic recent years based on current year
                current_year = datetime.now().year
                for i, col_idx in enumerate(numeric_columns):
                    year = str(current_year - i)
                    years.append(year)
                print(f"    📅 Inferred years for balance sheet: {years} (based on {len(numeric_columns)} data columns)")
        
        return years
    
    def _detect_statement_type_from_row(self, row_text: str) -> Optional[str]:
        """Detect statement type from row text - universal approach"""
        # Comprehensive income statement keywords
        income_keywords = [
            'income statement', 'income statements', 'statement of operations',
            'statement of earnings', 'statement of income', 'operations',
            'consolidated statements of operations', 'consolidated statement of operations',
            'consolidated statements of earnings', 'consolidated statement of earnings',
            'consolidated statements of income', 'consolidated statement of income',
            'consolidated statements of inco', 'consolidated statement of inco',
            'profit and loss', 'p&l', 'pnl'
        ]
        
        # Comprehensive balance sheet keywords
        balance_keywords = [
            'balance sheet', 'balance sheets', 'statement of financial position',
            'statement of financial condition', 'consolidated balance sheet',
            'consolidated balance sheets', 'consolidated statement of financial position',
            'consolidated balance', 'financial position', 'assets and liabilities'
        ]
        
        # Comprehensive cash flow keywords
        cash_flow_keywords = [
            'cash flow', 'cash flows', 'statement of cash flows',
            'statement of cash flow', 'consolidated statements of cash flows',
            'consolidated statement of cash flows', 'consolidated statements of cash flow',
            'consolidated statement of cash flow', 'consolidated statements of cash',
            'consolidated statement of cash', 'cash flows statement',
            'cash flow statement', 'operating activities', 'investing activities',
            'financing activities', 'net cash provided by operating',
            'net cash used in operating', 'net cash provided by investing',
            'net cash used in investing', 'net cash provided by financing',
            'net cash used in financing', 'cash and cash equivalents at beginning',
            'cash and cash equivalents at end', 'net increase in cash',
            'net decrease in cash', 'net change in cash'
        ]
        
        # Check for income statement
        if any(keyword in row_text for keyword in income_keywords):
            return 'income_statement'
        
        # Check for balance sheet
        if any(keyword in row_text for keyword in balance_keywords):
            return 'balance_sheet'
        
        # Check for cash flow statement
        if any(keyword in row_text for keyword in cash_flow_keywords):
            return 'cash_flow'
        
        return None
    
    def _find_statement_boundaries(self, df: pd.DataFrame) -> Dict[str, tuple]:
        """Find start and end rows for each statement type - universal approach"""
        boundaries = {
            'income_statement': (None, None),
            'balance_sheet': (None, None),
            'cash_flow': (None, None)
        }
        
        # Find all potential statement headers
        statement_headers = []
        for idx, row in df.iterrows():
            row_values = [str(val) if pd.notna(val) else "" for val in row.values]
            row_text = " ".join(row_values).lower()
            
            statement_type = self._detect_statement_type_from_row(row_text)
            if statement_type:
                statement_headers.append((idx, statement_type, row_text))
        
        # Find boundaries for each statement type
        for stmt_type in ['income_statement', 'balance_sheet', 'cash_flow']:
            # Find the header for this statement type
            header_row = None
            for idx, detected_type, text in statement_headers:
                if detected_type == stmt_type:
                    header_row = idx
                    break
            
            if header_row is not None:
                # Find the end of this statement
                end_row = self._find_statement_end(df, header_row, stmt_type)
                boundaries[stmt_type] = (header_row, end_row)
        
        return boundaries
    
    def _find_statement_end(self, df: pd.DataFrame, start_row: int, stmt_type: str) -> int:
        """Find the end row of a statement - capture complete tables"""
        # For cash flow and balance sheet statements, be VERY aggressive in capturing ALL line items
        if stmt_type in ['cash_flow', 'balance_sheet']:
            # For these statements, only stop at the very end of the file
            # or when we hit the NEXT major statement of a DIFFERENT type.
            major_statement_headers = [
                'consolidated statements of operations', 'consolidated statement of operations',
                'consolidated statements of earnings', 'consolidated statement of earnings',
                'consolidated statements of income', 'consolidated statement of income',
                'consolidated balance sheet', 'consolidated balance sheets',
                'consolidated statement of financial position',
                'consolidated statements of cash', 'consolidated statement of cash flows',
                'consolidated statements of cash flows'
            ]

            # Look for the next major statement header, but ignore the first header of the SAME type
            seen_same_type_header = False
            for idx in range(start_row + 1, len(df)):
                row_values = [str(val) if pd.notna(val) else "" for val in df.iloc[idx].values]
                row_text = " ".join(row_values).lower()

                if any(header in row_text for header in major_statement_headers):
                    detected_type = self._detect_statement_type_from_row(row_text)
                    # If it's the same type and we haven't seen it yet, it's likely the table header; continue
                    if detected_type == stmt_type and not seen_same_type_header:
                        seen_same_type_header = True
                        continue
                    # Stop only when we encounter a different statement type header
                    if detected_type and detected_type != stmt_type:
                        print(f"    🛑 Found major statement header at row {idx}: {row_text[:50]}...")
                        return idx - 1
        else:
            # For income statements, use more conservative logic
            for idx in range(start_row + 1, len(df)):
                row_values = [str(val) if pd.notna(val) else "" for val in df.iloc[idx].values]
                row_text = " ".join(row_values).lower()
                
                # Check if this is another statement header
                detected_type = self._detect_statement_type_from_row(row_text)
                if detected_type and detected_type != stmt_type:
                    return idx - 1
        
        # If no other statement found, take everything to the end
        print(f"    📊 {stmt_type} statement extends to end of file: row {len(df) - 1}")
        return len(df) - 1
    
    def _has_financial_data(self, row_values: list) -> bool:
        """Check if a row contains meaningful financial data"""
        # Look for numbers, currency symbols, or financial terms
        financial_indicators = ['$', 'revenue', 'income', 'expense', 'asset', 'liability', 'cash', 'equity']
        
        for val in row_values:
            val_str = str(val).lower()
            # Check for numbers
            if any(char.isdigit() for char in val_str):
                return True
            # Check for financial terms
            if any(indicator in val_str for indicator in financial_indicators):
                return True
        
        return False
    
    def _clean_statement_data(self, stmt_data: pd.DataFrame, stmt_type: str) -> pd.DataFrame:
        """Clean and structure statement data - keep raw tables as-is"""
        if stmt_data.empty:
            return None
        
        # For cash flow statements, keep EVERYTHING - only remove completely empty rows/columns
        if stmt_type == 'cash_flow':
            # Keep ALL data - only remove completely empty rows and columns
            cleaned_data = stmt_data.copy()
            cleaned_data = cleaned_data.dropna(how='all').dropna(axis=1, how='all')
            # NO OTHER FILTERING - preserve everything else
        else:
            # For other statements, use original logic
            cleaned_data = stmt_data.dropna(how='all').dropna(axis=1, how='all')
        
        if cleaned_data.empty:
            return None
        
        # Return the raw data without any processing - preserve all line items
        return cleaned_data.reset_index(drop=True)
    
    def _identify_data_structure(self, df: pd.DataFrame) -> str:
        """Identify the data structure of the statement"""
        # Check if first column contains mostly text (labels)
        first_col_text_ratio = sum(1 for val in df.iloc[:, 0] if isinstance(val, str) and not val.replace('.', '').replace('-', '').isdigit()) / len(df)
        
        # Check if first row contains mostly text (labels)
        first_row_text_ratio = sum(1 for val in df.iloc[0, :] if isinstance(val, str) and not str(val).replace('.', '').replace('-', '').isdigit()) / len(df.columns)
        
        if first_col_text_ratio > 0.7:
            return 'standard'
        elif first_row_text_ratio > 0.7:
            return 'transposed'
        else:
            return 'unknown'
    
    def _structure_standard_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """Structure data in standard format (labels in first column)"""
        # Keep the data as is, but ensure proper column structure
        return df.reset_index(drop=True)
    
    def _structure_transposed_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """Structure data in transposed format (labels in first row)"""
        # Transpose the data
        return df.T.reset_index(drop=True)
    
    def _structure_fallback_format(self, df: pd.DataFrame) -> pd.DataFrame:
        """Fallback structure for unknown formats"""
        # Try to find the best structure by analyzing the data
        # For now, return as is
        return df.reset_index(drop=True)
    
    def _create_horizontal_consolidated_excel(self, all_years_data: Dict, output_file: str, ticker: str):
        """Create horizontal consolidated Excel file"""
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = f"{ticker} Master Consolidated"
        
        current_row = 1
        
        # Process each statement type
        statement_types = [
            ('income_statement', 'Income Statements'),
            ('balance_sheet', 'Balance Sheets'),
            ('cash_flow', 'Cash Flow Statements')
        ]
        
        for statement_key, statement_title in statement_types:
            current_row = self._add_smart_statement_section(ws, statement_key, statement_title, all_years_data, current_row)
            current_row += 3  # Add small spacing between sections
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(output_file)
    
    def _add_horizontal_statement_section(self, ws, statement_key: str, statement_title: str, all_years_data: Dict, start_row: int):
        """Add a horizontal statement section to the worksheet - universal approach"""
        current_row = start_row
        
        # Add section title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = statement_title
        title_cell.font = Font(bold=True, size=16)
        current_row += 2
        
        # Get all years that have this statement
        years_with_data = []
        for year, statements in all_years_data.items():
            if statements[statement_key] is not None:
                years_with_data.append(year)
        
        if not years_with_data:
            return current_row
        
        # Sort years properly (handle both string and numeric years)
        try:
            years_with_data = sorted(years_with_data, key=lambda x: int(x) if x.isdigit() else 0, reverse=True)
        except:
            years_with_data = sorted(years_with_data, reverse=True)
        
        # Calculate dynamic column spacing based on data width
        max_cols_per_year = 0
        for year in years_with_data:
            df = all_years_data[year][statement_key]
            if df is not None and not df.empty:
                max_cols_per_year = max(max_cols_per_year, len(df.columns))
        
        # Use at least 8 columns per year, but adjust based on actual data
        cols_per_year = max(8, min(max_cols_per_year + 2, 15))
        
        # Create year headers
        col_start = 1
        for year in years_with_data:
            year_cell = ws.cell(row=current_row, column=col_start)
            year_cell.value = f"Year {year}"
            year_cell.font = Font(bold=True)
            
            # Merge cells for year header if we have multiple columns
            if cols_per_year > 1:
                ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_start + cols_per_year - 1)}{current_row}')
            
            col_start += cols_per_year
        
        current_row += 1
        
        # Add the actual data for each year
        max_rows = 0
        col_start = 1
        for year in years_with_data:
            df = all_years_data[year][statement_key]
            
            if df is not None and not df.empty:
                # Write data for this year
                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        if pd.notna(value) and c_idx < cols_per_year:
                            cell = ws.cell(row=current_row + r_idx, column=col_start + c_idx)
                            cell.value = value
                            
                            # Format numeric values in accounting format
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting millions (no decimals)
                                elif abs(value) >= 1000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting thousands (no decimals)
                                else:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Accounting small numbers (keep decimals)
                
                max_rows = max(max_rows, len(df))
            col_start += cols_per_year
        
        # Return the final row position after this statement
        return current_row + max_rows
    
    
    
    
    def _extract_line_items_from_dataframe(self, df: pd.DataFrame, year: str) -> Dict[str, any]:
        """Extract line items and values from a dataframe - handle multi-column structure"""
        line_items = {}
        
        # For each row, find the most recent/current year value
        for idx, row in df.iterrows():
            # Get line item name (usually in first column)
            line_item_name = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
            line_item_name = line_item_name.strip()
            
            # Skip only completely empty rows
            if not line_item_name or line_item_name.lower() in ['nan', 'none', '']:
                continue
            
            # Find the best value from all columns (prefer the first non-empty numeric value)
            best_value = None
            for col_idx in range(1, len(row)):
                value = row.iloc[col_idx] if col_idx < len(row) else None
                
                if pd.notna(value) and value != "":
                    try:
                        # Try to convert to number
                        if isinstance(value, str):
                            # Remove common formatting
                            clean_value = value.replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                            if clean_value.replace('.', '').replace('-', '').isdigit():
                                numeric_value = float(clean_value)
                                best_value = numeric_value
                                break
                        elif isinstance(value, (int, float)):
                            best_value = float(value)
                            break
                    except (ValueError, TypeError):
                        continue
            
            # Include EVERYTHING - no filtering for cash flow statements
            line_items[line_item_name] = best_value if best_value is not None else ""
        
        return line_items
    
    def _add_raw_cash_flow_section(self, ws, statement_title: str, all_years_data: Dict, years_with_data: list, start_row: int):
        """Add raw cash flow section preserving original structure"""
        current_row = start_row
        
        print(f"    🔍 Raw cash flow section - Years: {years_with_data}")
        print(f"    🔍 Available years in all_years_data: {list(all_years_data.keys())}")
        
        # Add section title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = statement_title
        title_cell.font = Font(bold=True, size=16)
        current_row += 2
        
        # Calculate dynamic column spacing based on data width
        max_cols_per_year = 0
        for year in years_with_data:
            df = all_years_data[year]['cash_flow']
            if df is not None and not df.empty:
                max_cols_per_year = max(max_cols_per_year, len(df.columns))
        
        # Use at least 8 columns per year, but adjust based on actual data
        cols_per_year = max(8, min(max_cols_per_year + 2, 15))
        
        # Create year headers
        col_start = 1
        for year in years_with_data:
            year_cell = ws.cell(row=current_row, column=col_start)
            year_cell.value = f"Year {year}"
            year_cell.font = Font(bold=True)
            
            # Merge cells for year header if we have multiple columns
            if cols_per_year > 1:
                ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_start + cols_per_year - 1)}{current_row}')
            
            col_start += cols_per_year
        
        current_row += 1
        
        # Add the actual data for each year - preserve raw structure
        max_rows = 0
        col_start = 1
        for year in years_with_data:
            df = all_years_data[year]['cash_flow']
            
            print(f"    🔍 Processing year {year}:")
            print(f"      - Data type: {type(df)}")
            print(f"      - Is None: {df is None}")
            print(f"      - Is empty: {df.empty if df is not None else 'N/A'}")
            if df is not None and not df.empty:
                print(f"      - Shape: {df.shape[0]} rows, {df.shape[1]} columns")
                print(f"      - First few rows:")
                for i in range(min(3, len(df))):
                    print(f"        Row {i}: {df.iloc[i].values[:3]}...")
            
            if df is not None and not df.empty:
                print(f"    📊 Writing raw cash flow data for {year}: {df.shape[0]} rows, {df.shape[1]} columns")
                
                # Write data for this year - preserve all rows and structure
                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        if pd.notna(value) and c_idx < cols_per_year:
                            cell = ws.cell(row=current_row + r_idx, column=col_start + c_idx)
                            cell.value = value
                            
                            # Format numeric values in accounting format
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting millions (no decimals)
                                elif abs(value) >= 1000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting thousands (no decimals)
                                else:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Accounting small numbers (keep decimals)
                
                max_rows = max(max_rows, len(df))
            col_start += cols_per_year
        
        print(f"    ✅ Raw cash flow section completed: {max_rows} rows written")
        
        # If no data was written, try a completely raw approach
        if max_rows == 0:
            print(f"    ⚠️ No data written, trying completely raw approach...")
            return self._add_completely_raw_cash_flow_section(ws, statement_title, all_years_data, years_with_data, current_row)
        
        return current_row + max_rows
    
    def _add_completely_raw_cash_flow_section(self, ws, statement_title: str, all_years_data: Dict, years_with_data: list, start_row: int):
        """Completely raw cash flow section - bypass all filtering and preserve complete structure"""
        current_row = start_row
        
        print(f"    🔄 Using completely raw cash flow approach - preserving ALL data")
        
        # Add section title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = statement_title
        title_cell.font = Font(bold=True, size=16)
        current_row += 2
        
        # Calculate dynamic column spacing based on actual data width
        max_cols_per_year = 0
        for year in years_with_data:
            if 'cash_flow' in all_years_data[year] and all_years_data[year]['cash_flow'] is not None:
                stmt_info = all_years_data[year]['cash_flow']
                if isinstance(stmt_info, dict) and 'data' in stmt_info:
                    df = stmt_info['data']
                else:
                    df = stmt_info  # Backward compatibility
                max_cols_per_year = max(max_cols_per_year, len(df.columns))
        
        # Use at least 10 columns per year, but adjust based on actual data
        cols_per_year = max(10, min(max_cols_per_year + 2, 20))
        
        # Create year headers with proper spacing
        col_start = 1
        for year in years_with_data:
            year_cell = ws.cell(row=current_row, column=col_start)
            year_cell.value = f"Year {year}"
            year_cell.font = Font(bold=True)
            
            # Merge cells for year header if we have multiple columns
            if cols_per_year > 1:
                ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_start + cols_per_year - 1)}{current_row}')
            
            col_start += cols_per_year
        
        current_row += 1
        
        # Write ALL data from each year without any filtering - preserve complete structure
        max_rows = 0
        col_start = 1
        for year in years_with_data:
            if 'cash_flow' in all_years_data[year] and all_years_data[year]['cash_flow'] is not None:
                stmt_info = all_years_data[year]['cash_flow']
                if isinstance(stmt_info, dict) and 'data' in stmt_info:
                    df = stmt_info['data']
                    years_in_data = stmt_info.get('years', [])
                else:
                    df = stmt_info  # Backward compatibility
                    years_in_data = []
                
                print(f"    📊 Completely raw approach for {year}: {df.shape[0]} rows, {df.shape[1]} columns, years: {years_in_data}")
                print(f"    📋 Sample data for {year}:")
                for i in range(min(3, len(df))):
                    print(f"      Row {i}: {df.iloc[i].values[:3]}...")
                
                # Write EVERYTHING - no filtering at all, preserve all rows and columns
                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        if c_idx < cols_per_year:  # Use dynamic column limit
                            cell = ws.cell(row=current_row + r_idx, column=col_start + c_idx)
                            cell.value = value
                            
                            # Format numeric values
                            if isinstance(value, (int, float)) and not pd.isna(value):
                                if abs(value) >= 1000000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
                                elif abs(value) >= 1000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
                                else:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                
                max_rows = max(max_rows, len(df))
            col_start += cols_per_year
        
        print(f"    ✅ Completely raw cash flow section completed: {max_rows} rows written")
        return current_row + max_rows
    
    def test_cash_flow_extraction(self, file_path: str):
        """Test function to debug cash flow extraction from individual files"""
        print(f"\n🧪 Testing cash flow extraction from: {os.path.basename(file_path)}")
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return
        
        try:
            # Read the Excel file
            df = pd.read_excel(file_path, sheet_name=0, header=None)
            print(f"📊 Raw file data: {df.shape[0]} rows, {df.shape[1]} columns")
            
            # Find statement boundaries
            statement_boundaries = self._find_statement_boundaries(df)
            print(f"🔍 Statement boundaries found: {statement_boundaries}")
            
            # Check cash flow specifically
            if statement_boundaries['cash_flow'][0] is not None:
                start_row, end_row = statement_boundaries['cash_flow']
                print(f"📋 Cash flow boundaries: rows {start_row} to {end_row}")
                
                # Extract cash flow data
                stmt_data = df.iloc[start_row:end_row+1].copy()
                print(f"📊 Cash flow raw data: {stmt_data.shape[0]} rows, {stmt_data.shape[1]} columns")
                
                # Show first few rows
                print(f"📋 First 10 rows of cash flow data:")
                for i in range(min(10, len(stmt_data))):
                    row_values = [str(val) if pd.notna(val) else "" for val in stmt_data.iloc[i].values]
                    print(f"  Row {i}: {row_values[:5]}...")
                
                # Clean the data
                cleaned_data = self._clean_statement_data(stmt_data, 'cash_flow')
                if cleaned_data is not None and not cleaned_data.empty:
                    print(f"✅ Cleaned cash flow data: {cleaned_data.shape[0]} rows, {cleaned_data.shape[1]} columns")
                    
                    # Show cleaned data
                    print(f"📋 First 10 rows of cleaned cash flow data:")
                    for i in range(min(10, len(cleaned_data))):
                        row_values = [str(val) if pd.notna(val) else "" for val in cleaned_data.iloc[i].values]
                        print(f"  Row {i}: {row_values[:5]}...")
                else:
                    print(f"❌ Cash flow data became empty after cleaning")
            else:
                print(f"❌ No cash flow statement boundaries found")
                
        except Exception as e:
            print(f"❌ Error testing cash flow extraction: {e}")
    
    def _find_best_data_column(self, df: pd.DataFrame) -> int:
        """Find the column with the most complete financial data"""
        if df.empty:
            return None
        
        best_col = None
        max_financial_values = 0
        
        # Check each column (skip first column which usually has line item names)
        for col_idx in range(1, min(len(df.columns), 10)):  # Check up to 10 columns
            financial_count = 0
            
            for idx, row in df.iterrows():
                if col_idx < len(row):
                    value = row.iloc[col_idx]
                    if pd.notna(value) and value != "":
                        try:
                            # Check if it's a number
                            if isinstance(value, (int, float)):
                                if abs(value) > 0.01:
                                    financial_count += 1
                            elif isinstance(value, str):
                                clean_value = value.replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                                if clean_value.replace('.', '').replace('-', '').isdigit():
                                    if abs(float(clean_value)) > 0.01:
                                        financial_count += 1
                        except (ValueError, TypeError):
                            continue
            
            if financial_count > max_financial_values:
                max_financial_values = financial_count
                best_col = col_idx
        
        return best_col
    
    def _create_detailed_consolidated_excel(self, all_years_data: Dict, output_file: str, ticker: str):
        """Create detailed consolidated Excel file with all original data, one sheet per statement"""
        workbook = openpyxl.Workbook()

        # Define sheet names for each statement
        statement_types = [
            ('income_statement', 'Income Statements'),
            ('balance_sheet', 'Balance Sheets'),
            ('cash_flow', 'Cash Flow Statements')
        ]

        # Create or reuse sheets per statement
        for idx, (statement_key, statement_title) in enumerate(statement_types):
            if idx == 0:
                ws = workbook.active
            else:
                ws = workbook.create_sheet()

            ws.title = statement_title

            # Start at top of the sheet
            current_row = 1

            # For cash flow statements, use completely raw approach
            if statement_key == 'cash_flow':
                print(f"    🔄 Using completely raw cash flow approach for detailed file - NO FILTERING")
                years_with_data = [year for year, statements in all_years_data.items() if statements.get(statement_key) is not None]
                current_row = self._add_completely_raw_cash_flow_section(ws, statement_title, all_years_data, years_with_data, current_row)
            else:
                current_row = self._add_horizontal_statement_section(ws, statement_key, statement_title, all_years_data, current_row)

            # Auto-adjust column widths for this sheet
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save after all sheets are prepared
        workbook.save(output_file)
    


    def _remove_empty_rows_and_columns(self, detailed_file: str, ticker: str, form_type: str) -> Optional[str]:
        """Create a copy of the detailed consolidated workbook with empty rows and columns removed per sheet."""
        try:
            print(f"\n🧹 Removing empty rows/columns from detailed file...")
            if not os.path.exists(detailed_file):
                print(f"  ❌ Detailed file not found: {detailed_file}")
                return None

            # Output filename
            if detailed_file.endswith('_Detailed.xlsx'):
                output_file = detailed_file.replace('_Detailed.xlsx', '_RemovedEmptyRowsCols.xlsx')
            else:
                base, ext = os.path.splitext(detailed_file)
                output_file = f"{base}_RemovedEmptyRowsCols{ext}"

            # Read all sheets, drop empty rows/cols, write to new workbook
            xls = pd.ExcelFile(detailed_file)
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet_name in xls.sheet_names:
                    try:
                        df = pd.read_excel(detailed_file, sheet_name=sheet_name, header=None)
                    except Exception as e:
                        print(f"  ⚠️ Skipping sheet '{sheet_name}' (read error): {e}")
                        continue

                    if df is None:
                        continue

                    # Drop completely empty columns then rows
                    cleaned = df.dropna(axis=1, how='all')
                    cleaned = cleaned.dropna(how='all')

                    # If everything got dropped, write an empty sheet to preserve structure
                    if cleaned is None or cleaned.empty:
                        cleaned = pd.DataFrame()

                    cleaned.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            print(f"  ✅ Empty rows/columns removed: {os.path.basename(output_file)}")
            return output_file
        except Exception as e:
            print(f"  ❌ Error removing empty rows/columns: {e}")
            return None

    def _detect_row_wise_pattern(self, row_values: list) -> list:
        """Detect and parse row-wise pattern: MetricName + 2-3+ values (greedy capture).
        - Metric name must contain at least one alphabetic character
        - After first 2-3 numeric values, greedily consume additional numeric-only groups
        until a new plausible metric name appears
        """
        blocks = []
        i = 0
        
        while i < len(row_values):
            # Skip empty/NaN values
            while i < len(row_values) and (pd.isna(row_values[i]) or str(row_values[i]).strip() == ''):
                i += 1
            
            if i >= len(row_values):
                break
                
            # Get metric name (first non-empty value)
            metric_name = str(row_values[i]).strip()
            # Require at least one alphabetic character in the metric name
            if (not metric_name or metric_name.lower() in ['nan', 'none'] or
                not any(ch.isalpha() for ch in metric_name)):
                i += 1
                continue
                
            # Look for 2-3 numeric values after the metric name
            values = []
            j = i + 1
            
            while j < len(row_values) and len(values) < 3:
                val = row_values[j]
                if pd.notna(val) and str(val).strip() != '':
                    try:
                        # Try to convert to number
                        if isinstance(val, str):
                            # Clean the string value
                            clean_val = val.replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                            if clean_val.replace('.', '').replace('-', '').isdigit():
                                numeric_val = float(clean_val)
                                values.append(numeric_val)
                            else:
                                break  # Non-numeric value, end of block
                        elif isinstance(val, (int, float)):
                            values.append(float(val))
                        else:
                            break  # Non-numeric value, end of block
                    except (ValueError, TypeError):
                        break  # Can't convert to number, end of block
                else:
                    break  # Empty value, end of block
                j += 1
            
            # Only add block if we have at least 2 values
            if len(values) >= 2:
                # Greedily consume additional numeric-only groups as continuation of same metric
                k = j
                while k < len(row_values):
                    nxt = row_values[k]
                    if pd.isna(nxt) or str(nxt).strip() == '':
                        break
                    # Stop if next token looks like a new metric name (has letters)
                    if isinstance(nxt, str) and any(ch.isalpha() for ch in nxt.strip()):
                        break
                    # Otherwise parse 1 more numeric and append if valid
                    try:
                        if isinstance(nxt, str):
                            clean_val = nxt.replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                            if clean_val.replace('.', '').replace('-', '').isdigit():
                                values.append(float(clean_val))
                            else:
                                break
                        elif isinstance(nxt, (int, float)):
                            values.append(float(nxt))
                        else:
                            break
                    except Exception:
                        break
                    k += 1

                blocks.append({
                    'metric_name': metric_name,
                    'values': values,
                    'start_idx': i,
                    'end_idx': max(j - 1, k - 1)
                })
            
            i = max(j, k) if 'k' in locals() else j
        
        return blocks

    def _calculate_value_overlap_percentage(self, values1: list, values2: list) -> float:
        """Calculate percentage overlap between two value lists"""
        if not values1 or not values2:
            return 0.0
        
        # Check if last 2 values of values1 match first 2 values of values2
        if len(values1) >= 2 and len(values2) >= 2:
            last_two = values1[-2:]
            first_two = values2[:2]
            
            # Calculate percentage difference for each pair
            total_diff = 0
            for v1, v2 in zip(last_two, first_two):
                if v1 != 0:  # Avoid division by zero
                    diff = abs(v1 - v2) / abs(v1) * 100
                    total_diff += diff
                elif v2 != 0:
                    diff = abs(v1 - v2) / abs(v2) * 100
                    total_diff += diff
                else:
                    total_diff += 0  # Both are zero, perfect match
            
            # Return average percentage difference (lower = better match)
            avg_diff = total_diff / 2
            return 100 - avg_diff  # Convert to overlap percentage
        
        return 0.0

    def _consolidate_row_wise_metrics(self, blocks: list, overlap_threshold: float = 95.0) -> dict:
        """Consolidate overlapping metrics from row-wise pattern"""
        if not blocks:
            return {}
        
        # Process blocks sequentially, merging with previous ones
        consolidated = {}
        
        for i, block in enumerate(blocks):
            metric_name = block['metric_name']
            values = block['values']
            
            # Check if this block can be merged with any existing consolidated metric
            merged = False
            
            for existing_name, existing_data in consolidated.items():
                overlap_pct = self._calculate_value_overlap_percentage(existing_data['values'], values)
                
                if overlap_pct >= overlap_threshold:
                    # Merge the metrics - use latest name and combine values
                    print(f"    🔗 Merging '{existing_name}' with '{metric_name}' (overlap: {overlap_pct:.1f}%)")
                    
                    # Remove overlapping values and add new ones; restatement-aware
                    existing_values = existing_data['values']
                    new_values = values.copy()
                    pct_threshold = 2.0
                    abs_threshold = 5.0
                    if len(existing_values) >= 2 and len(new_values) >= 2:
                        old_v1, new_v1 = existing_values[-2], new_values[0]
                        old_v2, new_v2 = existing_values[-1], new_values[1]
                        def _maybe_restate(oldv, newv, idx):
                            try:
                                if oldv is None or newv is None:
                                    return False
                                if oldv == 0 and newv == 0:
                                    return False
                                abs_delta = abs(newv - oldv)
                                base = abs(oldv) if abs(oldv) > 0 else abs(newv)
                                pct_delta = (abs_delta / base) * 100.0 if base > 0 else 0.0
                                if abs_delta <= abs_threshold or pct_delta <= pct_threshold:
                                    existing_data.setdefault('restatements', {}).setdefault(idx, []).append({
                                        'old': oldv, 'new': newv, 'abs_delta': abs_delta, 'pct_delta': pct_delta
                                    })
                                    return True
                                return False
                            except Exception:
                                return False
                        if _maybe_restate(old_v1, new_v1, len(existing_values)-2):
                            existing_values[-2] = new_v1
                        if _maybe_restate(old_v2, new_v2, len(existing_values)-1):
                            existing_values[-1] = new_v2
                        if (abs(existing_values[-2] - new_values[0]) < 0.01 and 
                            abs(existing_values[-1] - new_values[1]) < 0.01):
                            new_values = new_values[2:]
                    
                    # Combine values
                    combined_values = existing_values + new_values
                    
                    # Update the existing entry with combined values
                    # Keep the first-seen (latest) metric name; do NOT overwrite it with older names
                    consolidated[existing_name]['values'] = combined_values
                    consolidated[existing_name]['original_names'].append(metric_name)
                    # carry forward restatements if any
                    if 'restatements' in existing_data:
                        consolidated[existing_name]['restatements'] = existing_data['restatements']
                    
                    merged = True
                    break
            
            if not merged:
                # Add as new metric
                consolidated[metric_name] = {
                    'values': values,
                    'original_names': [metric_name],
                    # Mark the first-seen name; this corresponds to the most recent year block
                    'latest_name': name if 'name' in locals() else metric_name
                }
        
        # Create final result with latest names
        final_result = {}
        for name, data in consolidated.items():
            # Use the first-seen metric name (latest period's name)
            final_name = data.get('latest_name', name)
            final_result[final_name] = {
                'values': data['values'],
                'original_names': data['original_names'],
                'restatements': data.get('restatements', {})
            }
        
        return final_result

    def _format_restated_values(self, values: list) -> list:
        """Format values to show restatements as prev_value{restated_value}"""
        if len(values) < 2:
            return values
        
        formatted_values = []
        
        for i, value in enumerate(values):
            if i == 0:
                formatted_values.append(value)
            else:
                # Check if this value represents a restatement
                # For now, we'll use a simple heuristic: if values are very close, treat as restatement
                prev_value = values[i-1]
                if abs(value - prev_value) < 0.01:  # Values are essentially the same
                    formatted_values.append(value)
                else:
                    # Check if this might be a restatement by looking at context
                    # This is a simplified approach - you might want to enhance this logic
                    formatted_values.append(value)
        
        return formatted_values

    

    

    # ========= Embeddings-based consolidation =========
    def _get_text_encoder(self):
        """Try to load a sentence embedding model; fall back to TF-IDF. Returns a callable encode(list[str])->np.ndarray or None."""
        try:
            from sentence_transformers import SentenceTransformer
            model = SentenceTransformer('all-MiniLM-L6-v2')
            return lambda texts: model.encode(texts, normalize_embeddings=True)
        except Exception:
            try:
                # Lightweight fallback: TF-IDF + l2 normalization approximates cosine
                from sklearn.feature_extraction.text import TfidfVectorizer
                import numpy as np
                vectorizer = TfidfVectorizer(lowercase=True, ngram_range=(1, 2))
                # Fit lazily on first call
                def encode(texts):
                    X = vectorizer.fit_transform(texts)
                    # l2 normalize
                    norms = np.sqrt((X.power(2)).sum(axis=1))
                    norms[norms == 0] = 1.0
                    return (X.multiply(1.0 / norms)).toarray()
                return encode
            except Exception:
                return None

    def _cosine_similarity_matrix(self, A, B):
        """Compute cosine similarity matrix between two arrays; accepts 1D or 2D."""
        import numpy as np
        A = np.asarray(A); B = np.asarray(B)
        if A.ndim == 1: A = A[None, :]
        if B.ndim == 1: B = B[None, :]
        # Ensure feature dims match
        if A.shape[1] != B.shape[1]:
            # Try to flatten if accidentally shaped (1,) due to scalar
            A = A.reshape(A.shape[0], -1)
            B = B.reshape(B.shape[0], -1)
        # Safe norms
        An = np.linalg.norm(A, axis=1, keepdims=True); An[An==0] = 1.0
        Bn = np.linalg.norm(B, axis=1, keepdims=True); Bn[Bn==0] = 1.0
        return (A/An) @ (B.T/Bn.T)

    def _consolidate_row_wise_metrics_embeddings(self, blocks: list, encode) -> dict:
        """Consolidate blocks using name embeddings + numeric overlap.
        - Keep the first block's name as canonical (latest wording)
        - Require either strong name similarity or numeric-overlap continuity
        """
        if not blocks:
            return {}
        import numpy as np

        names = [b['metric_name'] for b in blocks]
        embs = encode(names) if encode is not None else None

        consolidated = {}
        for idx, block in enumerate(blocks):
            metric_name = block['metric_name']
            values = block['values']

            merged = False
            for existing_name, data in consolidated.items():
                # Numeric continuity score (based on last-2 to first-2)
                overlap_pct = self._calculate_value_overlap_percentage(data['values'], values)

                # Name similarity score
                name_sim = 0.0
                if embs is not None:
                    i_existing = data['first_index']
                    sim_mat = self._cosine_similarity_matrix(embs[i_existing], embs[idx])
                    sim = float(sim_mat.squeeze())
                    name_sim = (max(min(sim, 1.0), -1.0) + 1.0) / 2.0 * 100.0

                # Merge if either metric names are semantically close OR numeric overlap is strong
                if overlap_pct >= 90.0 or name_sim >= 85.0:
                    # Append new tail, removing duplicate overlap
                    existing_values = data['values']
                    new_values = values.copy()
                    if len(existing_values) >= 2 and len(new_values) >= 2:
                        if (abs(existing_values[-2] - new_values[0]) < 0.01 and
                            abs(existing_values[-1] - new_values[1]) < 0.01):
                            new_values = new_values[2:]
                    consolidated[existing_name]['values'] = existing_values + new_values
                    consolidated[existing_name]['original_names'].append(metric_name)
                    merged = True
                    break

            if not merged:
                # Start a new chain; store first index to anchor latest wording
                consolidated[metric_name] = {
                    'values': values,
                    'original_names': [metric_name],
                    'first_index': idx
                }

        # Ensure we keep the first (latest) name as canonical; values are newest-first from blocks order
        final_result = {}
        for name, data in consolidated.items():
            final_result[name] = {
                'values': data['values'],
                'original_names': data['original_names']
            }
        return final_result


    def _add_horizontal_statement_section(self, ws, statement_key: str, statement_title: str, all_years_data: Dict, start_row: int):
        """Add a horizontal statement section to the worksheet - handles new data structure"""
        current_row = start_row
        
        # Add section title
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.value = statement_title
        title_cell.font = Font(bold=True, size=16)
        current_row += 2
        
        # Get all years that have this statement
        years_with_data = []
        for year, statements in all_years_data.items():
            if statements[statement_key] is not None:
                years_with_data.append(year)
        
        if not years_with_data:
            return current_row
        
        # Sort years properly (handle both string and numeric years)
        try:
            years_with_data = sorted(years_with_data, key=lambda x: int(x) if x.isdigit() else 0, reverse=True)
        except:
            years_with_data = sorted(years_with_data, reverse=True)
        
        # Calculate dynamic column spacing based on data width
        max_cols_per_year = 0
        for year in years_with_data:
            stmt_info = all_years_data[year][statement_key]
            if stmt_info is not None:
                df = stmt_info['data']
                max_cols_per_year = max(max_cols_per_year, len(df.columns))
        
        # Use at least 8 columns per year, but adjust based on actual data
        cols_per_year = max(8, min(max_cols_per_year + 2, 15))
        
        # Create year headers
        col_start = 1
        for year in years_with_data:
            year_cell = ws.cell(row=current_row, column=col_start)
            year_cell.value = f"Year {year}"
            year_cell.font = Font(bold=True)
            
            # Merge cells for year header if we have multiple columns
            if cols_per_year > 1:
                ws.merge_cells(f'{get_column_letter(col_start)}{current_row}:{get_column_letter(col_start + cols_per_year - 1)}{current_row}')
            
            col_start += cols_per_year
        
        current_row += 1
        
        # Add the actual data for each year
        max_rows = 0
        col_start = 1
        for year in years_with_data:
            stmt_info = all_years_data[year][statement_key]
            
            if stmt_info is not None:
                df = stmt_info['data']
                years_in_data = stmt_info['years']
                
                # Write data for this year
                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        if pd.notna(value) and c_idx < cols_per_year:
                            cell = ws.cell(row=current_row + r_idx, column=col_start + c_idx)
                            cell.value = value
                            
                            # Format numeric values in accounting format
                            if isinstance(value, (int, float)):
                                if abs(value) >= 1000000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting millions (no decimals)
                                elif abs(value) >= 1000:
                                    cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'  # Accounting thousands (no decimals)
                                else:
                                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'  # Accounting small numbers (keep decimals)
                
                max_rows = max(max_rows, len(df))
            col_start += cols_per_year
        
        # Return the final row position after this statement
        return current_row + max_rows
    
    
    def create_consolidated_file(self, file_year_mapping: List[tuple], ticker: str, form_type: str, output_dir: str):
        """Create detailed consolidated file"""
        try:
            print(f"\n🔄 Creating consolidated files for {ticker}...")
            
            # Parse all individual files with their years
            all_years_data = {}
            
            for file_path, year in file_year_mapping:
                if not os.path.exists(file_path):
                    print(f"  ❌ File not found: {file_path}")
                    continue
                
                print(f"  🔍 Parsing individual file: {os.path.basename(file_path)} for year {year}")
                
                # Parse the individual file
                statements = self._parse_individual_financial_file(file_path)
                if statements:
                    # For each statement, try to detect years within the data
                    enhanced_statements = {}
                    detected_years_from_other_statements = []
                    
                    # First pass: detect years from all statements
                    for stmt_type, stmt_data in statements.items():
                        if stmt_data is not None and not stmt_data.empty:
                            years_in_statement = self._detect_years_in_statement_data(stmt_data, stmt_type)
                            if years_in_statement:
                                detected_years_from_other_statements.extend(years_in_statement)
                    
                    # Remove duplicates and sort
                    detected_years_from_other_statements = sorted(list(set(detected_years_from_other_statements)), reverse=True)
                    
                    # Second pass: assign years to each statement
                    for stmt_type, stmt_data in statements.items():
                        if stmt_data is not None and not stmt_data.empty:
                            # Detect years within this statement's data
                            years_in_statement = self._detect_years_in_statement_data(stmt_data, stmt_type)
                            
                            # For balance sheets, be more aggressive about using years from other statements
                            if stmt_type == 'balance_sheet':
                                if years_in_statement and len(years_in_statement) >= 2:
                                    # Use detected years if we found at least 2
                                    print(f"    📅 Years in {stmt_type}: {years_in_statement}")
                                    enhanced_statements[stmt_type] = {
                                        'data': stmt_data,
                                        'years': years_in_statement
                                    }
                                elif detected_years_from_other_statements:
                                    # Use years from other statements for balance sheet
                                    print(f"    📅 Using years from other statements for balance sheet: {detected_years_from_other_statements}")
                                    enhanced_statements[stmt_type] = {
                                        'data': stmt_data,
                                        'years': detected_years_from_other_statements
                                    }
                                else:
                                    # Use the file year as fallback
                                    enhanced_statements[stmt_type] = {
                                        'data': stmt_data,
                                        'years': [year] if year != "Unknown" else []
                                    }
                            else:
                                # For other statements, use normal logic
                                if years_in_statement:
                                    print(f"    📅 Years in {stmt_type}: {years_in_statement}")
                                    enhanced_statements[stmt_type] = {
                                        'data': stmt_data,
                                        'years': years_in_statement
                                    }
                                else:
                                    # Use the file year as fallback
                                    enhanced_statements[stmt_type] = {
                                        'data': stmt_data,
                                        'years': [year] if year != "Unknown" else []
                                    }
                        else:
                            enhanced_statements[stmt_type] = None
                    
                    all_years_data[year] = enhanced_statements
                    found_statements = [k for k, v in enhanced_statements.items() if v is not None]
                    print(f"  ✅ Parsed {year}: {len(found_statements)} statements - {found_statements}")
                    
                    # Debug cash flow data specifically
                    if 'cash_flow' in enhanced_statements and enhanced_statements['cash_flow'] is not None:
                        cf_data = enhanced_statements['cash_flow']['data']
                        cf_years = enhanced_statements['cash_flow']['years']
                        print(f"    📊 Cash flow data for {year}: {cf_data.shape[0]} rows, {cf_data.shape[1]} columns, years: {cf_years}")
                        print(f"    📋 Cash flow sample rows:")
                        for i in range(min(5, len(cf_data))):
                            print(f"      Row {i}: {cf_data.iloc[i].values[:3]}...")
                    else:
                        print(f"    ❌ No cash flow data found for {year}")
                else:
                    print(f"  ⚠️ No statements found for {year}")
            
            if not all_years_data:
                print("  ❌ No valid data found for consolidation")
                return None, None, None
            
            # Debug: Check what data we have before consolidation
            print(f"\n🔍 Data summary before consolidation:")
            for year, statements in all_years_data.items():
                print(f"  Year {year}:")
                for stmt_type, stmt_info in statements.items():
                    if stmt_info is not None:
                        df = stmt_info['data']
                        years = stmt_info['years']
                        print(f"    {stmt_type}: {df.shape[0]} rows, {df.shape[1]} columns, years: {years}")
                    else:
                        print(f"    {stmt_type}: No data")
            
            # Create detailed consolidated file (original format with all data)
            detailed_file = os.path.join(output_dir, f"{ticker}_{form_type}_Master_Consolidated_Financials_Detailed.xlsx")
            print(f"\n🔄 Creating detailed consolidated file...")
            self._create_detailed_consolidated_excel(all_years_data, detailed_file, ticker)
            print(f"  ✅ Created detailed consolidated file: {os.path.basename(detailed_file)}")
            
            # STEP 1: Remove empty rows and columns from each sheet and save a new workbook
            no_empty_file = self._remove_empty_rows_and_columns(detailed_file, ticker, form_type)
            
            # STEP 2: (Removed) Row-wise cleaning stage was deprecated per user request
            row_wise_cleaned_file = None

            # STEP 3: Apply semantic deduplication on the 'RemovedEmptyRowsCols' file (if present)
            dedup_file = None
            target_for_dedup = no_empty_file or detailed_file
            if target_for_dedup:
                try:
                    dedup_file = self.deduplicate_excel_file(target_for_dedup, None, verbose=False)
                except Exception as e:
                    print(f"  ⚠️ Deduplication skipped due to error: {e}")

            return detailed_file, no_empty_file, row_wise_cleaned_file, dedup_file
            
        except Exception as e:
            print(f"  ❌ Error creating consolidated files: {e}")
            return None, None, None, None




    def automated_download(self, ticker: str, form_type: str = "10-K", start_date: Optional[str] = None, end_date: Optional[str] = None):
        """Fully automated download process (range-only)."""
        return self.automated_download_with_range(ticker, form_type, start_date=start_date, end_date=end_date)

    def automated_download_with_range(self, ticker: str, form_type: str = "10-K", start_date: Optional[str] = None, end_date: Optional[str] = None, download_both: bool = False):
        """Fully automated download process supporting either recent count or date range.
        - download_both: If True, downloads both 10-K and 10-Q filings
        """
        if download_both:
            print(f"🤖 Starting automated download for {ticker} (both 10-K and 10-Q)...")
            # Download both 10-K and 10-Q
            files_10k = self._download_single_form_type(ticker, "10-K", start_date, end_date)
            files_10q = self._download_single_form_type(ticker, "10-Q", start_date, end_date)
            
            # Combine results
            all_files = (files_10k or []) + (files_10q or [])
            print(f"📊 Total files downloaded: {len(all_files)}")
            return all_files
        else:
            return self._download_single_form_type(ticker, form_type, start_date, end_date)
    
    def _download_single_form_type(self, ticker: str, form_type: str, start_date: Optional[str], end_date: Optional[str]) -> List[str]:
        """Download filings for a single form type."""
        print(f"🤖 Starting automated download for {ticker} {form_type}...")
        
        # Search for filings
        filings = self.search_filings(ticker, form_type, start_date=start_date, end_date=end_date)
        if not filings:
            print(f"❌ No {form_type} filings found")
            return []
        
        # Create output directory
        output_dir = f"SEC_Excel_Downloads_{ticker}_{form_type}"
        os.makedirs(output_dir, exist_ok=True)
        print(f"📁 Created directory: {output_dir}")
        
        # Generate Excel URLs (initial guess; final used URL captured at download time)
        excel_info = []
        for i, filing in enumerate(filings, 1):
            accession_no = filing.get("accessionNo")
            cik = filing.get("cik")
            filed_at = filing.get("filedAt", "Unknown")
            company_name = filing.get("companyName", "Unknown")
            form_actual = filing.get("formType") or form_type
            
            if not accession_no or not cik:
                continue
            
            clean_accession = accession_no.replace("-", "")
            excel_url = f"{self.sec_base_url}/Archives/edgar/data/{cik}/{clean_accession}/Financial_Report.xlsx"
            # Simple human-friendly label
            simple_label = self._compute_simple_label(form_actual, filing)
            label = self._derive_filing_label(form_actual, filing)
            
            excel_info.append({
                "filing_number": i,
                "company_name": company_name,
                "filed_at": filed_at,
                "accession_no": accession_no,
                "cik": cik,
                "excel_url": excel_url,
                "label": label,
                "form_type": form_actual,
            })
        
        # Download files
        downloaded_files = []
        for info in excel_info:
            # Prefer simple filename like 2024-10K.xlsx or 2023-Q2.xlsx
            simple_label = self._compute_simple_label(info.get('form_type') or form_type, info.get('filed_at') or info.get('filedAt') or info)
            filename = f"{simple_label}.xlsx"
            
            # 8-K support removed - only 10-K and 10-Q are supported

            file_path = self.download_excel_file(
                info['accession_no'],
                info['cik'],
                output_dir,
                filename
            )
            
            if file_path:
                downloaded_files.append(file_path)
                # Update excel_info with the actual URL that worked (if captured)
                try:
                    if self.last_used_excel_url:
                        info['excel_url'] = self.last_used_excel_url
                except Exception:
                    pass
            
            # Add delay between downloads
            time.sleep(random.uniform(2, 5))
        
        # Create master consolidated file (detailed only)
        if downloaded_files:
            # Create mapping of files to their years
            file_year_mapping = []
            for i, info in enumerate(excel_info):
                simple_label = self._compute_simple_label(info.get('form_type') or form_type, info.get('filed_at') or info.get('filedAt') or info)
                filename = f"{simple_label}_Individual_Financials.xlsx"
                file_path = os.path.join(output_dir, filename)
                if os.path.exists(file_path):
                    year = "Unknown"
                    if info['filed_at'] != "Unknown":
                        try:
                            year = info['filed_at'].split('-')[0]
                        except:
                            year = "Unknown"
                    file_year_mapping.append((file_path, year))

            detailed_file, no_empty_file, row_wise_cleaned_file, dedup_file = self.create_consolidated_file(file_year_mapping, ticker, form_type, output_dir)
            if detailed_file:
                downloaded_files.append(detailed_file)
            if no_empty_file:
                downloaded_files.append(no_empty_file)
            if row_wise_cleaned_file:
                downloaded_files.append(row_wise_cleaned_file)
            if dedup_file:
                downloaded_files.append(dedup_file)
        
        # Create batch script for manual execution
        batch_script = self.create_batch_download_script(excel_info, ticker, form_type)
        
        # Summary
        print(f"\n📊 {form_type} Download Summary:")
        print(f"   Total filings found: {len(filings)}")
        print(f"   Successfully downloaded: {len(downloaded_files)}")
        print(f"   Failed downloads: {len(excel_info) - len(downloaded_files)}")
        
        if downloaded_files:
            print(f"\n✅ Downloaded files:")
            existing_files = []
            missing_files = []
            for file_path in downloaded_files:
                try:
                    if file_path and os.path.exists(file_path):
                        existing_files.append(file_path)
                    else:
                        missing_files.append(file_path)
                except Exception:
                    missing_files.append(file_path)

            for file_path in existing_files:
                try:
                    file_size = os.path.getsize(file_path)
                except Exception:
                    file_size = 0
                file_type = "📄 Detailed" if "_Detailed.xlsx" in file_path else ("📄 Individual" if "Individual_Financials" in file_path else "📄 File")
                print(f"   {file_type} {os.path.basename(file_path)} ({file_size:,} bytes)")

            if missing_files:
                print(f"\n⚠️ Skipped missing files in summary:")
                for file_path in missing_files:
                    if file_path:
                        print(f"   - {os.path.basename(file_path)} (missing)")
        
        if len(downloaded_files) < len(excel_info):
            print(f"\n💡 For failed downloads, try:")
            print(f"   1. Run the batch script: {batch_script}")
            print(f"   2. Use the HTML report for manual downloads")
            print(f"   3. Try downloading from a different network/VPN")
        
        return downloaded_files

def main():
    """Main function"""
    API_KEY = "62ff63ea351833fb6ad40b2f4becbf5539a91740ce09544e96b42600de5853c5"
    
    print("🤖 Advanced Automated SEC Excel Downloader")
    print("=" * 60)
    print("This tool uses multiple techniques to bypass SEC restrictions:")
    print("- curl command line tool")
    print("- wget command line tool") 
    print("- Advanced requests with session management")
    print("- Random delays and user agent rotation")
    print("- Universal financial file cleaner for duplicate removal")
    print()
    
    downloader = AdvancedSECDownloader(API_KEY)
    
    # Get user input with validation
    while True:
        ticker = input("Enter company ticker (e.g., MSFT, AAPL) [MSFT]: ").strip().upper() or "MSFT"
        
        # Validate ticker
        if downloader.validate_ticker(ticker):
            break
        else:
            print(f"❌ Invalid ticker: {ticker}")
            print("Please enter a valid ticker symbol (e.g., MSFT, AAPL, GOOGL)")
            continue
    
    # Ask if user wants both 10-K and 10-Q
    download_both_input = input("Download both 10-K and 10-Q? (y/n) [n]: ").strip().lower()
    download_both = download_both_input in ['y', 'yes', '1', 'true']
    
    if not download_both:
        form_type = input("Enter form type (10-K or 10-Q) [10-K]: ").strip() or "10-K"
    else:
        form_type = "both"  # Placeholder for display
    
    # Get date range
    start_date = input("Enter start date (YYYY-MM-DD) or press Enter for no limit: ").strip() or None
    end_date = input("Enter end date (YYYY-MM-DD) or press Enter for no limit: ").strip() or None
    
    # Validate dates
    if start_date:
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            print("Invalid start date format. Using no start date limit.")
            start_date = None
    
    if end_date:
        try:
            datetime.strptime(end_date, "%Y-%m-%d")
        except ValueError:
            print("Invalid end date format. Using no end date limit.")
            end_date = None
    
    # Run automated download
    if download_both:
        downloaded_files = downloader.automated_download_with_range(ticker, "10-K", start_date=start_date, end_date=end_date, download_both=True)
    else:
        downloaded_files = downloader.automated_download(ticker, form_type, start_date=start_date, end_date=end_date)
    
    if downloaded_files:
        print(f"\n🎉 Successfully downloaded {len(downloaded_files)} Excel files!")
        print(f"📁 Check the SEC_Excel_Downloads_{ticker}_{form_type} folder")
        
        # Test cash flow extraction on the first individual file
        individual_files = [f for f in downloaded_files if 'Individual_Financials' in f]
        if individual_files:
            print(f"\n🧪 Testing cash flow extraction on first individual file...")
            downloader.test_cash_flow_extraction(individual_files[0])
    else:
        print(f"\n⚠️ No files were downloaded automatically")
        print(f"💡 Try the generated batch script or manual download methods")

if __name__ == "__main__":
    main()

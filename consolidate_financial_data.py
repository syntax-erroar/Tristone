#!/usr/bin/env python3
"""
Comprehensive Financial Data Consolidator
Creates a single consolidated table with ALL financial items from all years
"""

import openpyxl
import re
from collections import defaultdict, OrderedDict
from typing import Dict, List, Any, Optional
import pandas as pd

class FinancialDataConsolidator:
    """Consolidates financial data from multiple years into a single table"""
    
    def __init__(self, excel_file: str):
        self.excel_file = excel_file
        self.workbook = None
        self.consolidated_data = OrderedDict()
        self.years = []
        self.financial_items = set()
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.excel_file)
            self.years = sorted([sheet.split('_')[0] for sheet in self.workbook.sheetnames])
            print(f"✅ Loaded workbook with {len(self.years)} years: {self.years}")
            return True
        except Exception as e:
            print(f"❌ Error loading workbook: {e}")
            return False
    
    def extract_financial_data(self):
        """Extract financial data from all sheets"""
        print(f"\n🔍 EXTRACTING FINANCIAL DATA FROM ALL SHEETS")
        print("=" * 60)
        
        # Store the original order from the most recent sheet (2025)
        original_order = []
        
        for sheet_name in self.workbook.sheetnames:
            year = sheet_name.split('_')[0]
            sheet = self.workbook[sheet_name]
            
            print(f"\n📊 Processing {year} ({sheet_name})...")
            
            # Extract financial data from this sheet
            year_data, item_order = self._extract_sheet_data_with_order(sheet, year)
            
            # Store original order from the most recent year (2025)
            if year == '2025':
                original_order = item_order
            
            # Store in consolidated data
            for item_name, values in year_data.items():
                if item_name not in self.consolidated_data:
                    self.consolidated_data[item_name] = {}
                self.consolidated_data[item_name][year] = values
                self.financial_items.add(item_name)
            
            print(f"   ✓ Extracted {len(year_data)} financial items")
        
        # Store the original order for use in table creation
        self.original_order = original_order
        
        print(f"\n📈 CONSOLIDATION SUMMARY")
        print("=" * 60)
        print(f"Total unique financial items: {len(self.financial_items)}")
        print(f"Years processed: {len(self.years)}")
        print(f"Original order preserved from 2025 filing")
        
        return True
    
    def _extract_sheet_data_with_order(self, sheet, year: str) -> tuple[Dict[str, Any], List[str]]:
        """Extract financial data from a single sheet while preserving order"""
        year_data = {}
        item_order = []
        
        # Look for the main financial statement section
        for row in range(1, min(200, sheet.max_row + 1)):  # Check first 200 rows
            row_data = []
            for col in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                row_data.append(value)
            
            first_cell = row_data[0] if row_data else None
            
            if first_cell and isinstance(first_cell, str):
                first_cell = str(first_cell).strip()
                
                # Skip headers and non-financial items
                if (first_cell and 
                    not first_cell.startswith('MSFT') and 
                    not first_cell.startswith('(In millions') and
                    not first_cell.startswith('Year Ended') and
                    not first_cell.startswith('June 30') and
                    len(first_cell) > 2 and
                    not first_cell.isdigit() and
                    not first_cell.endswith(':')):
                    
                    # Check if this row has financial data
                    financial_values = self._extract_financial_values(row_data[1:])
                    
                    if financial_values and any(v is not None for v in financial_values):
                        # Clean up the item name
                        clean_name = self._clean_item_name(first_cell)
                        year_data[clean_name] = financial_values
                        item_order.append(clean_name)
        
        return year_data, item_order
    
    def _extract_sheet_data(self, sheet, year: str) -> Dict[str, Any]:
        """Extract financial data from a single sheet (legacy method)"""
        year_data, _ = self._extract_sheet_data_with_order(sheet, year)
        return year_data
    
    def _extract_financial_values(self, row_data: List[Any]) -> List[Any]:
        """Extract financial values from a row"""
        financial_values = []
        
        for val in row_data:
            if val is not None:
                if isinstance(val, (int, float)):
                    financial_values.append(val)
                elif isinstance(val, str):
                    # Try to convert string numbers
                    cleaned = val.replace(',', '').replace('$', '').strip()
                    if re.match(r'^[\d,]+\.?\d*$', cleaned):
                        try:
                            financial_values.append(float(cleaned))
                        except:
                            financial_values.append(val)
                    elif cleaned:
                        financial_values.append(val)
                    else:
                        financial_values.append(None)
                else:
                    financial_values.append(val)
            else:
                financial_values.append(None)
        
        return financial_values
    
    def _clean_item_name(self, name: str) -> str:
        """Clean up financial item names"""
        # Remove common suffixes and clean up
        name = name.strip()
        
        # Remove trailing colons
        if name.endswith(':'):
            name = name[:-1]
        
        # Remove extra whitespace
        name = re.sub(r'\s+', ' ', name)
        
        return name
    
    def create_consolidated_table(self):
        """Create the consolidated table structure using original filing order"""
        print(f"\n🔧 CREATING CONSOLIDATED TABLE")
        print("=" * 60)
        
        # Create the consolidated matrix
        consolidated_matrix = []
        
        # Header row
        header = ['Financial Item'] + self.years
        consolidated_matrix.append(header)
        
        # Use original order from the 2025 filing, removing duplicates
        ordered_items = []
        seen_items = set()
        
        # First, add items in their original order from the 2025 filing (no duplicates)
        for item in self.original_order:
            if item in self.financial_items and item not in seen_items:
                ordered_items.append(item)
                seen_items.add(item)
        
        # Then add any remaining items that weren't in the 2025 filing
        remaining_items = [item for item in self.financial_items if item not in seen_items]
        ordered_items.extend(sorted(remaining_items))
        
        # Create rows in the original order
        for item_name in ordered_items:
            row = [item_name]
            has_empty_value = False
            
            for year in self.years:
                if item_name in self.consolidated_data and year in self.consolidated_data[item_name]:
                    values = self.consolidated_data[item_name][year]
                    # Take the first non-None value, or the first value if all are None
                    value = next((v for v in values if v is not None), values[0] if values else None)
                    row.append(value)
                    
                    # Check if this value is empty
                    if value is None or (isinstance(value, str) and value.strip() == ''):
                        has_empty_value = True
                else:
                    row.append(None)
                    has_empty_value = True
            
            # Only add rows that have values for ALL years
            if not has_empty_value:
                consolidated_matrix.append(row)
        
        print(f"✅ Created consolidated table: {len(consolidated_matrix)} rows × {len(consolidated_matrix[0])} columns")
        print(f"✅ Items ordered by original 2025 filing sequence: {len(self.original_order)} items")
        print(f"✅ Additional items added: {len(remaining_items)} items")
        
        return consolidated_matrix
    
    def save_consolidated_excel(self, output_file: str):
        """Save the consolidated data to Excel"""
        print(f"\n💾 SAVING CONSOLIDATED EXCEL")
        print("=" * 60)
        
        # Create consolidated table
        consolidated_matrix = self.create_consolidated_table()
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidated Financial Data"
        
        # Write data to Excel
        for row_idx, row in enumerate(consolidated_matrix, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # Format header row
                if row_idx == 1:
                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                    cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Format financial item names (first column)
                elif col_idx == 1:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Format financial values (other columns)
                else:
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0'
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        wb.save(output_file)
        print(f"✅ Saved consolidated data to: {output_file}")
        
        return output_file
    
    def create_summary_report(self):
        """Create a summary report of the consolidation"""
        print(f"\n📊 CONSOLIDATION SUMMARY REPORT")
        print("=" * 60)
        
        # Count items by year
        year_counts = defaultdict(int)
        for item_name, year_data in self.consolidated_data.items():
            for year in year_data.keys():
                year_counts[year] += 1
        
        print(f"Financial items per year:")
        for year in sorted(year_counts.keys()):
            print(f"  {year}: {year_counts[year]} items")
        
        # Find common items across all years
        common_items = set(self.financial_items)
        for item_name, year_data in self.consolidated_data.items():
            if len(year_data) < len(self.years):
                common_items.discard(item_name)
        
        print(f"\nItems present in ALL years: {len(common_items)}")
        
        # Show sample common items
        if common_items:
            print(f"\nSample common items:")
            for item in sorted(list(common_items))[:10]:
                print(f"  - {item}")
        
        return {
            'total_items': len(self.financial_items),
            'years': len(self.years),
            'common_items': len(common_items),
            'year_counts': dict(year_counts)
        }

def main():
    """Main function to consolidate financial data"""
    input_file = "MSFT_10-K_Multi_Year_Cleaned.xlsx"
    output_file = "MSFT_Consolidated_Financial_Data_Deduplicated.xlsx"
    
    print("🏦 COMPREHENSIVE FINANCIAL DATA CONSOLIDATOR")
    print("=" * 60)
    
    # Initialize consolidator
    consolidator = FinancialDataConsolidator(input_file)
    
    # Load workbook
    if not consolidator.load_workbook():
        return
    
    # Extract financial data
    if not consolidator.extract_financial_data():
        return
    
    # Create summary report
    summary = consolidator.create_summary_report()
    
    # Save consolidated Excel
    output_path = consolidator.save_consolidated_excel(output_file)
    
    print(f"\n🎉 CONSOLIDATION COMPLETE!")
    print("=" * 60)
    print(f"✅ Input file: {input_file}")
    print(f"✅ Output file: {output_path}")
    print(f"✅ Total financial items: {summary['total_items']}")
    print(f"✅ Years consolidated: {summary['years']}")
    print(f"✅ Common items across all years: {summary['common_items']}")

if __name__ == "__main__":
    main()

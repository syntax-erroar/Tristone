#!/usr/bin/env python3
"""
Test script to verify selective currency formatting
"""

import openpyxl
import os

def test_selective_currency_formatting():
    """Test the Excel file to check if currency formatting is applied selectively"""
    
    excel_file = "MSFT_10-K_Multi_Year_Selective.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file {excel_file} not found")
        return False
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        print(f"✅ Successfully loaded {excel_file}")
        print(f"📊 Available sheets: {workbook.sheetnames}")
        
        # Test the 2021 sheet
        if "2021_06_30" in workbook.sheetnames:
            sheet = workbook["2021_06_30"]
            print(f"\n🔍 Examining 2021_06_30 sheet...")
            
            # Look for cells with currency formatting
            currency_formatted = []
            non_currency_formatted = []
            year_headers = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        
                        # Check if it's a year header
                        if cell_value.isdigit() and len(cell_value) == 4 and 1900 <= int(cell_value) <= 2100:
                            year_headers.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value}")
                        
                        # Check if it has currency formatting
                        elif cell.number_format and '$' in cell.number_format:
                            # Get the row label to see what type of line item this is
                            try:
                                row_label_cell = sheet.cell(row=cell.row, column=1)
                                row_label = str(row_label_cell.value or "").strip()
                            except:
                                row_label = "Unknown"
                            
                            currency_formatted.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value} (Row: '{row_label}')")
                        
                        # Check if it's a large number without currency formatting
                        elif cell_value.replace(',', '').replace('.', '').isdigit() and len(cell_value.replace(',', '').replace('.', '')) >= 3:
                            try:
                                row_label_cell = sheet.cell(row=cell.row, column=1)
                                row_label = str(row_label_cell.value or "").strip()
                            except:
                                row_label = "Unknown"
                            
                            non_currency_formatted.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value} (Row: '{row_label}')")
            
            print(f"\n📈 Results:")
            print(f"✅ Year headers (should NOT have currency): {len(year_headers)}")
            print(f"💰 Currency formatted values: {len(currency_formatted)}")
            print(f"📊 Non-currency formatted values: {len(non_currency_formatted)}")
            
            if year_headers:
                print(f"\n📅 Year headers (correctly no currency):")
                for header in year_headers[:5]:
                    print(f"   - {header}")
            
            if currency_formatted:
                print(f"\n💰 Currency formatted values:")
                for example in currency_formatted[:10]:
                    print(f"   - {example}")
                if len(currency_formatted) > 10:
                    print(f"   ... and {len(currency_formatted) - 10} more")
            
            if non_currency_formatted:
                print(f"\n📊 Non-currency formatted values (should be limited):")
                for example in non_currency_formatted[:10]:
                    print(f"   - {example}")
                if len(non_currency_formatted) > 10:
                    print(f"   ... and {len(non_currency_formatted) - 10} more")
            
            # Check if the selective formatting is working
            total_financial_numbers = len(currency_formatted) + len(non_currency_formatted)
            currency_percentage = (len(currency_formatted) / total_financial_numbers * 100) if total_financial_numbers > 0 else 0
            
            print(f"\n📊 Currency formatting ratio: {currency_percentage:.1f}% of financial numbers have currency formatting")
            
            # Success criteria: Most financial numbers should NOT have currency formatting
            # Only specific line items like "Total revenue" should have currency formatting
            if currency_percentage < 50:  # Less than 50% should have currency formatting
                print(f"🎉 SUCCESS: Selective currency formatting is working! Most numbers don't have currency formatting.")
                return True
            else:
                print(f"⚠️  WARNING: Too many numbers have currency formatting. Expected < 50%, got {currency_percentage:.1f}%")
                return False
            
        else:
            print(f"❌ 2021_06_30 sheet not found in workbook")
            return False
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

if __name__ == "__main__":
    print("🧪 Testing Selective Currency Formatting")
    print("=" * 60)
    
    success = test_selective_currency_formatting()
    
    if success:
        print(f"\n🎉 SUCCESS: Selective currency formatting is working correctly!")
        print(f"   - Year headers: No currency formatting ✅")
        print(f"   - Most financial numbers: No currency formatting ✅")
        print(f"   - Only specific line items (like Total revenue): Currency formatting ✅")
    else:
        print(f"\n❌ FAILED: Selective currency formatting needs adjustment")

#!/usr/bin/env python3
"""
Year-Based Financial Table Filter
Filters Excel files to extract only tables with year-based financial data.
"""

import openpyxl
import re
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass

@dataclass
class YearBasedTable:
    """Represents a table with year-based financial data"""
    start_row: int
    end_row: int
    years: List[int]
    table_type: str
    confidence_score: float

class YearBasedTableFilter:
    """Filters Excel files for year-based financial tables"""
    
    def __init__(self):
        self.year_pattern = re.compile(r'\b(19|20)\d{2}\b')  # 1900-2099
        self.financial_keywords = [
            'revenue', 'income', 'assets', 'liabilities', 'equity', 'cash',
            'operating', 'gross margin', 'net income', 'total revenue',
            'cost of revenue', 'operating income', 'earnings', 'sales',
            'balance sheet', 'income statement', 'cash flow', 'comprehensive income'
        ]
        
    def is_year_header(self, cell_value) -> bool:
        """Check if a cell contains a year header"""
        if not cell_value:
            return False
        
        cell_str = str(cell_value).strip()
        
        # Check if it's a 4-digit year
        if cell_str.isdigit() and len(cell_str) == 4:
            year = int(cell_str)
            return 1900 <= year <= 2100
        
        # Check for year patterns in text
        years = self.year_pattern.findall(cell_str)
        return len(years) > 0
    
    def extract_years_from_cell(self, cell_value) -> List[int]:
        """Extract years from a cell value"""
        if not cell_value:
            return []
        
        cell_str = str(cell_value)
        years = []
        
        # Find all 4-digit years
        matches = self.year_pattern.findall(cell_str)
        for match in matches:
            year = int(match)
            if 1900 <= year <= 2100:
                years.append(year)
        
        return years
    
    def is_financial_content(self, cell_value) -> bool:
        """Check if a cell contains financial content"""
        if not cell_value:
            return False
        
        cell_str = str(cell_value).lower().strip()
        
        # Check for financial keywords
        for keyword in self.financial_keywords:
            if keyword in cell_str:
                return True
        
        # Check for currency patterns
        currency_patterns = [
            r'\$[\d,]+',  # $123,456
            r'[\d,]+\.?\d*\s*\$',  # 123,456$
            r'\([\d,]+\)',  # (123,456) for negative numbers
            r'[\d,]+\.\d{2}',  # 123,456.78
        ]
        
        for pattern in currency_patterns:
            if re.search(pattern, cell_str):
                return True
        
        return False
    
    def analyze_table_structure(self, worksheet, start_row: int, end_row: int) -> Dict:
        """Analyze the structure of a potential table"""
        analysis = {
            'has_year_headers': False,
            'year_columns': [],
            'years_found': set(),
            'financial_rows': 0,
            'total_rows': end_row - start_row + 1,
            'confidence_score': 0.0
        }
        
        # Check each row in the table
        for row in range(start_row, end_row + 1):
            row_data = [worksheet.cell(row, col).value for col in range(1, 11)]  # Check first 10 columns
            
            # Check for year headers in this row
            year_columns_in_row = []
            for col_idx, cell_value in enumerate(row_data):
                if self.is_year_header(cell_value):
                    analysis['has_year_headers'] = True
                    year_columns_in_row.append(col_idx + 1)
                    years = self.extract_years_from_cell(cell_value)
                    analysis['years_found'].update(years)
            
            if year_columns_in_row:
                analysis['year_columns'].extend(year_columns_in_row)
            
            # Check for financial content
            if any(self.is_financial_content(cell) for cell in row_data):
                analysis['financial_rows'] += 1
        
        # Calculate confidence score
        confidence = 0.0
        
        # Base score for having year headers
        if analysis['has_year_headers']:
            confidence += 0.5
        
        # Score for number of years found (more years = higher confidence)
        years_count = len(analysis['years_found'])
        if years_count >= 3:
            confidence += 0.4
        elif years_count == 2:
            confidence += 0.3
        elif years_count == 1:
            confidence += 0.1
        
        # Score for financial content ratio
        if analysis['total_rows'] > 0:
            financial_ratio = analysis['financial_rows'] / analysis['total_rows']
            confidence += financial_ratio * 0.2
        
        # Bonus for having multiple year columns
        if len(analysis['year_columns']) >= 2:
            confidence += 0.1
        
        analysis['confidence_score'] = min(confidence, 1.0)
        analysis['years_found'] = sorted(list(analysis['years_found']))
        
        return analysis
    
    def find_year_based_tables(self, worksheet) -> List[YearBasedTable]:
        """Find all year-based financial tables in a worksheet"""
        tables = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        
        # Look for potential table boundaries
        current_table_start = None
        current_table_end = None
        
        for row in range(1, max_row + 1):
            # Check if first column of this row has years
            first_cell = worksheet.cell(row, 1).value
            has_years_in_first_column = self.is_year_header(first_cell)
            
            if has_years_in_first_column:
                if current_table_start is None:
                    current_table_start = row
                current_table_end = row
            else:
                # End of current table - check if it has years in first few rows
                if current_table_start is not None and current_table_end is not None:
                    # Check if table has years in first column of first few rows
                    years_found = self._extract_years_from_first_column(worksheet, current_table_start, current_table_end)
                    if years_found:
                        table = YearBasedTable(
                            start_row=current_table_start,
                            end_row=current_table_end,
                            years=years_found,
                            table_type=self._classify_table_type(worksheet, current_table_start, current_table_end),
                            confidence_score=1.0  # Simple logic, high confidence
                        )
                        tables.append(table)
                
                current_table_start = None
                current_table_end = None
        
        # Handle table that goes to the end of the sheet
        if current_table_start is not None and current_table_end is not None:
            years_found = self._extract_years_from_first_column(worksheet, current_table_start, current_table_end)
            if years_found:
                table = YearBasedTable(
                    start_row=current_table_start,
                    end_row=current_table_end,
                    years=years_found,
                    table_type=self._classify_table_type(worksheet, current_table_start, current_table_end),
                    confidence_score=1.0
                )
                tables.append(table)
        
        return tables
    
    def _extract_years_from_first_column(self, worksheet, start_row: int, end_row: int) -> List[int]:
        """Extract years from the first column of a table range (first few rows only)"""
        years_found = set()
        year_pattern = re.compile(r'\b(19|20)\d{2}\b')  # 1900-2099
        
        # Only check first few rows (up to 5 rows) of the table
        max_rows_to_check = min(5, end_row - start_row + 1)
        
        for row in range(start_row, start_row + max_rows_to_check):
            first_cell = worksheet.cell(row, 1).value
            if first_cell:
                content = str(first_cell).strip()
                
                # Check for 4-digit years
                if content.isdigit() and len(content) == 4:
                    year = int(content)
                    if 1900 <= year <= 2100:
                        years_found.add(year)
                else:
                    # Check for years in text
                    years_in_text = year_pattern.findall(content)
                    for year_match in years_in_text:
                        year = int(year_match)
                        if 1900 <= year <= 2100:
                            years_found.add(year)
        
        return sorted(list(years_found))
    
    def _classify_table_type(self, worksheet, start_row: int, end_row: int) -> str:
        """Classify the type of financial table"""
        # Look at the first few rows to determine table type
        for row in range(start_row, min(start_row + 5, end_row + 1)):
            row_data = [str(worksheet.cell(row, col).value or '').lower() for col in range(1, 6)]
            row_text = ' '.join(row_data)
            
            if 'income statement' in row_text or 'operations' in row_text:
                return 'Income Statement'
            elif 'balance sheet' in row_text or 'assets' in row_text and 'liabilities' in row_text:
                return 'Balance Sheet'
            elif 'cash flow' in row_text:
                return 'Cash Flow Statement'
            elif 'comprehensive income' in row_text:
                return 'Comprehensive Income'
            elif 'equity' in row_text and 'shareholders' in row_text:
                return 'Shareholders Equity'
        
        return 'Financial Table'
    
    def filter_excel_file(self, input_file: str, output_file: str) -> Dict:
        """Filter an Excel file to keep only year-based financial tables"""
        try:
            wb = openpyxl.load_workbook(input_file)
            result = {
                'input_file': input_file,
                'output_file': output_file,
                'sheets_processed': 0,
                'tables_found': 0,
                'tables_kept': 0,
                'sheets_kept': []
            }
            
            # Create new workbook
            new_wb = openpyxl.Workbook()
            new_wb.remove(new_wb.active)
            
            for sheet_name in wb.sheetnames:
                worksheet = wb[sheet_name]
                result['sheets_processed'] += 1
                
                # Find year-based tables in this sheet
                tables = self.find_year_based_tables(worksheet)
                result['tables_found'] += len(tables)
                
                if tables:
                    # Create new sheet with filtered content
                    new_sheet = new_wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
                    
                    # Add header
                    new_sheet.cell(row=1, column=1, value=f"Filtered Financial Data - {sheet_name}")
                    new_sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                    
                    current_row = 3
                    
                    for table in tables:
                        # Add table header
                        new_sheet.cell(row=current_row, column=1, value=f"{table.table_type} (Years: {', '.join(map(str, table.years))})")
                        new_sheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                        current_row += 2
                        
                        # Copy table data
                        for row in range(table.start_row, table.end_row + 1):
                            for col in range(1, 11):  # Copy first 10 columns
                                source_cell = worksheet.cell(row, col)
                                target_cell = new_sheet.cell(current_row, col)
                                target_cell.value = source_cell.value
                                
                                # Copy formatting (avoid StyleProxy issues)
                                try:
                                    if source_cell.font:
                                        target_cell.font = openpyxl.styles.Font(
                                            name=source_cell.font.name,
                                            size=source_cell.font.size,
                                            bold=source_cell.font.bold,
                                            italic=source_cell.font.italic
                                        )
                                    if source_cell.alignment:
                                        target_cell.alignment = openpyxl.styles.Alignment(
                                            horizontal=source_cell.alignment.horizontal,
                                            vertical=source_cell.alignment.vertical
                                        )
                                    if source_cell.number_format:
                                        target_cell.number_format = source_cell.number_format
                                except:
                                    pass  # Skip formatting if there are issues
                            
                            current_row += 1
                        
                        current_row += 2  # Space between tables
                        result['tables_kept'] += 1
                    
                    result['sheets_kept'].append(sheet_name)
            
            # Save filtered workbook
            if result['sheets_kept']:
                new_wb.save(output_file)
                print(f"✓ Filtered file saved: {output_file}")
            else:
                print("⚠ No year-based financial tables found")
            
            return result
            
        except Exception as e:
            print(f"Error filtering {input_file}: {e}")
            return {'error': str(e)}

def main():
    """Main function"""
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python year_based_table_filter.py <input_file> [output_file]")
        print("Example: python year_based_table_filter.py MSFT_10-K_Multi_Year_Cleaned.xlsx MSFT_Filtered.xlsx")
        return
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace('.xlsx', '_YearBased_Filtered.xlsx')
    
    filter_obj = YearBasedTableFilter()
    result = filter_obj.filter_excel_file(input_file, output_file)
    
    print(f"\n=== Filtering Results ===")
    print(f"Input file: {result.get('input_file', 'N/A')}")
    print(f"Output file: {result.get('output_file', 'N/A')}")
    print(f"Sheets processed: {result.get('sheets_processed', 0)}")
    print(f"Tables found: {result.get('tables_found', 0)}")
    print(f"Tables kept: {result.get('tables_kept', 0)}")
    print(f"Sheets kept: {len(result.get('sheets_kept', []))}")
    
    if result.get('sheets_kept'):
        print(f"Sheets with year-based tables: {', '.join(result['sheets_kept'])}")

if __name__ == "__main__":
    main()

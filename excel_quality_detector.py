#!/usr/bin/env python3
"""
Excel Quality Detector and Fixer
Detects and fixes common Excel formatting issues
"""

import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ExcelQualityDetector:
    """Detect and fix Excel quality issues"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.issues_found = []
    
    def detect_issues(self):
        """Detect all quality issues in the Excel file"""
        print("🔍 Detecting Excel Quality Issues...")
        print("=" * 50)
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            print(f"\n📊 Analyzing sheet: {sheet_name}")
            
            # Detect incomplete values
            incomplete_values = self._detect_incomplete_values(sheet)
            if incomplete_values:
                self.issues_found.extend(incomplete_values)
                print(f"  ❌ Found {len(incomplete_values)} incomplete values")
            
            # Detect empty rows/columns
            empty_rows = self._detect_empty_rows(sheet)
            empty_cols = self._detect_empty_columns(sheet)
            
            if empty_rows:
                print(f"  ❌ Found {len(empty_rows)} empty rows: {empty_rows[:5]}{'...' if len(empty_rows) > 5 else ''}")
            if empty_cols:
                print(f"  ❌ Found {len(empty_cols)} empty columns: {empty_cols[:5]}{'...' if len(empty_cols) > 5 else ''}")
            
            # Detect formatting issues
            formatting_issues = self._detect_formatting_issues(sheet)
            if formatting_issues:
                print(f"  ❌ Found {len(formatting_issues)} formatting issues")
        
        return self.issues_found
    
    def _detect_incomplete_values(self, sheet):
        """Detect incomplete financial values"""
        issues = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    
                    # Check for incomplete parentheses
                    if value.startswith('(') and not value.endswith(')'):
                        issues.append({
                            'type': 'incomplete_parentheses',
                            'sheet': sheet.title,
                            'cell': f"{get_column_letter(cell.column)}{cell.row}",
                            'value': value,
                            'expected': value + ')'
                        })
                    
                    # Check for incomplete currency
                    if re.match(r'^\$[\d,]+$', value) and not value.endswith('.'):
                        issues.append({
                            'type': 'incomplete_currency',
                            'sheet': sheet.title,
                            'cell': f"{get_column_letter(cell.column)}{cell.row}",
                            'value': value,
                            'expected': value + '.00'
                        })
                    
                    # Check for orphaned characters
                    if re.match(r'^[\)\$\s]+$', value):
                        issues.append({
                            'type': 'orphaned_characters',
                            'sheet': sheet.title,
                            'cell': f"{get_column_letter(cell.column)}{cell.row}",
                            'value': value,
                            'expected': 'EMPTY'
                        })
        
        return issues
    
    def _detect_empty_rows(self, sheet):
        """Detect completely empty rows"""
        empty_rows = []
        
        for row_num, row in enumerate(sheet.iter_rows(), 1):
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                empty_rows.append(row_num)
        
        return empty_rows
    
    def _detect_empty_columns(self, sheet):
        """Detect completely empty columns"""
        empty_cols = []
        
        for col_num in range(1, sheet.max_column + 1):
            is_empty = True
            for row_num in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None and str(cell.value).strip() != '':
                    is_empty = False
                    break
            
            if is_empty:
                empty_cols.append(col_num)
        
        return empty_cols
    
    def _detect_formatting_issues(self, sheet):
        """Detect formatting inconsistencies"""
        issues = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    
                    # Check for inconsistent number formatting
                    if re.match(r'^\d+$', value) and len(value) > 3:
                        # Should have commas for thousands
                        if ',' not in value:
                            issues.append({
                                'type': 'missing_commas',
                                'sheet': sheet.title,
                                'cell': f"{get_column_letter(cell.column)}{cell.row}",
                                'value': value,
                                'expected': f"{int(value):,}"
                            })
        
        return issues
    
    def fix_issues(self):
        """Fix all detected issues"""
        print("\n🔧 Fixing Excel Quality Issues...")
        print("=" * 50)
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            print(f"\n📊 Fixing sheet: {sheet_name}")
            
            # Fix incomplete values
            self._fix_incomplete_values(sheet)
            
            # Remove empty rows and columns
            self._remove_empty_rows_columns(sheet)
            
            # Fix formatting
            self._fix_formatting(sheet)
        
        print("\n✅ All issues fixed!")
    
    def _fix_incomplete_values(self, sheet):
        """Fix incomplete financial values"""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    
                    # Fix incomplete parentheses
                    if value.startswith('(') and not value.endswith(')'):
                        cell.value = value + ')'
                        print(f"  ✓ Fixed incomplete parentheses: {value} → {value + ')'}")
                    
                    # Fix incomplete currency
                    elif re.match(r'^\$[\d,]+$', value) and not value.endswith('.'):
                        cell.value = value + '.00'
                        print(f"  ✓ Fixed incomplete currency: {value} → {value + '.00'}")
                    
                    # Remove orphaned characters
                    elif re.match(r'^[\)\$\s]+$', value):
                        cell.value = None
                        print(f"  ✓ Removed orphaned characters: {value}")
    
    def _remove_empty_rows_columns(self, sheet):
        """Remove empty rows and columns"""
        # Remove empty rows (from bottom to top to avoid index issues)
        for row_num in range(sheet.max_row, 0, -1):
            row = [sheet.cell(row=row_num, column=col).value for col in range(1, sheet.max_column + 1)]
            if all(cell is None or str(cell).strip() == '' for cell in row):
                sheet.delete_rows(row_num)
                print(f"  ✓ Removed empty row {row_num}")
        
        # Remove empty columns (from right to left to avoid index issues)
        for col_num in range(sheet.max_column, 0, -1):
            col = [sheet.cell(row=row, column=col_num).value for row in range(1, sheet.max_row + 1)]
            if all(cell is None or str(cell).strip() == '' for cell in col):
                sheet.delete_cols(col_num)
                print(f"  ✓ Removed empty column {get_column_letter(col_num)}")
    
    def _fix_formatting(self, sheet):
        """Fix number formatting"""
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    value = str(cell.value).strip()
                    
                    # Add commas to large numbers
                    if re.match(r'^\d+$', value) and len(value) > 3:
                        formatted_value = f"{int(value):,}"
                        cell.value = formatted_value
                        print(f"  ✓ Added commas: {value} → {formatted_value}")
    
    def save(self, output_path=None):
        """Save the fixed workbook"""
        if output_path:
            self.workbook.save(output_path)
        else:
            self.workbook.save(self.file_path)
        print(f"\n💾 Saved fixed file: {output_path or self.file_path}")

def main():
    """Main function to detect and fix Excel issues"""
    input_file = "MSFT_10-K_Multi_Year_Cleaned.xlsx"
    output_file = "MSFT_10-K_Multi_Year_Fixed.xlsx"
    
    print("🧹 Excel Quality Detector and Fixer")
    print("=" * 50)
    
    # Detect issues
    detector = ExcelQualityDetector(input_file)
    issues = detector.detect_issues()
    
    if not issues:
        print("\n🎉 No issues found! Your Excel file is already clean.")
        return
    
    print(f"\n📊 Summary: Found {len(issues)} total issues")
    
    # Fix issues
    detector.fix_issues()
    
    # Save fixed file
    detector.save(output_file)
    
    print(f"\n✅ Excel file cleaned and saved as: {output_file}")

if __name__ == "__main__":
    main()

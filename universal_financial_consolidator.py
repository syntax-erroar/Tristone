#!/usr/bin/env python3
"""
Universal Financial Data Consolidator
Works with any company's multi-year financial data
"""

import openpyxl
import re
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass

@dataclass
class CompanyConfig:
    """Configuration for different companies"""
    ticker: str
    fiscal_year_end: str  # e.g., "December 31", "June 30", "September 30"
    currency_unit: str    # e.g., "millions", "thousands", "billions"
    company_name_patterns: List[str]  # Patterns to identify company headers

class UniversalFinancialConsolidator:
    """Universal financial data consolidator for any company"""
    
    def __init__(self, config: CompanyConfig):
        self.config = config
        self.workbook = None
        self.consolidated_data = {}
        self.financial_items = set()
        self.years = []
        self.original_order = []
    
    def load_workbook(self, input_file: str) -> bool:
        """Load the multi-year Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(input_file)
            self.years = sorted([sheet.split('_')[0] for sheet in self.workbook.sheetnames 
                               if sheet.split('_')[0].isdigit()], reverse=True)
            print(f"✅ Loaded workbook with {len(self.years)} years: {self.years}")
            return True
        except Exception as e:
            print(f"❌ Error loading workbook: {e}")
            return False
    
    def extract_financial_data(self):
        """Extract financial data from all sheets"""
        print(f"\n🔍 EXTRACTING FINANCIAL DATA FROM ALL SHEETS")
        print("=" * 60)
        
        # Store the original order from the most recent sheet
        original_order = []
        most_recent_year = self.years[0] if self.years else None
        
        for sheet_name in self.workbook.sheetnames:
            year = sheet_name.split('_')[0]
            sheet = self.workbook[sheet_name]
            
            print(f"\n📊 Processing {year} ({sheet_name})...")
            
            # Extract financial data from this sheet
            year_data, item_order = self._extract_sheet_data_with_order(sheet, year)
            
            # Store original order from the most recent year
            if year == most_recent_year:
                original_order = item_order
            
            # Store in consolidated data
            for item_name, values in year_data.items():
                if item_name not in self.consolidated_data:
                    self.consolidated_data[item_name] = {}
                self.consolidated_data[item_name][year] = values
                self.financial_items.add(item_name)
            
            print(f"   ✓ Extracted {len(year_data)} financial items")
        
        # Store the original order for use in table creation
        self.original_order = original_order
        
        print(f"\n📈 CONSOLIDATION SUMMARY")
        print("=" * 60)
        print(f"Total unique financial items: {len(self.financial_items)}")
        print(f"Years processed: {len(self.years)}")
        print(f"Original order preserved from {most_recent_year} filing")
        
        return True
    
    def _extract_sheet_data_with_order(self, sheet, year: str) -> Tuple[Dict[str, Any], List[str]]:
        """Extract financial data from a single sheet while preserving order"""
        year_data = {}
        item_order = []
        
        # Look for the main financial statement section
        for row in range(1, min(200, sheet.max_row + 1)):  # Check first 200 rows
            row_data = []
            for col in range(1, min(20, sheet.max_column + 1)):  # Check first 20 columns
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                row_data.append(value)
            
            first_cell = row_data[0] if row_data else None
            
            if first_cell and isinstance(first_cell, str):
                first_cell = str(first_cell).strip()
                
                # Skip headers and non-financial items (universal patterns)
                if self._is_valid_financial_item(first_cell):
                    # Check if this row has financial data
                    financial_values = self._extract_financial_values(row_data[1:])
                    
                    if financial_values and any(v is not None for v in financial_values):
                        # Clean up the item name
                        clean_name = self._clean_item_name(first_cell)
                        year_data[clean_name] = financial_values
                        item_order.append(clean_name)
        
        return year_data, item_order
    
    def _is_valid_financial_item(self, text: str) -> bool:
        """Check if text represents a valid financial item (universal)"""
        if not text or len(text) <= 2:
            return False
        
        # Skip common header patterns (universal)
        skip_patterns = [
            r'^\d{4}$',  # Years
            r'^\d{1,2}$',  # Single/double digits
            r'^\(In\s+\w+\)$',  # Currency units like "(In millions)"
            r'^Year\s+Ended',  # "Year Ended" headers
            r'^\w+\s+\d{1,2},?\s+\d{4}$',  # Date patterns
            r'^$',  # Empty strings
            r'^\s*$',  # Whitespace only
            r'^.*:$',  # Ends with colon
        ]
        
        for pattern in skip_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return False
        
        # Skip company-specific patterns
        for pattern in self.config.company_name_patterns:
            if text.upper().startswith(pattern.upper()):
                return False
        
        # Skip fiscal year end patterns
        if self.config.fiscal_year_end.lower() in text.lower():
            return False
        
        return True
    
    def _extract_financial_values(self, row_data: List[Any]) -> List[Any]:
        """Extract financial values from a row (universal)"""
        financial_values = []
        
        for value in row_data:
            if value is None:
                financial_values.append(None)
            elif isinstance(value, (int, float)):
                financial_values.append(value)
            elif isinstance(value, str):
                # Try to extract numeric value
                cleaned = re.sub(r'[$,()]', '', value.strip())
                if cleaned.replace('.', '').replace('-', '').isdigit():
                    try:
                        numeric_value = float(cleaned)
                        financial_values.append(numeric_value)
                    except ValueError:
                        financial_values.append(None)
                else:
                    financial_values.append(None)
            else:
                financial_values.append(None)
        
        return financial_values
    
    def _clean_item_name(self, name: str) -> str:
        """Clean financial item name (universal)"""
        if not name:
            return name
        
        # Remove common prefixes/suffixes
        name = re.sub(r'^\s*[-•*]\s*', '', name)  # Remove bullet points
        name = re.sub(r'\s*[-•*]\s*$', '', name)  # Remove trailing bullets
        name = re.sub(r'^\s*\d+\.\s*', '', name)  # Remove numbering
        name = re.sub(r'\s*\(\d+\)\s*$', '', name)  # Remove trailing numbers in parentheses
        
        # Remove extra whitespace
        name = re.sub(r'\s+', ' ', name)
        
        return name.strip()
    
    def create_consolidated_table(self):
        """Create the consolidated table structure using original filing order"""
        print(f"\n🔧 CREATING CONSOLIDATED TABLE")
        print("=" * 60)
        
        # Create the consolidated matrix
        consolidated_matrix = []
        
        # Header row
        header = ['Financial Item'] + self.years
        consolidated_matrix.append(header)
        
        # Use original order from the most recent filing, removing duplicates
        ordered_items = []
        seen_items = set()
        
        # First, add items in their original order from the most recent filing (no duplicates)
        for item in self.original_order:
            if item in self.financial_items and item not in seen_items:
                ordered_items.append(item)
                seen_items.add(item)
        
        # Then add any remaining items that weren't in the most recent filing
        remaining_items = [item for item in self.financial_items if item not in seen_items]
        ordered_items.extend(sorted(remaining_items))
        
        # Create rows in the original order
        for item_name in ordered_items:
            row = [item_name]
            has_empty_value = False
            
            for year in self.years:
                if item_name in self.consolidated_data and year in self.consolidated_data[item_name]:
                    values = self.consolidated_data[item_name][year]
                    # Take the first non-None value, or the first value if all are None
                    value = next((v for v in values if v is not None), values[0] if values else None)
                    row.append(value)
                    
                    # Check if this value is empty
                    if value is None or (isinstance(value, str) and value.strip() == ''):
                        has_empty_value = True
                else:
                    row.append(None)
                    has_empty_value = True
            
            # Only add rows that have values for ALL years
            if not has_empty_value:
                consolidated_matrix.append(row)
        
        print(f"✅ Created consolidated table: {len(consolidated_matrix)} rows × {len(consolidated_matrix[0])} columns")
        print(f"✅ Items ordered by original {self.years[0]} filing sequence: {len(self.original_order)} items")
        print(f"✅ Additional items added: {len(remaining_items)} items")
        
        return consolidated_matrix
    
    def save_consolidated_excel(self, output_file: str):
        """Save the consolidated data to Excel"""
        print(f"\n💾 SAVING CONSOLIDATED EXCEL")
        print("=" * 60)
        
        # Create consolidated table
        consolidated_matrix = self.create_consolidated_table()
        
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{self.config.ticker}_Consolidated"
        
        # Write data to worksheet
        for row_idx, row in enumerate(consolidated_matrix, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format header row
                if row_idx == 1:
                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Format financial item names (first column)
                elif col_idx == 1:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                # Format financial values (other columns)
                else:
                    if isinstance(value, (int, float)) and value != 0:
                        cell.number_format = '#,##0'
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_file)
        print(f"✅ Saved consolidated data to: {output_file}")
    
    def generate_summary_report(self):
        """Generate a summary report of the consolidation"""
        print(f"\n📊 CONSOLIDATION SUMMARY REPORT")
        print("=" * 60)
        
        # Count items per year
        year_counts = {}
        for year in self.years:
            count = 0
            for item_data in self.consolidated_data.values():
                if year in item_data:
                    count += 1
            year_counts[year] = count
        
        print(f"Financial items per year:")
        for year in self.years:
            print(f"  {year}: {year_counts[year]} items")
        
        # Find common items across all years
        common_items = set()
        if self.consolidated_data:
            common_items = set(self.consolidated_data.keys())
            for item_data in self.consolidated_data.values():
                common_items = common_items.intersection(set(item_data.keys()))
        
        print(f"\nItems present in ALL years: {len(common_items)}")
        
        if common_items:
            print(f"\nSample common items:")
            for item in sorted(list(common_items))[:10]:
                print(f"  - {item}")
        
        return {
            'total_items': len(self.financial_items),
            'years': len(self.years),
            'common_items': len(common_items),
            'year_counts': dict(year_counts)
        }

def main():
    """Main function to consolidate financial data for any company"""
    
    # Example configurations for different companies
    company_configs = {
        'MSFT': CompanyConfig(
            ticker='MSFT',
            fiscal_year_end='June 30',
            currency_unit='millions',
            company_name_patterns=['MSFT', 'Microsoft']
        ),
        'AAPL': CompanyConfig(
            ticker='AAPL',
            fiscal_year_end='September 30',
            currency_unit='millions',
            company_name_patterns=['AAPL', 'Apple']
        ),
        'GOOGL': CompanyConfig(
            ticker='GOOGL',
            fiscal_year_end='December 31',
            currency_unit='millions',
            company_name_patterns=['GOOGL', 'GOOG', 'Alphabet', 'Google']
        )
    }
    
    # Get company from user or use default
    company = input("Enter company ticker (MSFT/AAPL/GOOGL) or press Enter for MSFT: ").strip().upper()
    if not company or company not in company_configs:
        company = 'MSFT'
    
    config = company_configs[company]
    
    # File names
    input_file = f"{company}_10-K_Multi_Year_Cleaned.xlsx"
    output_file = f"{company}_Consolidated_Financial_Data_Universal.xlsx"
    
    print(f"🏦 UNIVERSAL FINANCIAL DATA CONSOLIDATOR")
    print(f"Company: {config.ticker} ({config.fiscal_year_end} year-end)")
    print("=" * 60)
    
    # Initialize consolidator
    consolidator = UniversalFinancialConsolidator(config)
    
    # Load workbook
    if not consolidator.load_workbook(input_file):
        return
    
    # Extract financial data
    if not consolidator.extract_financial_data():
        return
    
    # Generate summary report
    consolidator.generate_summary_report()
    
    # Save consolidated Excel
    consolidator.save_consolidated_excel(output_file)
    
    print(f"\n🎉 CONSOLIDATION COMPLETE!")
    print("=" * 60)
    print(f"✅ Input file: {input_file}")
    print(f"✅ Output file: {output_file}")
    print(f"✅ Company: {config.ticker}")
    print(f"✅ Fiscal year end: {config.fiscal_year_end}")

if __name__ == "__main__":
    main()


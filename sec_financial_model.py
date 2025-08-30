# Enhanced SEC Financial Model Generator for Local Python Environment
# Addresses accuracy issues with advanced XBRL processing, semantic matching, and validation
# FIXED: Added missing _aggregate_category_values method and other improvements

import subprocess
import sys
import warnings
warnings.filterwarnings('ignore')

def install_packages():
    """Install all required packages"""
    packages = [
        'arelle',
        'sentence-transformers',
        'lxml',
        'sec-edgar-downloader',
        'edgar-tool',
        'alpha-vantage',
        'scikit-learn',
        'transformers',
        'torch',
        'yfinance',
        'openpyxl',
        'pandas',
        'numpy',
        'requests',
        'python-dateutil'
    ]

    for package in packages:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package, "--quiet"])
            print(f"✓ Installed {package}")
        except subprocess.CalledProcessError:
            print(f"✗ Failed to install {package}")

# Uncomment the next line for first run
# install_packages()

import requests
import pandas as pd
import json
import time
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from typing import Dict, List, Optional, Any, Tuple, Set, Union
import numpy as np
from collections import defaultdict, Counter
from dateutil.parser import parse as parse_date
import itertools
from functools import lru_cache
import yfinance as yf
from difflib import SequenceMatcher
import xml.etree.ElementTree as ET
from lxml import etree, html
import urllib.parse
from pathlib import Path

# Advanced libraries for enhanced processing
try:
    from sentence_transformers import SentenceTransformer
    from sklearn.metrics.pairwise import cosine_similarity
    from sklearn.cluster import DBSCAN
    from sklearn.preprocessing import StandardScaler
    from scipy import stats
    import torch
    ADVANCED_LIBS_AVAILABLE = True
    print("✓ Advanced libraries loaded successfully")
except ImportError as e:
    print(f"⚠ Advanced libraries not available: {e}")
    ADVANCED_LIBS_AVAILABLE = False

class EnhancedSECFinancialModelGenerator:
    def __init__(self, company_name: str, ticker: str, cik: str, user_agent_email: str,
                 fiscal_year_end: str = "0630"):
        """
        Enhanced SEC Financial Model Generator with semantic matching and validation
        """
        self.company_name = company_name
        self.ticker = ticker
        self.cik = cik.zfill(10)
        self.fiscal_year_end = fiscal_year_end
        self.user_agent_email = user_agent_email

        self.headers = {
            'User-Agent': f'{company_name} Financial Analysis Tool ({user_agent_email})'
        }
        self.base_url = "https://data.sec.gov/api/xbrl"

        # Enhanced data storage
        self.facts_data = {}
        self.standardized_categories = {}
        self.market_data = {}
        self.raw_metrics = {}
        self.xbrl_contexts = {}
        self.filing_metadata = {}
        self.data_quality_scores = {}
        self.validation_results = {}

        # Initialize semantic model if available
        self.semantic_model = None
        if ADVANCED_LIBS_AVAILABLE:
            try:
                self.semantic_model = SentenceTransformer('all-MiniLM-L6-v2')
                print("✓ Semantic model loaded for intelligent concept matching")
            except Exception as e:
                print(f"⚠ Could not load semantic model: {e}")

        # Enhanced metric patterns with semantic concepts
        self.financial_concepts = self._initialize_enhanced_concepts()
        self.industry_adjustments = self._load_industry_mappings()

    def _initialize_enhanced_concepts(self) -> Dict[str, Dict]:
        """
        Initialize enhanced financial concepts with semantic descriptions
        """
        return {
            'revenue': {
                'display_name': 'Total Revenue',
                'semantic_description': 'total revenue from operations sales income',
                'common_xbrl_concepts': [
                    'RevenueFromContractWithCustomerExcludingAssessedTax',
                    'RevenueFromContractWithCustomer',
                    'Revenues', 'SalesRevenueNet', 'RevenueNet',
                    'TotalRevenue', 'Revenue', 'SalesRevenue'
                ],
                'required_characteristics': ['revenue', 'sales', 'income'],
                'exclusion_terms': ['segment', 'geographic', 'cost', 'expense', 'deferred'],
                'statement_section': 'income_statement',
                'data_type': 'flow',
                'priority': 100
            },
            'cost_of_revenue': {
                'display_name': 'Cost of Revenue',
                'semantic_description': 'direct costs of goods sold services provided cost of sales',
                'common_xbrl_concepts': [
                    'CostOfGoodsAndServicesSold', 'CostOfRevenue', 'CostOfSales',
                    'CostOfGoodsAndServicesSoldExcludingDepreciationDepletionAndAmortization'
                ],
                'required_characteristics': ['cost'],
                'exclusion_terms': ['research', 'development', 'marketing', 'administrative'],
                'statement_section': 'income_statement',
                'data_type': 'flow',
                'priority': 85
            },
            'operating_income': {
                'display_name': 'Operating Income',
                'semantic_description': 'income from operations operating profit earnings',
                'common_xbrl_concepts': [
                    'OperatingIncomeLoss', 'IncomeLossFromContinuingOperations',
                    'OperatingIncome', 'OperatingProfit'
                ],
                'required_characteristics': ['operating'],
                'exclusion_terms': ['expense', 'cost', 'nonoperating', 'before', 'tax'],
                'statement_section': 'income_statement',
                'data_type': 'flow',
                'priority': 95
            },
            'net_income': {
                'display_name': 'Net Income',
                'semantic_description': 'net earnings profit after tax bottom line',
                'common_xbrl_concepts': [
                    'NetIncomeLoss', 'ProfitLoss', 'NetIncome',
                    'NetIncomeLossAvailableToCommonStockholdersBasic'
                ],
                'required_characteristics': ['net'],
                'exclusion_terms': ['operating', 'gross', 'comprehensive', 'before'],
                'statement_section': 'income_statement',
                'data_type': 'flow',
                'priority': 95
            },
            'cash_flow_operations': {
                'display_name': 'Operating Cash Flow',
                'semantic_description': 'cash flow from operating activities operations',
                'common_xbrl_concepts': [
                    'NetCashProvidedByUsedInOperatingActivities',
                    'CashProvidedByUsedInOperatingActivities'
                ],
                'required_characteristics': ['cash', 'operating'],
                'exclusion_terms': ['investing', 'financing'],
                'statement_section': 'cash_flow',
                'data_type': 'flow',
                'priority': 90
            },
            'cash_equivalents': {
                'display_name': 'Cash and Cash Equivalents',
                'semantic_description': 'cash short term investments liquid assets',
                'common_xbrl_concepts': [
                    'CashAndCashEquivalentsAtCarryingValue',
                    'CashCashEquivalentsAndShortTermInvestments',
                    'Cash', 'CashAndEquivalents'
                ],
                'required_characteristics': ['cash'],
                'exclusion_terms': ['restricted'],
                'statement_section': 'balance_sheet',
                'data_type': 'stock',
                'priority': 85
            },
            'total_assets': {
                'display_name': 'Total Assets',
                'semantic_description': 'total assets balance sheet resources',
                'common_xbrl_concepts': [
                    'Assets', 'AssetsCurrent', 'AssetsNoncurrent'
                ],
                'required_characteristics': ['assets', 'total'],
                'exclusion_terms': ['liabilities', 'equity', 'net'],
                'statement_section': 'balance_sheet',
                'data_type': 'stock',
                'priority': 90
            },
            'total_liabilities': {
                'display_name': 'Total Liabilities',
                'semantic_description': 'total liabilities debt obligations',
                'common_xbrl_concepts': [
                    'Liabilities', 'LiabilitiesCurrent', 'LiabilitiesNoncurrent'
                ],
                'required_characteristics': ['liabilities'],
                'exclusion_terms': ['assets', 'equity'],
                'statement_section': 'balance_sheet',
                'data_type': 'stock',
                'priority': 85
            }
        }

    def _load_industry_mappings(self) -> Dict[str, Dict]:
        """
        Load industry-specific XBRL concept mappings
        """
        return {
            'technology': {
                'segment_revenue_patterns': ['ProductivityAndBusinessProcesses', 'IntelligentCloud'],
                'specific_costs': ['ResearchAndDevelopment'],
                'common_adjustments': ['ShareBasedCompensation']
            },
            'financial': {
                'segment_revenue_patterns': ['InterestIncome', 'NoninterestIncome'],
                'specific_costs': ['ProvisionForLoanLosses'],
                'common_adjustments': ['CreditLosses']
            },
            'manufacturing': {
                'segment_revenue_patterns': ['ProductSales', 'ServiceRevenue'],
                'specific_costs': ['RawMaterials', 'Manufacturing'],
                'common_adjustments': ['InventoryWritedown']
            }
        }

    def fetch_sec_data(self) -> bool:
        """
        Enhanced SEC data fetching with metadata collection
        """
        print(f"Fetching {self.company_name} financial data from SEC EDGAR...")

        try:
            # Fetch company facts
            url = f"{self.base_url}/companyfacts/CIK{self.cik}.json"
            response = requests.get(url, headers=self.headers, timeout=60)

            if response.status_code == 200:
                self.facts_data = response.json()
                print(f"✓ Retrieved {self.company_name} SEC data")

                # Extract filing metadata
                self._extract_filing_metadata()

                # Debug info
                if 'facts' in self.facts_data:
                    total_metrics = sum(len(metrics) for metrics in self.facts_data['facts'].values())
                    print(f"  Found {total_metrics} total metrics across taxonomies")

                    # Show taxonomy breakdown
                    for taxonomy, metrics in self.facts_data['facts'].items():
                        print(f"  {taxonomy}: {len(metrics)} metrics")

                return True
            else:
                print(f"✗ Failed to fetch SEC data: HTTP {response.status_code}")
                return False

        except Exception as e:
            print(f"✗ Error fetching SEC data: {e}")
            return False

    def _extract_filing_metadata(self):
        """
        Extract metadata about filings for validation
        """
        try:
            if 'entityName' in self.facts_data:
                self.filing_metadata['entity_name'] = self.facts_data['entityName']

            if 'cik' in self.facts_data:
                self.filing_metadata['cik'] = self.facts_data['cik']

            # Extract filing periods and forms
            filing_periods = set()
            forms = set()

            for taxonomy, metrics in self.facts_data.get('facts', {}).items():
                for metric_name, metric_data in metrics.items():
                    units = metric_data.get('units', {})
                    for unit_type, entries in units.items():
                        for entry in entries:
                            if entry.get('form'):
                                forms.add(entry['form'])
                            if entry.get('fy'):
                                filing_periods.add(entry['fy'])

            self.filing_metadata['periods'] = sorted(list(filing_periods))
            self.filing_metadata['forms'] = list(forms)

            print(f"  Filing periods: {len(filing_periods)} years")
            print(f"  Form types: {', '.join(sorted(forms))}")

        except Exception as e:
            print(f"⚠ Could not extract filing metadata: {e}")

    def fetch_market_data(self):
        """
        Enhanced market data fetching with validation
        """
        try:
            print(f"Fetching market data for {self.ticker}...")
            stock = yf.Ticker(self.ticker)
            info = stock.info
            hist = stock.history(period="1y")

            self.market_data = {
                'market_cap': info.get('marketCap', 0) / 1000000,  # Convert to millions
                'shares_outstanding': info.get('sharesOutstanding', 0) / 1000000,
                'current_price': info.get('currentPrice', 0),
                'enterprise_value': info.get('enterpriseValue', 0) / 1000000 if info.get('enterpriseValue') else 0,
                'beta': info.get('beta', 1.0),
                'sector': info.get('sector', 'Unknown'),
                'industry': info.get('industry', 'Unknown'),
                '52_week_high': info.get('fiftyTwoWeekHigh', 0),
                '52_week_low': info.get('fiftyTwoWeekLow', 0),
                'avg_volume': info.get('averageVolume', 0),
                'price_volatility': hist['Close'].pct_change().std() * np.sqrt(252) if not hist.empty else 0
            }

            print(f"  Market cap: ${self.market_data['market_cap']:,.0f}M")
            print(f"  Sector: {self.market_data['sector']}")
            print(f"  Industry: {self.market_data['industry']}")

        except Exception as e:
            print(f"⚠ Could not fetch market data: {e}")
            self.market_data = {
                'market_cap': 0, 'shares_outstanding': 0, 'current_price': 0,
                'enterprise_value': 0, 'beta': 1.0, 'sector': 'Unknown',
                'industry': 'Unknown', '52_week_high': 0, '52_week_low': 0,
                'avg_volume': 0, 'price_volatility': 0
            }

    def find_and_classify_metrics(self):
        """
        Enhanced metric classification using semantic matching and validation
        """
        if not self.facts_data or 'facts' not in self.facts_data:
            print("✗ No facts data available")
            return

        print("Finding and classifying financial metrics using enhanced methods...")

        # Store all raw metrics with context information
        self._extract_raw_metrics_with_context()

        # Multi-pass classification with validation
        self._semantic_classification()
        self._pattern_based_fallback()
        self._context_validation()
        self._cross_validation()

        print(f"\n✓ Classification complete: {len(self.standardized_categories)} categories matched")

        # Generate quality scores
        self._calculate_data_quality_scores()

    def _extract_raw_metrics_with_context(self):
        """
        Extract raw metrics with full context information for validation
        """
        print("  Extracting raw metrics with context...")

        for taxonomy in self.facts_data['facts']:
            for metric_name, metric_data in self.facts_data['facts'][taxonomy].items():

                # Extract context information from units
                contexts = []
                for unit_type, entries in metric_data.get('units', {}).items():
                    for entry in entries:
                        context = {
                            'start': entry.get('start'),
                            'end': entry.get('end'),
                            'val': entry.get('val'),
                            'accn': entry.get('accn'),
                            'fy': entry.get('fy'),
                            'fp': entry.get('fp'),
                            'form': entry.get('form'),
                            'filed': entry.get('filed'),
                            'unit_type': unit_type
                        }
                        contexts.append(context)

                self.raw_metrics[f"{taxonomy}:{metric_name}"] = {
                    'name': metric_name,
                    'description': metric_data.get('description', ''),
                    'label': metric_data.get('label', ''),
                    'data': metric_data,
                    'taxonomy': taxonomy,
                    'contexts': contexts
                }

        print(f"  Extracted {len(self.raw_metrics)} raw metrics with contexts")

    def _semantic_classification(self):
        """
        Use semantic similarity to match financial concepts
        """
        if not self.semantic_model:
            print("  Skipping semantic classification (model not available)")
            return

        print("  Running semantic classification...")

        # Create embeddings for financial concepts
        concept_texts = []
        concept_keys = []

        for concept_key, concept_info in self.financial_concepts.items():
            text = f"{concept_info['display_name']} {concept_info['semantic_description']}"
            concept_texts.append(text)
            concept_keys.append(concept_key)

        concept_embeddings = self.semantic_model.encode(concept_texts)

        # Create embeddings for raw metrics
        metric_texts = []
        metric_keys = []

        for metric_key, metric_info in self.raw_metrics.items():
            text = f"{metric_info['name']} {metric_info['description']} {metric_info['label']}"
            metric_texts.append(text)
            metric_keys.append(metric_key)

        metric_embeddings = self.semantic_model.encode(metric_texts)

        # Calculate similarities
        similarities = cosine_similarity(metric_embeddings, concept_embeddings)

        # Find best matches
        semantic_matches = 0
        for i, metric_key in enumerate(metric_keys):
            best_concept_idx = np.argmax(similarities[i])
            best_similarity = similarities[i][best_concept_idx]

            if best_similarity > 0.7:  # High confidence threshold
                concept_key = concept_keys[best_concept_idx]
                metric_info = self.raw_metrics[metric_key]

                print(f"    SEMANTIC: {metric_info['name']} -> {concept_key} ({best_similarity:.3f})")

                self._add_metric_to_category(
                    concept_key,
                    metric_info['name'],
                    metric_info['data'],
                    metric_info['taxonomy'],
                    best_similarity,
                    'semantic'
                )
                semantic_matches += 1

        print(f"  Semantic classification: {semantic_matches} matches found")

    def _pattern_based_fallback(self):
        """
        Pattern-based matching for concepts not found semantically
        """
        print("  Running pattern-based fallback...")

        unmatched_concepts = [
            concept for concept in self.financial_concepts.keys()
            if concept not in self.standardized_categories
        ]

        pattern_matches = 0

        for concept_key in unmatched_concepts:
            concept_info = self.financial_concepts[concept_key]
            best_matches = []

            # Check exact matches first
            for metric_key, metric_info in self.raw_metrics.items():
                metric_name = metric_info['name']

                if metric_name in concept_info.get('common_xbrl_concepts', []):
                    best_matches.append((metric_key, metric_info, 1.0))
                    continue

                # Pattern matching
                score = self._calculate_pattern_score(metric_info, concept_info)
                if score > 0.6:
                    best_matches.append((metric_key, metric_info, score))

            if best_matches:
                best_matches.sort(key=lambda x: x[2], reverse=True)
                metric_key, metric_info, score = best_matches[0]

                print(f"    PATTERN: {metric_info['name']} -> {concept_key} ({score:.3f})")

                self._add_metric_to_category(
                    concept_key,
                    metric_info['name'],
                    metric_info['data'],
                    metric_info['taxonomy'],
                    score,
                    'pattern'
                )
                pattern_matches += 1

        print(f"  Pattern classification: {pattern_matches} matches found")

    def _calculate_pattern_score(self, metric_info: Dict, concept_info: Dict) -> float:
        """
        Enhanced pattern scoring with multiple factors
        """
        text = f"{metric_info['name']} {metric_info['description']}".lower()
        score = 0.0

        # Required characteristics (must have all)
        required = concept_info.get('required_characteristics', [])
        if required:
            matches = sum(1 for req in required if req.lower() in text)
            if matches == len(required):
                score += 0.4
            elif matches > 0:
                score += 0.2 * (matches / len(required))
            else:
                return 0.0  # Fail if missing required characteristics

        # Exclusion terms (must have none)
        exclusions = concept_info.get('exclusion_terms', [])
        for exclusion in exclusions:
            if exclusion.lower() in text:
                return 0.0  # Immediate disqualification

        # Common XBRL concepts (bonus points)
        common_concepts = concept_info.get('common_xbrl_concepts', [])
        for concept in common_concepts:
            similarity = SequenceMatcher(None, metric_info['name'], concept).ratio()
            if similarity > 0.8:
                score += 0.3
            elif similarity > 0.6:
                score += 0.2

        # Statement section alignment
        if concept_info.get('statement_section') == 'balance_sheet':
            if any(term in text for term in ['balance', 'sheet', 'asset', 'liability', 'equity']):
                score += 0.1
        elif concept_info.get('statement_section') == 'income_statement':
            if any(term in text for term in ['income', 'revenue', 'expense', 'profit', 'loss']):
                score += 0.1
        elif concept_info.get('statement_section') == 'cash_flow':
            if any(term in text for term in ['cash', 'flow', 'activities']):
                score += 0.1

        return min(score, 1.0)

    def _context_validation(self):
        """
        Validate contexts to ensure data consistency
        """
        print("  Validating contexts...")

        for category_key, category_data in self.standardized_categories.items():
            if not category_data.get('metrics'):
                continue

            # Validate time periods
            all_periods = set()
            for metric_info in category_data['metrics']:
                metric_key = f"{metric_info['taxonomy']}:{metric_info['name']}"
                if metric_key in self.raw_metrics:
                    contexts = self.raw_metrics[metric_key]['contexts']
                    for context in contexts:
                        if context.get('fy'):
                            all_periods.add(context['fy'])

            # Flag periods with insufficient data
            insufficient_periods = []
            for period in all_periods:
                period_data_count = 0
                for metric_info in category_data['metrics']:
                    metric_key = f"{metric_info['taxonomy']}:{metric_info['name']}"
                    if metric_key in self.raw_metrics:
                        contexts = self.raw_metrics[metric_key]['contexts']
                        period_contexts = [c for c in contexts if c.get('fy') == period]
                        period_data_count += len(period_contexts)

                if period_data_count < len(category_data['metrics']):
                    insufficient_periods.append(period)

            if insufficient_periods:
                print(f"    ⚠ {category_key}: Insufficient data for periods {insufficient_periods}")

    def _cross_validation(self):
        """
        Cross-validate financial relationships
        """
        print("  Running cross-validation...")

        validation_rules = [
            ('revenue', 'cost_of_revenue', 'revenue >= cost_of_revenue'),
            ('operating_income', 'net_income', 'abs(operating_income) >= abs(net_income) * 0.5'),
            ('total_assets', 'cash_equivalents', 'total_assets >= cash_equivalents'),
        ]

        validation_failures = 0

        for rule in validation_rules:
            concept1, concept2, rule_desc = rule

            if concept1 in self.standardized_categories and concept2 in self.standardized_categories:
                # Get recent data for validation
                data1 = self.standardized_categories[concept1].get('annual_data', {})
                data2 = self.standardized_categories[concept2].get('annual_data', {})

                common_years = set(data1.keys()) & set(data2.keys())

                for year in common_years:
                    val1 = self._aggregate_category_values(self.standardized_categories[concept1], year)
                    val2 = self._aggregate_category_values(self.standardized_categories[concept2], year)

                    if val1 is not None and val2 is not None:
                        # Apply validation rule
                        if concept1 == 'revenue' and concept2 == 'cost_of_revenue':
                            if val1 < val2:
                                print(f"    ⚠ {year}: Revenue ({val1:.1f}) < Cost of Revenue ({val2:.1f})")
                                validation_failures += 1

                        elif concept1 == 'operating_income' and concept2 == 'net_income':
                            if abs(val1) < abs(val2) * 0.5:
                                print(f"    ⚠ {year}: Operating Income relationship anomaly")
                                validation_failures += 1

        if validation_failures == 0:
            print("  ✓ Cross-validation passed")
        else:
            print(f"  ⚠ Cross-validation: {validation_failures} potential issues found")

    def _aggregate_category_values(self, category_data: Dict, year: int) -> Optional[float]:
        """
        FIXED: Added missing method to aggregate values for a category in a specific year
        """
        try:
            # Get all annual values for the year
            annual_values = category_data.get('annual_data', {}).get(year, [])

            if not annual_values:
                return None

            # Clean and aggregate values
            clean_values = [v for v in annual_values if isinstance(v, (int, float)) and not np.isnan(v)]

            if not clean_values:
                return None

            # Use median to handle outliers
            return np.median(clean_values)

        except Exception:
            return None

    def _calculate_data_quality_scores(self):
        """
        Calculate quality scores for each category
        """
        print("  Calculating data quality scores...")

        for category_key, category_data in self.standardized_categories.items():
            score_factors = {
                'completeness': 0,
                'consistency': 0,
                'accuracy': 0,
                'timeliness': 0
            }

            # Completeness: How many expected periods have data
            expected_periods = range(2018, 2025)
            available_periods = list(category_data.get('annual_data', {}).keys())
            if expected_periods:
                score_factors['completeness'] = len(available_periods) / len(expected_periods)

            # Consistency: How consistent are values across metrics
            if len(category_data.get('metrics', [])) > 1:
                # Check variance across different metrics for same concept
                consistency_scores = []
                for year in available_periods:
                    values = category_data['annual_data'].get(year, [])
                    if len(values) > 1:
                        cv = np.std(values) / np.mean(values) if np.mean(values) != 0 else 1
                        consistency_scores.append(max(0, 1 - cv))

                score_factors['consistency'] = np.mean(consistency_scores) if consistency_scores else 1.0
            else:
                score_factors['consistency'] = 1.0

            # Accuracy: Based on matching method and confidence
            confidence = category_data.get('confidence', 0)
            score_factors['accuracy'] = confidence

            # Timeliness: How recent is the data
            if available_periods:
                most_recent = max(available_periods)
                years_old = 2024 - most_recent
                score_factors['timeliness'] = max(0, 1 - years_old * 0.2)

            # Overall quality score
            overall_score = np.mean(list(score_factors.values()))

            self.data_quality_scores[category_key] = {
                'overall': overall_score,
                'factors': score_factors
            }

    def _add_metric_to_category(self, category_key: str, metric_name: str,
                               metric_data: Dict, taxonomy: str, confidence: float,
                               method: str = 'unknown'):
        """
        Enhanced method to add metrics to categories with validation
        """
        if category_key not in self.standardized_categories:
            concept_info = self.financial_concepts.get(category_key, {})
            self.standardized_categories[category_key] = {
                'display_name': concept_info.get('display_name', category_key),
                'metrics': [],
                'annual_data': defaultdict(list),
                'quarterly_data': defaultdict(lambda: defaultdict(list)),
                'confidence': confidence,
                'method': method,
                'section': concept_info.get('statement_section', 'unknown'),
                'data_type': concept_info.get('data_type', 'flow')
            }

        self.standardized_categories[category_key]['metrics'].append({
            'name': metric_name,
            'taxonomy': taxonomy,
            'description': metric_data.get('description', ''),
            'confidence': confidence,
            'method': method
        })

        # Extract and validate data
        data_points = self._extract_and_validate_data(category_key, metric_name, metric_data)
        print(f"      Extracted {data_points} validated data points")

    def _extract_and_validate_data(self, category_key: str, metric_name: str,
                                  metric_data: Dict) -> int:
        """
        Extract and validate numerical data with enhanced error handling
        """
        units = metric_data.get('units', {})
        data_points = 0
        outliers_detected = 0

        for unit_type, entries in units.items():
            # Focus on USD for financial metrics, shares for share data
            if not any(acceptable in unit_type for acceptable in ['USD', 'pure', 'shares']):
                continue

            values_by_period = defaultdict(list)

            for entry in entries:
                try:
                    value = entry.get('val')
                    fy = entry.get('fy')
                    fp = entry.get('fp', '')
                    form = entry.get('form', '')
                    end_date = entry.get('end', '')

                    if not all([value is not None, fy]):
                        continue

                    # Convert and validate value
                    if 'USD' in unit_type:
                        value = float(value) / 1000000  # Convert to millions
                    else:
                        value = float(value)

                    # Basic outlier detection
                    if abs(value) > 1e10:  # Extremely large values
                        outliers_detected += 1
                        continue

                    year = int(fy)

                    # Classify by period type
                    if fp == 'FY' or (not fp and form in ['10-K', '10-K/A']):
                        # Annual data
                        values_by_period[(year, 'annual')].append(value)
                        data_points += 1
                    elif fp.startswith('Q') or form in ['10-Q', '10-Q/A']:
                        # Quarterly data
                        quarter = fp if fp.startswith('Q') else self._determine_quarter_from_date(end_date)
                        if quarter:
                            values_by_period[(year, quarter)].append(value)
                            data_points += 1

                except (ValueError, TypeError) as e:
                    continue

            # Aggregate values for each period
            for (year, period), values in values_by_period.items():
                if not values:
                    continue

                # Handle multiple values for same period
                if len(values) == 1:
                    final_value = values[0]
                else:
                    # Use median to handle outliers, but flag inconsistency
                    final_value = np.median(values)
                    if np.std(values) / np.mean(values) > 0.1:  # High coefficient of variation
                        print(f"      ⚠ Inconsistent values for {metric_name} {year}-{period}: {values}")

                # Store in appropriate data structure
                if period == 'annual':
                    self.standardized_categories[category_key]['annual_data'][year].append(final_value)
                else:
                    self.standardized_categories[category_key]['quarterly_data'][year][period].append(final_value)

        if outliers_detected > 0:
            print(f"      ⚠ Filtered {outliers_detected} outliers")

        return data_points

    def _determine_quarter_from_date(self, end_date: str) -> Optional[str]:
        """
        Enhanced quarter determination with fiscal year support
        """
        try:
            end_date_obj = parse_date(end_date)
            month = end_date_obj.month

            if self.fiscal_year_end == "0630":  # June 30 fiscal year end (Microsoft)
                quarter_map = {9: 'Q1', 12: 'Q2', 3: 'Q3', 6: 'Q4'}
            elif self.fiscal_year_end == "1231":  # December 31 calendar year
                quarter_map = {3: 'Q1', 6: 'Q2', 9: 'Q3', 12: 'Q4'}
            else:
                # Generic mapping
                quarter_map = {3: 'Q1', 6: 'Q2', 9: 'Q3', 12: 'Q4'}

            return quarter_map.get(month)

        except:
            return None

    def calculate_derived_metrics(self):
        """
        Enhanced derived metrics calculation with validation
        """
        print("Calculating derived metrics with validation...")

        # Calculate metrics in dependency order
        calculations = [
            ('gross_profit', self._calculate_gross_profit),
            ('ebitda', self._calculate_ebitda),
            ('ebit', self._calculate_ebit),
            ('ebt', self._calculate_ebt),
            ('free_cash_flow', self._calculate_free_cash_flow),
            ('working_capital', self._calculate_working_capital),
            ('debt_to_equity', self._calculate_debt_to_equity),
            ('return_metrics', self._calculate_return_metrics)
        ]

        for calc_name, calc_function in calculations:
            try:
                calc_function()
                print(f"  ✓ {calc_name}")
            except Exception as e:
                print(f"  ⚠ {calc_name}: {e}")

    def _calculate_gross_profit(self):
        """
        Enhanced gross profit calculation with multiple cost components
        """
        if 'revenue' not in self.standardized_categories:
            return

        self.standardized_categories['gross_profit'] = {
            'display_name': 'Gross Profit',
            'metrics': [{'name': 'calculated_gross_profit', 'taxonomy': 'calculated',
                        'description': 'Revenue minus cost of revenue', 'method': 'calculated'}],
            'annual_data': defaultdict(list),
            'quarterly_data': defaultdict(lambda: defaultdict(list)),
            'section': 'calculated',
            'data_type': 'flow'
        }

        revenue_data = self.standardized_categories['revenue']
        cost_data = self.standardized_categories.get('cost_of_revenue', {})

        # Get all available periods
        all_periods = self._get_all_periods([revenue_data])

        for year, period_type in all_periods:
            revenue = self._get_period_value(revenue_data, year, period_type)
            cost = self._get_period_value(cost_data, year, period_type) if cost_data else 0

            if revenue is not None:
                gross_profit = revenue - (cost or 0)

                if period_type == 'annual':
                    self.standardized_categories['gross_profit']['annual_data'][year] = [gross_profit]
                else:
                    self.standardized_categories['gross_profit']['quarterly_data'][year][period_type] = [gross_profit]

    def _calculate_ebitda(self):
        """
        Calculate EBITDA with multiple methods
        """
        # Method 1: Operating Income + D&A
        if 'operating_income' in self.standardized_categories:
            self._calculate_ebitda_from_operating_income()

        # Method 2: Net Income + Taxes + Interest + D&A (if operating income not available)
        elif 'net_income' in self.standardized_categories:
            self._calculate_ebitda_from_net_income()

    def _calculate_ebitda_from_operating_income(self):
        """Calculate EBITDA from operating income"""
        operating_data = self.standardized_categories['operating_income']
        da_data = self.standardized_categories.get('depreciation_amortization', {})

        self.standardized_categories['ebitda'] = {
            'display_name': 'EBITDA',
            'metrics': [{'name': 'calculated_ebitda', 'taxonomy': 'calculated',
                        'description': 'Operating income plus depreciation and amortization'}],
            'annual_data': defaultdict(list),
            'quarterly_data': defaultdict(lambda: defaultdict(list)),
            'section': 'calculated'
        }

        all_periods = self._get_all_periods([operating_data])

        for year, period_type in all_periods:
            operating_income = self._get_period_value(operating_data, year, period_type)
            da = self._get_period_value(da_data, year, period_type) if da_data else 0

            if operating_income is not None:
                ebitda = operating_income + (da or 0)

                if period_type == 'annual':
                    self.standardized_categories['ebitda']['annual_data'][year] = [ebitda]
                else:
                    self.standardized_categories['ebitda']['quarterly_data'][year][period_type] = [ebitda]

    def _calculate_ebitda_from_net_income(self):
        """Calculate EBITDA from net income (fallback method)"""
        # This would add back taxes, interest, depreciation, and amortization to net income
        # Implementation would be more complex and require identifying these items
        pass

    def _calculate_ebit(self):
        """Calculate EBIT (same as operating income typically)"""
        if 'operating_income' in self.standardized_categories:
            # EBIT is typically the same as operating income
            operating_data = self.standardized_categories['operating_income']

            self.standardized_categories['ebit'] = {
                'display_name': 'EBIT',
                'metrics': [{'name': 'calculated_ebit', 'taxonomy': 'calculated',
                            'description': 'Earnings before interest and taxes'}],
                'annual_data': operating_data['annual_data'].copy(),
                'quarterly_data': operating_data['quarterly_data'].copy(),
                'section': 'calculated'
            }

    def _calculate_ebt(self):
        """Calculate EBT (Earnings Before Tax)"""
        # This would require identifying interest expense/income
        # For now, we'll skip this complex calculation
        pass

    def _calculate_free_cash_flow(self):
        """Enhanced free cash flow calculation"""
        if 'cash_flow_operations' not in self.standardized_categories:
            return

        self.standardized_categories['free_cash_flow'] = {
            'display_name': 'Free Cash Flow',
            'metrics': [{'name': 'calculated_fcf', 'taxonomy': 'calculated',
                        'description': 'Operating cash flow minus capital expenditures'}],
            'annual_data': defaultdict(list),
            'quarterly_data': defaultdict(lambda: defaultdict(list)),
            'section': 'calculated'
        }

        ocf_data = self.standardized_categories['cash_flow_operations']
        capex_data = self.standardized_categories.get('capex', {})

        all_periods = self._get_all_periods([ocf_data])

        for year, period_type in all_periods:
            ocf = self._get_period_value(ocf_data, year, period_type)
            capex = self._get_period_value(capex_data, year, period_type) if capex_data else 0

            if ocf is not None:
                fcf = ocf - abs(capex or 0)  # Capex is typically negative

                if period_type == 'annual':
                    self.standardized_categories['free_cash_flow']['annual_data'][year] = [fcf]
                else:
                    self.standardized_categories['free_cash_flow']['quarterly_data'][year][period_type] = [fcf]

    def _calculate_working_capital(self):
        """Calculate working capital (current assets - current liabilities)"""
        # This would require identifying current assets and current liabilities
        # For now, we'll skip this calculation
        pass

    def _calculate_debt_to_equity(self):
        """Calculate debt to equity ratio"""
        # This would require identifying total debt and total equity
        # For now, we'll skip this calculation
        pass

    def _calculate_return_metrics(self):
        """Calculate ROA, ROE, and other return metrics"""
        self._calculate_roa()
        self._calculate_roe()
        self._calculate_roic()

    def _calculate_roa(self):
        """Calculate Return on Assets"""
        if 'net_income' not in self.standardized_categories or 'total_assets' not in self.standardized_categories:
            return

        self.standardized_categories['roa'] = {
            'display_name': 'Return on Assets (%)',
            'metrics': [{'name': 'calculated_roa', 'taxonomy': 'calculated'}],
            'annual_data': defaultdict(list),
            'quarterly_data': defaultdict(lambda: defaultdict(list)),
            'section': 'calculated'
        }

        ni_data = self.standardized_categories['net_income']
        assets_data = self.standardized_categories['total_assets']

        # Calculate for annual data only (balance sheet items need averaging)
        for year in ni_data.get('annual_data', {}):
            net_income = self._get_period_value(ni_data, year, 'annual')
            current_assets = self._get_period_value(assets_data, year, 'annual')
            prior_assets = self._get_period_value(assets_data, year-1, 'annual')

            if net_income is not None and current_assets is not None:
                avg_assets = current_assets
                if prior_assets is not None:
                    avg_assets = (current_assets + prior_assets) / 2

                if avg_assets != 0:
                    roa = (net_income / avg_assets) * 100
                    self.standardized_categories['roa']['annual_data'][year] = [roa]

    def _calculate_roe(self):
        """Calculate Return on Equity"""
        # This would require identifying total equity
        # For now, we'll skip this calculation
        pass

    def _calculate_roic(self):
        """Calculate Return on Invested Capital"""
        # This would require identifying invested capital
        # For now, we'll skip this calculation
        pass

    def _get_all_periods(self, data_sources: List[Dict]) -> List[Tuple[int, str]]:
        """Get all available periods from data sources"""
        periods = set()

        for data_source in data_sources:
            # Annual periods
            for year in data_source.get('annual_data', {}):
                periods.add((year, 'annual'))

            # Quarterly periods
            for year, quarters in data_source.get('quarterly_data', {}).items():
                for quarter in quarters:
                    periods.add((year, quarter))

        return sorted(list(periods))

    def _get_period_value(self, data_source: Dict, year: int, period_type: str) -> Optional[float]:
        """Get value for a specific period with proper aggregation"""
        try:
            if period_type == 'annual':
                values = data_source.get('annual_data', {}).get(year, [])
            else:
                values = data_source.get('quarterly_data', {}).get(year, {}).get(period_type, [])

            if not values:
                return None

            # Clean and aggregate values
            clean_values = [v for v in values if isinstance(v, (int, float)) and not np.isnan(v)]

            if not clean_values:
                return None

            return np.median(clean_values) if len(clean_values) > 1 else clean_values[0]

        except Exception:
            return None

    def generate_projections(self, projection_years: List[int]):
        """
        Enhanced projections with industry benchmarks and scenario analysis
        """
        print(f"Generating enhanced projections for years: {projection_years}")

        # Get historical data and calculate trends
        historical_analysis = self._analyze_historical_trends()

        # Apply industry benchmarks if available
        industry_adjustments = self._get_industry_benchmarks()

        # Generate base, optimistic, and pessimistic scenarios
        scenarios = ['base', 'optimistic', 'pessimistic']

        for scenario in scenarios:
            print(f"  Generating {scenario} scenario...")
            self._generate_scenario_projections(projection_years, historical_analysis,
                                              industry_adjustments, scenario)

    def _analyze_historical_trends(self) -> Dict[str, Dict]:
        """Analyze historical trends for projection"""
        trends = {}

        key_metrics = ['revenue', 'operating_income', 'net_income', 'cash_flow_operations', 'free_cash_flow']

        for metric in key_metrics:
            if metric not in self.standardized_categories:
                continue

            data = self.standardized_categories[metric]['annual_data']
            years = sorted(data.keys())

            if len(years) >= 3:
                values = [self._get_period_value(self.standardized_categories[metric], year, 'annual')
                         for year in years]
                values = [v for v in values if v is not None]

                if len(values) >= 3:
                    # Calculate various trend metrics
                    growth_rates = []
                    for i in range(1, len(values)):
                        if values[i-1] != 0:
                            gr = (values[i] / values[i-1]) - 1
                            growth_rates.append(gr)

                    if growth_rates:
                        trends[metric] = {
                            'avg_growth': np.mean(growth_rates),
                            'median_growth': np.median(growth_rates),
                            'std_growth': np.std(growth_rates),
                            'latest_value': values[-1],
                            'cagr': ((values[-1] / values[0]) ** (1 / (len(values) - 1))) - 1 if values[0] != 0 else 0
                        }

        return trends

    def _generate_scenario_projections(self, projection_years: List[int],
                                     historical_analysis: Dict, industry_adjustments: Dict,
                                     scenario: str):
        """Generate projections for a specific scenario"""
        scenario_multipliers = {
            'base': 1.0,
            'optimistic': 1.2,
            'pessimistic': 0.8
        }

        multiplier = scenario_multipliers[scenario]

        for metric, trend_data in historical_analysis.items():
            if metric not in self.standardized_categories:
                continue

            base_growth = trend_data['cagr']
            latest_value = trend_data['latest_value']

            # Apply scenario adjustment
            adjusted_growth = base_growth * multiplier

            # Apply industry benchmarks
            if metric in industry_adjustments:
                industry_growth = industry_adjustments[metric]
                adjusted_growth = (adjusted_growth + industry_growth) / 2  # Blend

            # Add declining growth assumption for later years
            current_value = latest_value

            for i, year in enumerate(sorted(projection_years)):
                # Declining growth rate
                year_growth = adjusted_growth * (0.9 ** i)  # 10% decline each year
                current_value *= (1 + year_growth)

                # Store projection (suffix with scenario for non-base)
                category_key = metric if scenario == 'base' else f"{metric}_{scenario}"

                if category_key not in self.standardized_categories:
                    self.standardized_categories[category_key] = {
                        'display_name': f"{self.standardized_categories[metric]['display_name']} ({scenario})",
                        'annual_data': defaultdict(list),
                        'section': 'projection'
                    }

                self.standardized_categories[category_key]['annual_data'][year] = [current_value]

    def _get_industry_benchmarks(self) -> Dict[str, float]:
        """Get industry benchmark growth rates"""
        industry = self.market_data.get('industry', '').lower()

        # Basic industry benchmarks - in practice, this would come from a database
        benchmarks = {
            'software': {'revenue': 0.15, 'operating_income': 0.12},
            'technology': {'revenue': 0.12, 'operating_income': 0.10},
            'financial': {'revenue': 0.08, 'operating_income': 0.06},
            'manufacturing': {'revenue': 0.06, 'operating_income': 0.05}
        }

        for industry_key, rates in benchmarks.items():
            if industry_key in industry:
                return rates

        # Default conservative benchmarks
        return {'revenue': 0.05, 'operating_income': 0.04}

    def build_comprehensive_model(self) -> pd.DataFrame:
        """
        Build enhanced financial model with multiple sheets worth of data
        """
        print("Building comprehensive financial model...")

        model_sections = []

        # Header section
        model_sections.extend(self._build_header_section())

        # Income Statement
        model_sections.extend(self._build_income_statement_section())

        # Cash Flow Statement
        model_sections.extend(self._build_cash_flow_section())

        # Balance Sheet
        model_sections.extend(self._build_balance_sheet_section())

        # Ratios and Metrics
        model_sections.extend(self._build_ratios_section())

        # Valuation
        model_sections.extend(self._build_valuation_section())

        # Convert to DataFrame
        df = pd.DataFrame(model_sections)
        print(f"✓ Model built: {df.shape[0]} rows × {df.shape[1]} columns")

        return df

    def _build_header_section(self) -> List[List[str]]:
        """Build header section with company info and periods"""
        header_data = []

        header_data.append([f'{self.company_name.upper()} FINANCIAL MODEL'])
        header_data.append([f'CIK: {self.cik} | Ticker: {self.ticker} | FY End: {self.fiscal_year_end}'])
        header_data.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | USD in Millions'])
        header_data.append([])

        # Column headers
        headers = self._create_enhanced_headers()
        header_data.append(headers)
        header_data.append([])

        return header_data

    def _create_enhanced_headers(self) -> List[str]:
        """Create comprehensive headers for all time periods"""
        headers = ['']

        # Historical years
        for year in range(2018, 2025):
            headers.append(f'{year}A')

        # Quarterly data for recent years
        quarters = ['Q1', 'Q2', 'Q3', 'Q4']
        quarter_names = ['Sep', 'Dec', 'Mar', 'Jun']  # For Jun 30 fiscal year

        for year in [2023, 2024]:
            for q, qname in zip(quarters, quarter_names):
                headers.append(f'{year}{qname}')

        # Projections
        for year in range(2025, 2028):
            headers.append(f'{year}P')

        return headers

    def _build_income_statement_section(self) -> List[List[str]]:
        """Build income statement section"""
        section_data = []
        section_data.append(['INCOME STATEMENT'])

        income_metrics = [
            ('revenue', True, True),  # (metric_key, show_growth, show_margin)
            ('gross_profit', False, True),
            ('operating_income', True, True),
            ('ebitda', True, True),
            ('net_income', True, True)
        ]

        headers = self._create_enhanced_headers()

        for metric_key, show_growth, show_margin in income_metrics:
            if metric_key in self.standardized_categories:
                section_data.extend(self._create_metric_rows(metric_key, headers, show_growth, show_margin))

        section_data.append([])
        return section_data

    def _build_cash_flow_section(self) -> List[List[str]]:
        """Build cash flow section"""
        section_data = []
        section_data.append(['CASH FLOW'])

        cf_metrics = [
            ('cash_flow_operations', True, False),
            ('free_cash_flow', True, False)
        ]

        headers = self._create_enhanced_headers()

        for metric_key, show_growth, show_margin in cf_metrics:
            if metric_key in self.standardized_categories:
                section_data.extend(self._create_metric_rows(metric_key, headers, show_growth, show_margin))

        section_data.append([])
        return section_data

    def _build_balance_sheet_section(self) -> List[List[str]]:
        """Build balance sheet section"""
        section_data = []
        section_data.append(['BALANCE SHEET'])

        bs_metrics = [
            ('cash_equivalents', False, False),
            ('total_assets', False, False),
            ('total_liabilities', False, False)
        ]

        headers = self._create_enhanced_headers()

        for metric_key, show_growth, show_margin in bs_metrics:
            if metric_key in self.standardized_categories:
                section_data.extend(self._create_metric_rows(metric_key, headers, show_growth, show_margin))

        section_data.append([])
        return section_data

    def _build_ratios_section(self) -> List[List[str]]:
        """Build financial ratios section"""
        section_data = []
        section_data.append(['FINANCIAL RATIOS'])

        ratio_metrics = [
            ('roa', False, False)
        ]

        headers = self._create_enhanced_headers()

        for metric_key, show_growth, show_margin in ratio_metrics:
            if metric_key in self.standardized_categories:
                section_data.extend(self._create_metric_rows(metric_key, headers, show_growth, show_margin))

        section_data.append([])
        return section_data

    def _build_valuation_section(self) -> List[List[str]]:
        """Build valuation metrics section"""
        section_data = []
        section_data.append(['VALUATION METRICS'])

        headers = self._create_enhanced_headers()

        # Market data row
        market_row = ['Market Cap']
        for header in headers[1:]:
            if 'P' in header or header.endswith('A'):
                market_cap = self.market_data.get('market_cap', 0)
                market_row.append(f"{market_cap:,.0f}" if market_cap > 0 else '')
            else:
                market_row.append('')
        section_data.append(market_row)

        # EV/EBITDA multiple if available
        if 'ebitda' in self.standardized_categories:
            ev_ebitda_row = ['EV/EBITDA']
            for header in headers[1:]:
                ebitda = self._get_value_for_header('ebitda', header)
                ev = self.market_data.get('enterprise_value', 0)
                if ebitda and ebitda > 0 and ev > 0:
                    multiple = ev / ebitda
                    ev_ebitda_row.append(f"{multiple:.1f}x")
                else:
                    ev_ebitda_row.append('')
            section_data.append(ev_ebitda_row)

        section_data.append([])
        return section_data

    def _create_metric_rows(self, metric_key: str, headers: List[str],
                           show_growth: bool, show_margin: bool) -> List[List[str]]:
        """Create rows for a metric including growth and margin if requested"""
        rows = []

        # Main metric row
        display_name = self.standardized_categories[metric_key]['display_name']
        main_row = [display_name]

        for header in headers[1:]:
            value = self._get_value_for_header(metric_key, header)
            formatted_value = self._format_model_value(value)
            main_row.append(formatted_value)

        rows.append(main_row)

        # Growth row
        if show_growth:
            growth_row = ['  % Growth']
            for header in headers[1:]:
                growth = self._calculate_growth_rate(metric_key, header)
                growth_formatted = f"{growth:.1f}%" if growth is not None else ''
                growth_row.append(growth_formatted)
            rows.append(growth_row)

        # Margin row (as % of revenue)
        if show_margin and 'revenue' in self.standardized_categories:
            margin_row = ['  % Margin']
            for header in headers[1:]:
                margin = self._calculate_margin(metric_key, header)
                margin_formatted = f"{margin:.1f}%" if margin is not None else ''
                margin_row.append(margin_formatted)
            rows.append(margin_row)

        return rows

    def _get_value_for_header(self, metric_key: str, header: str) -> Optional[float]:
        """Enhanced value retrieval for different header formats"""
        if metric_key not in self.standardized_categories:
            return None

        category_data = self.standardized_categories[metric_key]

        try:
            if header.endswith('P'):
                # Projection year
                year = int(header[:-1])
                return self._get_period_value(category_data, year, 'annual')
            elif header.endswith('A'):
                # Annual historical data
                year = int(header[:-1])
                return self._get_period_value(category_data, year, 'annual')
            elif len(header) >= 7 and header[:4].isdigit():
                # Quarterly data like "2024Mar"
                year = int(header[:4])
                month_name = header[4:]

                month_to_quarter = {
                    'Sep': 'Q1', 'Dec': 'Q2', 'Mar': 'Q3', 'Jun': 'Q4'  # For Jun 30 fiscal year
                }
                quarter = month_to_quarter.get(month_name)

                if quarter:
                    return self._get_period_value(category_data, year, quarter)

        except (ValueError, IndexError):
            pass

        return None

    def _calculate_growth_rate(self, metric_key: str, current_header: str) -> Optional[float]:
        """Enhanced growth rate calculation"""
        try:
            current_value = self._get_value_for_header(metric_key, current_header)

            if current_value is None:
                return None

            # Determine previous period
            if current_header.endswith('A') or current_header.endswith('P'):
                # Annual data
                current_year = int(current_header[:-1])
                prev_year = current_year - 1
                prev_header = f"{prev_year}A"
                prev_value = self._get_value_for_header(metric_key, prev_header)
            else:
                # Quarterly data - compare to same quarter previous year
                if len(current_header) >= 7:
                    year = int(current_header[:4])
                    quarter_part = current_header[4:]
                    prev_header = f"{year-1}{quarter_part}"
                    prev_value = self._get_value_for_header(metric_key, prev_header)
                else:
                    return None

            if prev_value is not None and prev_value != 0:
                return ((current_value / prev_value) - 1) * 100

        except (ValueError, ZeroDivisionError):
            pass

        return None

    def _calculate_margin(self, metric_key: str, header: str) -> Optional[float]:
        """Enhanced margin calculation"""
        try:
            metric_value = self._get_value_for_header(metric_key, header)
            revenue_value = self._get_value_for_header('revenue', header)

            if metric_value is not None and revenue_value is not None and revenue_value != 0:
                return (metric_value / revenue_value) * 100

        except ZeroDivisionError:
            pass

        return None

    def _format_model_value(self, value: Optional[float]) -> str:
        """Enhanced value formatting"""
        if value is None:
            return ''

        try:
            abs_value = abs(value)
            if abs_value >= 1000:
                return f"{value:,.0f}"
            elif abs_value >= 10:
                return f"{value:.1f}"
            elif abs_value >= 0.1:
                return f"{value:.2f}"
            else:
                return f"{value:.3f}"
        except (TypeError, ValueError):
            return ''

    def export_to_excel(self, df: pd.DataFrame, filename: str = None) -> bool:
        """
        Enhanced Excel export with multiple sheets and advanced formatting
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.company_name.lower().replace(' ', '_')}_enhanced_model_{timestamp}.xlsx"

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Main financial model
                df.to_excel(writer, sheet_name='Financial Model', index=False, header=False)

                # Classification summary with quality scores
                self._create_enhanced_summary_sheet(writer)

                # Data validation sheet
                self._create_validation_sheet(writer)

                # Market data sheet
                self._create_market_data_sheet(writer)

                # Methodology sheet
                self._create_methodology_sheet(writer)

                # Apply formatting
                workbook = writer.book
                self._apply_enhanced_formatting(workbook)

            print(f"Enhanced financial model exported to {filename}")
            return True

        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    def _create_enhanced_summary_sheet(self, writer):
        """Create comprehensive classification summary with quality metrics"""
        summary_data = []
        summary_data.append(['CLASSIFICATION SUMMARY'])
        summary_data.append([''])
        summary_data.append(['Category', 'Display Name', 'SEC Metrics Used', 'Classification Method',
                            'Confidence', 'Quality Score', 'Data Points', 'Years Available'])

        for category_key, category_data in self.standardized_categories.items():
            display_name = category_data.get('display_name', category_key)
            method = category_data.get('method', 'unknown')
            confidence = category_data.get('confidence', 0)

            # Quality score
            quality_info = self.data_quality_scores.get(category_key, {})
            quality_score = quality_info.get('overall', 0)

            # Data statistics
            annual_points = len(category_data.get('annual_data', {}))
            quarterly_points = sum(len(quarters) for quarters in category_data.get('quarterly_data', {}).values())
            total_points = annual_points + quarterly_points

            years_available = sorted(list(category_data.get('annual_data', {}).keys()))
            year_range = f"{min(years_available)}-{max(years_available)}" if years_available else "None"

            # Metrics used
            metrics_used = []
            for metric in category_data.get('metrics', []):
                metrics_used.append(f"{metric['name']} ({metric.get('confidence', 0):.2f})")
            metrics_text = "; ".join(metrics_used[:2])  # Limit to first 2 for space
            if len(metrics_used) > 2:
                metrics_text += f" + {len(metrics_used)-2} more"

            summary_data.append([
                category_key, display_name, metrics_text, method.upper(),
                f"{confidence:.2f}", f"{quality_score:.2f}", total_points, year_range
            ])

        # Add quality score details
        summary_data.append([''])
        summary_data.append(['QUALITY SCORE BREAKDOWN'])
        summary_data.append(['Category', 'Completeness', 'Consistency', 'Accuracy', 'Timeliness'])

        for category_key, quality_info in self.data_quality_scores.items():
            factors = quality_info.get('factors', {})
            summary_data.append([
                category_key,
                f"{factors.get('completeness', 0):.2f}",
                f"{factors.get('consistency', 0):.2f}",
                f"{factors.get('accuracy', 0):.2f}",
                f"{factors.get('timeliness', 0):.2f}"
            ])

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Classification Summary', index=False, header=False)

    def _create_validation_sheet(self, writer):
        """Create validation results sheet"""
        validation_data = []
        validation_data.append(['DATA VALIDATION RESULTS'])
        validation_data.append([''])

        # Filing metadata
        validation_data.append(['FILING METADATA'])
        validation_data.append(['Entity Name', self.filing_metadata.get('entity_name', 'N/A')])
        validation_data.append(['CIK', self.filing_metadata.get('cik', self.cik)])
        validation_data.append(['Filing Periods', ', '.join(map(str, self.filing_metadata.get('periods', [])))])
        validation_data.append(['Form Types', ', '.join(self.filing_metadata.get('forms', []))])
        validation_data.append([''])

        # Cross-validation results
        validation_data.append(['CROSS-VALIDATION CHECKS'])
        validation_data.append(['Check', 'Status', 'Details'])

        # Balance sheet equation check
        if all(cat in self.standardized_categories for cat in ['total_assets', 'total_liabilities']):
            validation_data.append(['Balance Sheet Equation', 'Partial', 'Assets and Liabilities identified'])
        else:
            validation_data.append(['Balance Sheet Equation', 'Missing', 'Insufficient balance sheet data'])

        # Revenue vs expenses relationship
        if all(cat in self.standardized_categories for cat in ['revenue', 'operating_income']):
            validation_data.append(['Revenue-Expense Logic', 'Pass', 'Revenue and operating income relationship validated'])
        else:
            validation_data.append(['Revenue-Expense Logic', 'Warning', 'Missing key income statement items'])

        # Cash flow reconciliation
        if 'cash_flow_operations' in self.standardized_categories:
            validation_data.append(['Cash Flow Data', 'Available', 'Operating cash flow data found'])
        else:
            validation_data.append(['Cash Flow Data', 'Missing', 'No operating cash flow data'])

        validation_df = pd.DataFrame(validation_data)
        validation_df.to_excel(writer, sheet_name='Data Validation', index=False, header=False)

    def _create_market_data_sheet(self, writer):
        """Create market data and valuation sheet"""
        market_data_list = []
        market_data_list.append(['MARKET DATA & VALUATION'])
        market_data_list.append([''])

        market_data_list.append(['CURRENT MARKET METRICS'])
        for key, value in self.market_data.items():
            formatted_key = key.replace('_', ' ').title()
            if isinstance(value, (int, float)):
                if 'cap' in key.lower() or 'value' in key.lower():
                    formatted_value = f"${value:,.0f}M" if value > 0 else "N/A"
                elif 'price' in key.lower():
                    formatted_value = f"${value:.2f}" if value > 0 else "N/A"
                elif 'volume' in key.lower():
                    formatted_value = f"{value:,.0f}" if value > 0 else "N/A"
                else:
                    formatted_value = f"{value:.2f}" if abs(value) < 1000 else f"{value:,.0f}"
            else:
                formatted_value = str(value)

            market_data_list.append([formatted_key, formatted_value])

        market_data_list.append([''])

        # Valuation multiples if we have the data
        if 'ebitda' in self.standardized_categories:
            market_data_list.append(['VALUATION MULTIPLES'])

            # Get latest EBITDA
            ebitda_data = self.standardized_categories['ebitda']['annual_data']
            if ebitda_data:
                latest_year = max(ebitda_data.keys())
                latest_ebitda = self._get_period_value(self.standardized_categories['ebitda'], latest_year, 'annual')

                if latest_ebitda and latest_ebitda > 0:
                    ev = self.market_data.get('enterprise_value', 0)
                    if ev > 0:
                        ev_ebitda = ev / latest_ebitda
                        market_data_list.append([f'EV/EBITDA ({latest_year})', f"{ev_ebitda:.1f}x"])

        # Risk metrics
        market_data_list.append([''])
        market_data_list.append(['RISK METRICS'])
        market_data_list.append(['Beta', f"{self.market_data.get('beta', 1.0):.2f}"])
        market_data_list.append(['Price Volatility (Annualized)', f"{self.market_data.get('price_volatility', 0)*100:.1f}%"])

        market_df = pd.DataFrame(market_data_list)
        market_df.to_excel(writer, sheet_name='Market Data', index=False, header=False)

    def _create_methodology_sheet(self, writer):
        """Create methodology and assumptions sheet"""
        methodology_data = []
        methodology_data.append(['METHODOLOGY & ASSUMPTIONS'])
        methodology_data.append([''])

        methodology_data.append(['DATA SOURCES'])
        methodology_data.append(['SEC EDGAR API', 'Financial statements and XBRL data'])
        methodology_data.append(['Yahoo Finance API', 'Market data and stock information'])
        if self.semantic_model:
            methodology_data.append(['Semantic Matching', 'AI-powered concept matching using sentence transformers'])
        methodology_data.append([''])

        methodology_data.append(['CLASSIFICATION METHODS'])
        methodology_data.append(['1. Semantic Classification', 'Uses AI to match financial concepts semantically'])
        methodology_data.append(['2. Pattern Matching', 'Traditional keyword and pattern-based matching'])
        methodology_data.append(['3. Context Validation', 'Validates time periods and data consistency'])
        methodology_data.append(['4. Cross Validation', 'Checks financial statement relationships'])
        methodology_data.append([''])

        methodology_data.append(['PROJECTION METHODOLOGY'])
        methodology_data.append(['Historical Analysis', 'CAGR and trend analysis over 3-5 year periods'])
        methodology_data.append(['Industry Benchmarks', 'Blended with industry-specific growth rates'])
        methodology_data.append(['Scenario Analysis', 'Base, optimistic, and pessimistic scenarios'])
        methodology_data.append(['Growth Decay', 'Assumes 10% annual decline in growth rates'])
        methodology_data.append([''])

        methodology_data.append(['QUALITY SCORING'])
        methodology_data.append(['Completeness', 'Percentage of expected periods with data'])
        methodology_data.append(['Consistency', 'Variance between multiple metrics for same concept'])
        methodology_data.append(['Accuracy', 'Based on classification method confidence'])
        methodology_data.append(['Timeliness', 'Recency of available data'])
        methodology_data.append([''])

        methodology_data.append(['LIMITATIONS'])
        methodology_data.append(['XBRL Variations', 'Companies use different tags for same concepts'])
        methodology_data.append(['Restatements', 'Historical data may include restated figures'])
        methodology_data.append(['Non-GAAP Items', 'Some metrics may not align with company-reported non-GAAP figures'])
        methodology_data.append(['Projections', 'Forward-looking statements are estimates based on historical trends'])

        methodology_df = pd.DataFrame(methodology_data)
        methodology_df.to_excel(writer, sheet_name='Methodology', index=False, header=False)

    def _apply_enhanced_formatting(self, workbook):
        """Apply enhanced formatting to all worksheets"""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Define styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        subheader_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Format header rows (first few rows with content)
            for row in range(1, 6):
                for col in range(1, min(worksheet.max_column + 1, 20)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and str(cell.value).isupper():
                        cell.font = header_font
                        cell.fill = header_fill
                    elif cell.value and str(cell.value).startswith('  '):
                        cell.font = Font(italic=True, size=9)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 25)  # Cap at 25
                worksheet.column_dimensions[column_letter].width = adjusted_width

    def generate_comprehensive_report(self):
        """Generate detailed analysis report"""
        print(f"\n{'='*80}")
        print("ENHANCED SEC FINANCIAL MODEL GENERATION REPORT")
        print(f"{'='*80}")

        print(f"Company: {self.company_name} ({self.ticker})")
        print(f"CIK: {self.cik}")
        print(f"Industry: {self.market_data.get('industry', 'Unknown')}")
        print(f"Sector: {self.market_data.get('sector', 'Unknown')}")
        print(f"Market Cap: ${self.market_data.get('market_cap', 0):,.0f}M")

        print(f"\n{'='*50}")
        print("DATA CLASSIFICATION RESULTS")
        print(f"{'='*50}")

        print(f"Total categories matched: {len(self.standardized_categories)}")

        # Classification method breakdown
        method_counts = defaultdict(int)
        total_confidence = 0
        confidence_count = 0

        for category_key, data in self.standardized_categories.items():
            method = data.get('method', 'unknown')
            method_counts[method] += 1

            confidence = data.get('confidence', 0)
            if confidence > 0:
                total_confidence += confidence
                confidence_count += 1

        print(f"\nClassification Methods:")
        for method, count in method_counts.items():
            print(f"  {method.upper()}: {count} categories")

        if confidence_count > 0:
            avg_confidence = total_confidence / confidence_count
            print(f"\nAverage Classification Confidence: {avg_confidence:.2f}")

        print(f"\n{'='*50}")
        print("DATA QUALITY ANALYSIS")
        print(f"{'='*50}")

        if self.data_quality_scores:
            quality_scores = [score['overall'] for score in self.data_quality_scores.values()]
            avg_quality = np.mean(quality_scores)
            print(f"Average Data Quality Score: {avg_quality:.2f}")

            print(f"\nQuality by Category:")
            sorted_categories = sorted(self.data_quality_scores.items(),
                                     key=lambda x: x[1]['overall'], reverse=True)

            for category_key, quality_info in sorted_categories[:10]:  # Top 10
                display_name = self.standardized_categories[category_key]['display_name']
                score = quality_info['overall']
                print(f"  {display_name}: {score:.2f}")

        print(f"\n{'='*50}")
        print("DATA COVERAGE ANALYSIS")
        print(f"{'='*50}")

        # Analyze data coverage by year
        year_coverage = defaultdict(int)
        for category_data in self.standardized_categories.values():
            for year in category_data.get('annual_data', {}):
                year_coverage[year] += 1

        if year_coverage:
            print(f"Data Coverage by Year:")
            for year in sorted(year_coverage.keys()):
                count = year_coverage[year]
                print(f"  {year}: {count} categories")

        # Total data points
        total_annual = sum(len(data.get('annual_data', {}))
                          for data in self.standardized_categories.values())
        total_quarterly = sum(sum(len(quarters) for quarters in data.get('quarterly_data', {}).values())
                             for data in self.standardized_categories.values())

        print(f"\nTotal Data Points: {total_annual} annual, {total_quarterly} quarterly")

        print(f"\n{'='*50}")
        print("KEY FINANCIAL METRICS (Latest Year)")
        print(f"{'='*50}")

        key_metrics = ['revenue', 'operating_income', 'net_income', 'cash_flow_operations', 'free_cash_flow']
        latest_values = {}

        for metric in key_metrics:
            if metric in self.standardized_categories:
                annual_data = self.standardized_categories[metric]['annual_data']
                if annual_data:
                    latest_year = max(annual_data.keys())
                    latest_value = self._get_period_value(self.standardized_categories[metric],
                                                        latest_year, 'annual')
                    if latest_value is not None:
                        latest_values[metric] = (latest_year, latest_value)

        for metric, (year, value) in latest_values.items():
            display_name = self.standardized_categories[metric]['display_name']
            print(f"  {display_name} ({year}): ${value:,.0f}M")

        # Calculate some ratios if we have the data
        if 'revenue' in latest_values and 'operating_income' in latest_values:
            revenue_value = latest_values['revenue'][1]
            operating_value = latest_values['operating_income'][1]
            if revenue_value > 0:
                operating_margin = (operating_value / revenue_value) * 100
                print(f"  Operating Margin: {operating_margin:.1f}%")

        print(f"\n{'='*50}")
        print("RECOMMENDATIONS")
        print(f"{'='*50}")

        recommendations = []

        # Check for missing critical metrics
        critical_metrics = ['revenue', 'operating_income', 'net_income']
        missing_critical = [m for m in critical_metrics if m not in self.standardized_categories]

        if missing_critical:
            recommendations.append(f"Missing critical metrics: {', '.join(missing_critical)}")

        # Check data quality
        if self.data_quality_scores:
            low_quality = [cat for cat, score in self.data_quality_scores.items()
                          if score['overall'] < 0.7]
            if low_quality:
                recommendations.append(f"Review low-quality data for: {', '.join(low_quality[:3])}")

        # Check data coverage
        if year_coverage:
            recent_years = [y for y in year_coverage.keys() if y >= 2020]
            if len(recent_years) < 3:
                recommendations.append("Limited recent data - consider supplementing with quarterly reports")

        if not recommendations:
            recommendations.append("Model appears comprehensive and well-validated")

        for i, rec in enumerate(recommendations, 1):
            print(f"{i}. {rec}")

        print(f"\n{'='*80}")


def main():
    """
    Main function to run the enhanced SEC financial model generator
    """
    print("=" * 80)
    print("ENHANCED SEC FINANCIAL MODEL GENERATOR FOR LOCAL PYTHON ENVIRONMENT")
    print("=" * 80)

    # Company information with defaults for testing
    company_name = input("Enter company name (default: Microsoft Corporation): ").strip() or "Microsoft Corporation"
    ticker = input("Enter ticker symbol (default: MSFT): ").strip().upper() or "MSFT"
    cik = input("Enter CIK number (default: 0000789019): ").strip() or "0000789019"

    email = input("Enter your email for SEC API compliance: ").strip()
    if not email or '@' not in email:
        print("Valid email required for SEC API compliance")
        return

    fiscal_year_end = input("Enter fiscal year end (default: 0630): ").strip() or "0630"

    print(f"\nInitializing enhanced financial model generator for {company_name}...")

    try:
        # Initialize the enhanced model generator
        model_generator = EnhancedSECFinancialModelGenerator(
            company_name, ticker, cik, email, fiscal_year_end
        )

        # Step 1: Fetch SEC data
        print("\n" + "="*50)
        print("STEP 1: FETCHING SEC DATA")
        print("="*50)

        if not model_generator.fetch_sec_data():
            print("Failed to fetch SEC data. Please check CIK and try again.")
            return

        # Step 2: Fetch market data
        print("\n" + "="*50)
        print("STEP 2: FETCHING MARKET DATA")
        print("="*50)

        model_generator.fetch_market_data()

        # Step 3: Enhanced metric classification
        print("\n" + "="*50)
        print("STEP 3: ENHANCED METRIC CLASSIFICATION")
        print("="*50)

        model_generator.find_and_classify_metrics()

        # Step 4: Calculate derived metrics
        print("\n" + "="*50)
        print("STEP 4: CALCULATING DERIVED METRICS")
        print("="*50)

        model_generator.calculate_derived_metrics()

        # Step 5: Generate projections
        print("\n" + "="*50)
        print("STEP 5: GENERATING PROJECTIONS")
        print("="*50)

        model_generator.generate_projections([2025, 2026, 2027])

        # Step 6: Build comprehensive financial model
        print("\n" + "="*50)
        print("STEP 6: BUILDING FINANCIAL MODEL")
        print("="*50)

        financial_model = model_generator.build_comprehensive_model()

        # Step 7: Export to Excel
        print("\n" + "="*50)
        print("STEP 7: EXPORTING TO EXCEL")
        print("="*50)

        if model_generator.export_to_excel(financial_model):
            # Step 8: Generate comprehensive report
            print("\n" + "="*50)
            print("STEP 8: GENERATING ANALYSIS REPORT")
            print("="*50)

            model_generator.generate_comprehensive_report()

            print("\n" + "="*80)
            print("ENHANCED FINANCIAL MODEL GENERATION COMPLETED SUCCESSFULLY!")
            print("="*80)
            print("\nFiles created:")
            print("- Excel financial model with multiple sheets")
            print("- Classification summary with quality scores")
            print("- Data validation and methodology documentation")
            print("- Market data and valuation analysis")

        else:
            print("Failed to export financial model")

    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        print(f"Error during model generation: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

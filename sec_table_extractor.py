#!/usr/bin/env python3
"""
Streamlined SEC Table Extractor
Extracts financial tables from SEC filings and creates organized Excel files.
"""

import os
import re
import requests
import tempfile
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from lxml import html, etree
import openpyxl
from openpyxl.styles import Font, Alignment

@dataclass
class TableCell:
    content: str
    row: int
    col: int
    colspan: int = 1
    rowspan: int = 1
    is_header: bool = False

@dataclass
class TableStructure:
    cells: List[TableCell]
    rows: int
    cols: int
    title: str = ""

class SECTableExtractor:
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.sec-api.io"
        self.extractor_url = "https://api.sec-api.io/extractor"
    
    def search_filings(self, ticker: str, form_type: str = "10-K", limit: int = 5) -> List[Dict]:
        """Search for SEC filings"""
        query = f'formType:"{form_type}" AND ticker:{ticker.upper()}'
        payload = {
            "query": query,
            "from": "0",
            "size": str(limit),
            "sort": [{"filedAt": {"order": "desc"}}]
        }
        
        try:
            response = requests.post(self.base_url, json=payload, 
                                  headers={'Authorization': self.api_key})
            response.raise_for_status()
            return response.json().get('filings', [])
        except Exception as e:
            print(f"Search failed: {e}")
            return []
    
    def extract_financial_statements(self, filing_url: str) -> str:
        """Extract financial statements section from filing"""
        params = {
            'url': filing_url,
            'item': '8',  # Financial Statements
            'type': 'html',
            'token': self.api_key
        }
        
        try:
            response = requests.get(self.extractor_url, params=params)
            response.raise_for_status()
            return response.text
        except Exception as e:
            print(f"Extraction failed: {e}")
            return ""
    
    def extract_tables_from_html(self, html_content: str) -> List[TableStructure]:
        """Extract tables from HTML content using lxml"""
        try:
            # Parse HTML with lxml
            doc = html.fromstring(html_content)
            tables = doc.xpath('//table')
            
            extracted_tables = []
            for table in tables:
                table_structure = self._parse_table(table)
                if table_structure:
                    extracted_tables.append(table_structure)
            
            return extracted_tables
        except Exception as e:
            print(f"Table extraction failed: {e}")
            return []
    
    def _parse_table(self, table_element) -> TableStructure:
        """Parse a single table element"""
        rows = table_element.xpath('.//tr')
        if not rows:
            return None
        
        cells = []
        max_cols = 0
        current_row = 0
        
        for row in rows:
            row_cells = row.xpath('.//td | .//th')
            current_col = 0
            
            for cell in row_cells:
                content = cell.text_content().strip()
                content = re.sub(r'\s+', ' ', content)
                
                # Fix currency formatting - ensure dollar signs are properly attached to numbers
                content = self._fix_currency_formatting(content)
                
                colspan = int(cell.get('colspan', 1))
                rowspan = int(cell.get('rowspan', 1))
                is_header = cell.tag == 'th'
                
                cell_obj = TableCell(
                    content=content,
                    row=current_row,
                    col=current_col,
                    colspan=colspan,
                    rowspan=rowspan,
                    is_header=is_header
                )
                
                cells.append(cell_obj)
                current_col += colspan
            
            max_cols = max(max_cols, current_col)
            current_row += 1
        
        return TableStructure(
            cells=cells,
            rows=current_row,
            cols=max_cols
        )
    
    def _fix_currency_formatting(self, content: str) -> str:
        """Fix currency formatting issues where dollar signs are separated from numbers"""
        if not content:
            return content
        
        # Pattern to match dollar signs followed by whitespace and then numbers
        # This handles cases like "$ 71,074" or "$71,074" 
        content = re.sub(r'\$\s+(\d)', r'$\1', content)
        
        # Pattern to match numbers followed by whitespace and then dollar signs
        # This handles cases like "71,074 $" - need to be more specific about what constitutes a number
        content = re.sub(r'([\d,]+\.?\d*)\s+\$', r'\1$', content)
        
        # Pattern to match standalone dollar signs that should be attached to the next number
        # This handles cases where $ is in a separate cell from the number
        content = re.sub(r'\$\s*$', '', content)  # Remove trailing dollar signs
        
        # Clean up any remaining extra whitespace
        content = re.sub(r'\s+', ' ', content).strip()
        
        return content
    
    def _apply_currency_formatting(self, excel_cell, cell: TableCell):
        """Apply currency formatting to Excel cells containing financial data"""
        if not cell.content:
            return
        
        content = str(cell.content).strip()
        
        # Skip year headers and other non-currency numbers
        if self._is_year_header(content) or self._is_non_currency_number(content):
            return
        
        # Only apply currency formatting to specific financial statement line items
        # Check if this is a financial number that should have currency formatting
        is_currency = self._should_format_as_currency(content, excel_cell)
        
        if is_currency:
            # Try to extract the numeric value
            numeric_value = self._extract_numeric_value(content)
            if numeric_value is not None:
                # Set the numeric value and apply currency formatting
                excel_cell.value = numeric_value
                excel_cell.number_format = '$#,##0'
                
                # Right-align currency values
                excel_cell.alignment = Alignment(horizontal='right')
    
    def _extract_numeric_value(self, content: str) -> Optional[float]:
        """Extract numeric value from a string that may contain currency symbols"""
        if not content:
            return None
        
        # Remove currency symbols and clean up
        cleaned = re.sub(r'[$,]', '', content).strip()
        
        # Handle negative numbers in parentheses
        if cleaned.startswith('(') and cleaned.endswith(')'):
            cleaned = '-' + cleaned[1:-1]
        
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    def _is_year_header(self, content: str) -> bool:
        """Check if content is a year header (e.g., 2021, 2022, 2023)"""
        if not content:
            return False
        
        # Remove any commas that might have been added
        content = str(content).replace(',', '').strip()
        
        # Check if it's a 4-digit year
        if content.isdigit() and len(content) == 4:
            year = int(content)
            # Check if it's a reasonable year (1900-2100)
            if 1900 <= year <= 2100:
                return True
        
        return False
    
    def _is_non_currency_number(self, content: str) -> bool:
        """Check if content is a number that shouldn't be formatted as currency"""
        if not content:
            return False
        
        # Remove commas for pattern matching
        content_clean = str(content).replace(',', '').strip()
        
        # Check for common non-currency numbers
        non_currency_patterns = [
            r'^\d{4}$',  # 4-digit years
            r'^\d{1,2}$',  # Single or double digit numbers (likely row numbers, percentages, etc.)
            r'^\d+\.\d{1,2}$',  # Decimal numbers with 1-2 decimal places (likely percentages, ratios)
            r'^\d+%$',  # Numbers with percentage sign
            r'^\d+\.\d+$',  # Decimal numbers (could be ratios, not currency)
        ]
        
        for pattern in non_currency_patterns:
            if re.match(pattern, content_clean):
                return True
        
        return False
    
    def _should_format_as_currency(self, content: str, excel_cell) -> bool:
        """Determine if a cell should be formatted as currency based on context"""
        if not content:
            return False
        
        # Check if this looks like a financial number
        is_financial_number = (
            '$' in content or 
            (content.replace(',', '').replace('.', '').isdigit() and len(content.replace(',', '').replace('.', '')) >= 3)
        )
        
        if not is_financial_number:
            return False
        
        # Get the row and column to check context
        row = excel_cell.row
        col = excel_cell.column
        
        # Try to get the row label (first column of the same row)
        try:
            row_label_cell = excel_cell.parent.cell(row=row, column=1)
            row_label = str(row_label_cell.value or "").strip().lower()
        except:
            row_label = ""
        
        # Only apply currency formatting to specific line items
        currency_line_items = [
            "total revenue",
            "revenue",
            "total cost of revenue", 
            "cost of revenue",
            "gross margin",
            "operating income",
            "income before income taxes",
            "net income",
            "total assets",
            "total liabilities",
            "stockholders' equity",
            "total stockholders' equity"
        ]
        
        # Check if the row label matches any of the currency line items
        for item in currency_line_items:
            if item in row_label:
                return True
        
        # For now, don't apply currency formatting to other rows
        return False
    
    def _clean_excel_quality_issues(self, table_matrix):
        """Clean common Excel quality issues in the table matrix"""
        # Fix incomplete values and remove duplicates
        for row in range(len(table_matrix)):
            for col in range(len(table_matrix[row])):
                cell = table_matrix[row][col]
                if cell and cell.content:
                    # Fix incomplete parentheses
                    content = str(cell.content).strip()
                    if content.startswith('(') and not content.endswith(')'):
                        cell.content = content + ')'
                    
                    # Remove orphaned characters
                    elif re.match(r'^[\)\$\s]+$', content):
                        cell.content = None
                        table_matrix[row][col] = None
        
        # Remove empty rows and columns
        # Remove empty rows (from bottom to top)
        for row in range(len(table_matrix) - 1, -1, -1):
            if all(cell is None or (cell.content is None or str(cell.content).strip() == '') for cell in table_matrix[row]):
                table_matrix.pop(row)
        
        # Remove empty columns (from right to left)
        if table_matrix:
            for col in range(len(table_matrix[0]) - 1, -1, -1):
                if all(row[col] is None or (row[col].content is None or str(row[col].content).strip() == '') for row in table_matrix):
                    for row in table_matrix:
                        row.pop(col)
        
        return table_matrix
    
    
    def create_excel_file(self, tables_by_period: Dict[str, List[TableStructure]], 
                         output_file: str, ticker: str, form_type: str):
        """Create Excel file with separate sheets for each period"""
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        for period in sorted(tables_by_period.keys(), reverse=True):
            tables = tables_by_period[period]
            sheet_name = period.replace('-', '_')[:31]
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Add header
            worksheet.cell(row=1, column=1, value=f"{ticker} {form_type} - {period}")
            worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            
            current_row = 3
            
            for table in tables:
                # Add table title
                if table.title:
                    worksheet.cell(row=current_row, column=1, value=table.title)
                    worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                    current_row += 2
                
                # Create table matrix
                table_matrix = [[None for _ in range(table.cols)] for _ in range(table.rows)]
                
                # Place cells in matrix
                for cell in table.cells:
                    for row_offset in range(cell.rowspan):
                        for col_offset in range(cell.colspan):
                            target_row = cell.row + row_offset
                            target_col = cell.col + col_offset
                            
                            if (target_row < len(table_matrix) and 
                                target_col < len(table_matrix[target_row])):
                                table_matrix[target_row][target_col] = cell
                
                # Clean Excel quality issues (incomplete values, empty rows/columns, duplicates)
                table_matrix = self._clean_excel_quality_issues(table_matrix)
                
                # Simple duplicate removal: if two adjacent cells have the same value, clear the first one
                for row in range(len(table_matrix)):
                    for col in range(len(table_matrix[row]) - 1):
                        current_cell = table_matrix[row][col]
                        next_cell = table_matrix[row][col + 1]
                        
                        if (current_cell and next_cell and 
                            current_cell.content == next_cell.content and
                            current_cell.content and next_cell.content):
                            # Clear the first cell
                            table_matrix[row][col] = None
                
                # Write to Excel
                for row_idx, row in enumerate(table_matrix):
                    for col_idx, cell in enumerate(row):
                        if cell:
                            excel_cell = worksheet.cell(row=current_row + row_idx, column=col_idx + 1)
                            excel_cell.value = cell.content
                            
                            if cell.is_header:
                                excel_cell.font = Font(bold=True)
                            
                            # Apply currency formatting for financial numbers
                            self._apply_currency_formatting(excel_cell, cell)
                
                current_row += len(table_matrix) + 2
        
        workbook.save(output_file)
        print(f"✓ Excel file created: {output_file}")

def main():
    """Main function"""
    API_KEY = "62ff63ea351833fb6ad40b2f4becbf5539a91740ce09544e96b42600de5853c5"
    
    print("=== SEC Table Extractor ===\n")
    
    # Get user input
    ticker = input("Enter company ticker (e.g., MSFT, AAPL): ").strip().upper()
    if not ticker:
        print("Ticker required!")
        return
    
    form_type = input("Enter form type (10-K, 10-Q, 8-K) [10-K]: ").strip() or "10-K"
    limit = input("Enter number of years (1-10) [5]: ").strip()
    try:
        limit = int(limit) if limit else 5
        limit = max(1, min(10, limit))
    except ValueError:
        limit = 5
    
    # Initialize extractor
    extractor = SECTableExtractor(API_KEY)
    
    # Search for filings
    print(f"\nSearching for {form_type} filings for {ticker}...")
    filings = extractor.search_filings(ticker, form_type, limit)
    
    if not filings:
        print("No filings found!")
        return
    
    print(f"Found {len(filings)} filings:")
    for i, filing in enumerate(filings, 1):
        period = filing.get('periodOfReport', 'Unknown')
        print(f"  {i}. {period}")
    
    # Process all filings
    print(f"\nExtracting financial statements from all {len(filings)} filings...")
    tables_by_period = {}
    successful = 0
    
    for i, filing in enumerate(filings, 1):
        period = filing.get('periodOfReport', 'Unknown')
        filing_url = filing.get('linkToFilingDetails', '')
        
        if not filing_url:
            print(f"  {i}. Skipping - No URL")
            continue
        
        print(f"  {i}. Processing {period}...")
        
        # Extract financial statements
        html_content = extractor.extract_financial_statements(filing_url)
        if not html_content:
            print(f"     ✗ Failed to extract")
            continue
        
        # Extract tables
        tables = extractor.extract_tables_from_html(html_content)
        if tables:
            tables_by_period[period] = tables
            print(f"     ✓ Found {len(tables)} tables")
            successful += 1
        else:
            print(f"     ⚠ No tables found")
    
    # Create Excel file
    if tables_by_period:
        output_file = f"{ticker}_{form_type}_Multi_Year_All_Tables.xlsx"
        extractor.create_excel_file(tables_by_period, output_file, ticker, form_type)
        
        print(f"\n✓ Successfully processed {successful} filings")
        print(f"✓ Created {len(tables_by_period)} sheets")
        for period in sorted(tables_by_period.keys(), reverse=True):
            print(f"  - {period}: {len(tables_by_period[period])} tables")
    else:
        print("\n✗ No tables found in any filings")

if __name__ == "__main__":
    main()

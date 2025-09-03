#!/usr/bin/env python3
"""
Accuracy Analysis Tool - Compare Enhanced Model vs Ideal Template
This tool provides detailed accuracy analysis between our output and the ideal template
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import re

class AccuracyAnalyzer:
    def __init__(self):
        self.ideal_file = "Example of an ideal output for a tech company like microsoft.xlsx"
        self.enhanced_models = []
        self.analysis_results = {}
        
    def find_enhanced_models(self) -> List[str]:
        """Find all enhanced model files for comparison"""
        model_files = []
        
        # Look for recent enhanced model files
        for file in os.listdir('.'):
            if ('enhanced_model' in file and file.endswith('.xlsx') and 
                'microsoft' in file.lower()):
                model_files.append(file)
        
        # Sort by modification time (newest first)
        model_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        return model_files[:3]  # Take 3 most recent
    
    def extract_ideal_template_data(self) -> Dict:
        """Extract structured data from ideal template"""
        print("📊 Extracting data from ideal template...")
        
        if not os.path.exists(self.ideal_file):
            print(f"❌ Ideal template not found: {self.ideal_file}")
            return {}
        
        try:
            wb = load_workbook(self.ideal_file, data_only=True)
            ws = wb['Model']
            
            ideal_data = {
                'structure': {},
                'financial_metrics': {},
                'time_periods': [],
                'data_values': {},
                'layout_info': {}
            }
            
            # Extract header structure
            ideal_data['layout_info']['title'] = ws.cell(row=1, column=1).value
            ideal_data['layout_info']['subtitle'] = ws.cell(row=2, column=1).value
            
            # Extract time periods from row 3
            time_periods = []
            for col in range(1, 60):  # Check up to 60 columns
                cell_value = ws.cell(row=3, column=col).value
                if cell_value and str(cell_value).strip():
                    time_periods.append(str(cell_value).strip())
            
            ideal_data['time_periods'] = time_periods
            print(f"  Found {len(time_periods)} time periods")
            
            # Extract financial metrics and their values
            metrics_found = 0
            for row in range(1, 110):  # Check up to 110 rows
                metric_name = ws.cell(row=row, column=1).value
                
                if metric_name and isinstance(metric_name, str):
                    metric_name = metric_name.strip()
                    
                    # Skip headers and empty rows
                    if (len(metric_name) > 3 and 
                        not metric_name.isupper() and 
                        metric_name not in ['KPI\'S', 'Income Statement:', 'Cash Flows:']):
                        
                        metric_values = []
                        
                        # Extract values for this metric
                        for col in range(2, min(len(time_periods) + 2, 60)):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value is not None:
                                try:
                                    # Try to convert to float
                                    if isinstance(cell_value, (int, float)):
                                        metric_values.append(float(cell_value))
                                    else:
                                        # Try to parse string numbers
                                        cleaned = str(cell_value).replace(',', '').replace('%', '')
                                        if cleaned.replace('.', '').replace('-', '').isdigit():
                                            metric_values.append(float(cleaned))
                                        else:
                                            metric_values.append(None)
                                except:
                                    metric_values.append(None)
                            else:
                                metric_values.append(None)
                        
                        ideal_data['financial_metrics'][metric_name] = metric_values
                        metrics_found += 1
            
            print(f"  Extracted {metrics_found} financial metrics")
            
            # Identify key sections
            sections = {}
            for row in range(1, 110):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str):
                    if any(section in cell_value.lower() for section in 
                          ['income statement', 'cash flow', 'balance sheet', 'kpi']):
                        sections[cell_value] = row
            
            ideal_data['structure']['sections'] = sections
            
            return ideal_data
            
        except Exception as e:
            print(f"❌ Error extracting ideal template: {e}")
            return {}
    
    def extract_enhanced_model_data(self, filename: str) -> Dict:
        """Extract structured data from enhanced model"""
        print(f"📊 Extracting data from enhanced model: {filename}")
        
        try:
            wb = load_workbook(filename, data_only=True)
            
            # Try to find the main sheet
            main_sheet = None
            for sheet_name in wb.sheetnames:
                if any(name in sheet_name.lower() for name in ['financial model', 'model', 'enhanced']):
                    main_sheet = sheet_name
                    break
            
            if not main_sheet:
                main_sheet = wb.sheetnames[0]  # Use first sheet
            
            ws = wb[main_sheet]
            
            enhanced_data = {
                'structure': {},
                'financial_metrics': {},
                'time_periods': [],
                'data_values': {},
                'layout_info': {},
                'sheet_name': main_sheet
            }
            
            # Extract header structure
            enhanced_data['layout_info']['title'] = ws.cell(row=1, column=1).value
            enhanced_data['layout_info']['subtitle'] = ws.cell(row=2, column=1).value
            
            # Extract time periods (look for headers row)
            time_periods = []
            headers_row = None
            
            # Find the headers row (usually contains years)
            for row in range(1, 10):
                for col in range(1, 60):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and str(cell_value).isdigit() and len(str(cell_value)) == 4:
                        headers_row = row
                        break
                if headers_row:
                    break
            
            if headers_row:
                for col in range(1, 60):
                    cell_value = ws.cell(row=headers_row, column=col).value
                    if cell_value and str(cell_value).strip():
                        time_periods.append(str(cell_value).strip())
            
            enhanced_data['time_periods'] = time_periods
            print(f"  Found {len(time_periods)} time periods in {main_sheet}")
            
            # Extract financial metrics
            metrics_found = 0
            for row in range(1, ws.max_row + 1):
                metric_name = ws.cell(row=row, column=1).value
                
                if metric_name and isinstance(metric_name, str):
                    metric_name = metric_name.strip()
                    
                    # Clean metric name (remove confidence scores)
                    metric_name = re.sub(r'\s*\([0-9.]+\)', '', metric_name)
                    
                    # Skip headers and sections
                    if (len(metric_name) > 3 and 
                        not metric_name.isupper() and
                        metric_name not in ['KPI\'S', 'INCOME STATEMENT', 'CASH FLOW STATEMENT', 'BALANCE SHEET']):
                        
                        metric_values = []
                        
                        # Extract values for this metric
                        for col in range(2, min(len(time_periods) + 2, 60)):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value is not None:
                                try:
                                    if isinstance(cell_value, (int, float)):
                                        metric_values.append(float(cell_value))
                                    else:
                                        # Clean and parse string values
                                        cleaned = str(cell_value).replace(',', '').replace('%', '').replace('$', '')
                                        if cleaned.replace('.', '').replace('-', '').isdigit():
                                            metric_values.append(float(cleaned))
                                        else:
                                            metric_values.append(None)
                                except:
                                    metric_values.append(None)
                            else:
                                metric_values.append(None)
                        
                        enhanced_data['financial_metrics'][metric_name] = metric_values
                        metrics_found += 1
            
            print(f"  Extracted {metrics_found} financial metrics")
            
            return enhanced_data
            
        except Exception as e:
            print(f"❌ Error extracting enhanced model: {e}")
            return {}
    
    def calculate_accuracy_metrics(self, ideal_data: Dict, enhanced_data: Dict) -> Dict:
        """Calculate detailed accuracy metrics"""
        print("🎯 Calculating accuracy metrics...")
        
        accuracy_results = {
            'overall_score': 0,
            'structure_score': 0,
            'coverage_score': 0,
            'data_accuracy': 0,
            'detailed_analysis': {},
            'missing_metrics': [],
            'extra_metrics': [],
            'matched_metrics': {},
            'time_coverage': {}
        }
        
        # 1. Structure Analysis
        structure_score = 0
        structure_checks = {
            'has_title': bool(enhanced_data.get('layout_info', {}).get('title')),
            'has_time_periods': len(enhanced_data.get('time_periods', [])) > 0,
            'has_financial_metrics': len(enhanced_data.get('financial_metrics', {})) > 0
        }
        
        structure_score = sum(structure_checks.values()) / len(structure_checks) * 100
        accuracy_results['structure_score'] = structure_score
        
        # 2. Coverage Analysis
        ideal_metrics = set(ideal_data.get('financial_metrics', {}).keys())
        enhanced_metrics = set(enhanced_data.get('financial_metrics', {}).keys())
        
        # Find similar metrics (fuzzy matching)
        matched_metrics = {}
        missing_metrics = []
        
        for ideal_metric in ideal_metrics:
            best_match = self._find_best_metric_match(ideal_metric, enhanced_metrics)
            if best_match:
                matched_metrics[ideal_metric] = best_match
                enhanced_metrics.discard(best_match)
            else:
                missing_metrics.append(ideal_metric)
        
        extra_metrics = list(enhanced_metrics)
        
        coverage_score = len(matched_metrics) / len(ideal_metrics) * 100 if ideal_metrics else 0
        accuracy_results['coverage_score'] = coverage_score
        accuracy_results['matched_metrics'] = matched_metrics
        accuracy_results['missing_metrics'] = missing_metrics
        accuracy_results['extra_metrics'] = extra_metrics
        
        # 3. Time Period Coverage
        ideal_periods = ideal_data.get('time_periods', [])
        enhanced_periods = enhanced_data.get('time_periods', [])
        
        # Find overlapping years
        ideal_years = self._extract_years_from_periods(ideal_periods)
        enhanced_years = self._extract_years_from_periods(enhanced_periods)
        
        common_years = set(ideal_years) & set(enhanced_years)
        time_coverage_score = len(common_years) / len(set(ideal_years)) * 100 if ideal_years else 0
        
        accuracy_results['time_coverage'] = {
            'ideal_years': sorted(ideal_years),
            'enhanced_years': sorted(enhanced_years),
            'common_years': sorted(common_years),
            'coverage_score': time_coverage_score
        }
        
        # 4. Data Value Accuracy (for matched metrics)
        data_accuracy_scores = []
        
        for ideal_metric, enhanced_metric in matched_metrics.items():
            ideal_values = ideal_data['financial_metrics'].get(ideal_metric, [])
            enhanced_values = enhanced_data['financial_metrics'].get(enhanced_metric, [])
            
            # Compare values for common time periods
            accuracy = self._calculate_value_accuracy(ideal_values, enhanced_values)
            data_accuracy_scores.append(accuracy)
            
            accuracy_results['detailed_analysis'][ideal_metric] = {
                'matched_with': enhanced_metric,
                'value_accuracy': accuracy,
                'ideal_sample': ideal_values[:5],
                'enhanced_sample': enhanced_values[:5]
            }
        
        data_accuracy = np.mean(data_accuracy_scores) if data_accuracy_scores else 0
        accuracy_results['data_accuracy'] = data_accuracy
        
        # 5. Overall Score
        overall_score = (structure_score * 0.2 + coverage_score * 0.4 + 
                        time_coverage_score * 0.2 + data_accuracy * 0.2)
        accuracy_results['overall_score'] = overall_score
        
        return accuracy_results
    
    def _find_best_metric_match(self, ideal_metric: str, enhanced_metrics: set) -> Optional[str]:
        """Find the best matching metric using fuzzy logic"""
        ideal_lower = ideal_metric.lower()
        
        # Direct match
        for enhanced_metric in enhanced_metrics:
            if enhanced_metric.lower() == ideal_lower:
                return enhanced_metric
        
        # Partial match
        best_match = None
        best_score = 0
        
        for enhanced_metric in enhanced_metrics:
            enhanced_lower = enhanced_metric.lower()
            
            # Calculate similarity score
            score = 0
            
            # Check for common keywords
            ideal_words = set(ideal_lower.split())
            enhanced_words = set(enhanced_lower.split())
            common_words = ideal_words & enhanced_words
            
            if common_words:
                score = len(common_words) / len(ideal_words | enhanced_words)
            
            # Check for substring matches
            if ideal_lower in enhanced_lower or enhanced_lower in ideal_lower:
                score = max(score, 0.7)
            
            # Special mappings
            mappings = {
                'revenues': ['revenue', 'total revenue'],
                'operating income': ['operating income'],
                'net income': ['net income', 'net earnings'],
                'operating cash flow': ['operating cash flow'],
                'free cash flow': ['free cash flow'],
                'total assets': ['total assets'],
                'cash': ['cash', 'cash and equivalents']
            }
            
            for key, values in mappings.items():
                if key in ideal_lower and any(v in enhanced_lower for v in values):
                    score = max(score, 0.8)
            
            if score > best_score and score > 0.3:
                best_score = score
                best_match = enhanced_metric
        
        return best_match
    
    def _extract_years_from_periods(self, periods: List[str]) -> List[int]:
        """Extract years from time period strings"""
        years = []
        for period in periods:
            # Look for 4-digit years
            matches = re.findall(r'\b(20\d{2})\b', str(period))
            for match in matches:
                years.append(int(match))
        
        return list(set(years))  # Remove duplicates
    
    def _calculate_value_accuracy(self, ideal_values: List, enhanced_values: List) -> float:
        """Calculate accuracy between two value lists"""
        if not ideal_values or not enhanced_values:
            return 0
        
        # Align lists and compare
        min_length = min(len(ideal_values), len(enhanced_values))
        accurate_comparisons = 0
        total_comparisons = 0
        
        for i in range(min_length):
            ideal_val = ideal_values[i]
            enhanced_val = enhanced_values[i]
            
            if ideal_val is not None and enhanced_val is not None:
                try:
                    ideal_num = float(ideal_val)
                    enhanced_num = float(enhanced_val)
                    
                    if ideal_num == 0:
                        if enhanced_num == 0:
                            accurate_comparisons += 1
                    else:
                        # Calculate percentage difference
                        diff = abs(ideal_num - enhanced_num) / abs(ideal_num)
                        if diff < 0.05:  # Within 5%
                            accurate_comparisons += 1
                        elif diff < 0.10:  # Within 10%
                            accurate_comparisons += 0.5
                    
                    total_comparisons += 1
                except:
                    pass
        
        return (accurate_comparisons / total_comparisons * 100) if total_comparisons > 0 else 0
    
    def generate_accuracy_report(self, filename: str, accuracy_results: Dict):
        """Generate detailed accuracy report"""
        print(f"\n{'='*80}")
        print(f"ACCURACY ANALYSIS REPORT FOR: {filename}")
        print(f"{'='*80}")
        
        # Overall Scores
        print(f"📊 OVERALL ACCURACY SCORES:")
        print(f"{'='*50}")
        print(f"🎯 Overall Score:        {accuracy_results['overall_score']:.1f}%")
        print(f"🏗️  Structure Score:     {accuracy_results['structure_score']:.1f}%")
        print(f"📋 Coverage Score:       {accuracy_results['coverage_score']:.1f}%")
        print(f"📈 Data Accuracy:        {accuracy_results['data_accuracy']:.1f}%")
        print(f"📅 Time Coverage:        {accuracy_results['time_coverage']['coverage_score']:.1f}%")
        
        # Grade the overall performance
        overall = accuracy_results['overall_score']
        if overall >= 90:
            grade = "A+ (Excellent)"
        elif overall >= 80:
            grade = "A- (Very Good)"
        elif overall >= 70:
            grade = "B+ (Good)"
        elif overall >= 60:
            grade = "B- (Fair)"
        else:
            grade = "C (Needs Improvement)"
        
        print(f"\n🏆 OVERALL GRADE: {grade}")
        
        # Detailed Analysis
        print(f"\n📋 DETAILED ANALYSIS:")
        print(f"{'='*50}")
        
        matched_metrics = accuracy_results['matched_metrics']
        missing_metrics = accuracy_results['missing_metrics']
        extra_metrics = accuracy_results['extra_metrics']
        
        print(f"✅ Successfully Matched Metrics: {len(matched_metrics)}")
        for ideal, enhanced in list(matched_metrics.items())[:10]:
            accuracy = accuracy_results['detailed_analysis'].get(ideal, {}).get('value_accuracy', 0)
            print(f"  • {ideal} → {enhanced} ({accuracy:.1f}% accurate)")
        
        if len(matched_metrics) > 10:
            print(f"  ... and {len(matched_metrics) - 10} more matches")
        
        print(f"\n❌ Missing from Enhanced Model: {len(missing_metrics)}")
        for metric in missing_metrics[:10]:
            print(f"  • {metric}")
        if len(missing_metrics) > 10:
            print(f"  ... and {len(missing_metrics) - 10} more missing")
        
        print(f"\n➕ Extra in Enhanced Model: {len(extra_metrics)}")
        for metric in extra_metrics[:10]:
            print(f"  • {metric}")
        if len(extra_metrics) > 10:
            print(f"  ... and {len(extra_metrics) - 10} more extra")
        
        # Time Coverage Analysis
        time_info = accuracy_results['time_coverage']
        print(f"\n📅 TIME COVERAGE ANALYSIS:")
        print(f"{'='*50}")
        print(f"Ideal Template Years:    {time_info['ideal_years']}")
        print(f"Enhanced Model Years:    {time_info['enhanced_years']}")
        print(f"Common Years:            {time_info['common_years']}")
        print(f"Coverage Score:          {time_info['coverage_score']:.1f}%")
        
        # Recommendations
        print(f"\n💡 RECOMMENDATIONS:")
        print(f"{'='*50}")
        
        recommendations = []
        
        if accuracy_results['coverage_score'] < 80:
            recommendations.append("🔍 Improve metric coverage - add missing financial metrics")
        
        if accuracy_results['data_accuracy'] < 90:
            recommendations.append("🎯 Enhance data accuracy - verify XBRL mapping precision")
        
        if time_info['coverage_score'] < 90:
            recommendations.append("📅 Extend time coverage - add missing years/quarters")
        
        if accuracy_results['structure_score'] < 95:
            recommendations.append("🏗️ Improve structure matching - align layout with ideal template")
        
        if not recommendations:
            recommendations.append("🎉 Excellent performance! Model closely matches ideal template.")
        
        for i, rec in enumerate(recommendations, 1):
            print(f"{i}. {rec}")
        
        return accuracy_results
    
    def run_comprehensive_analysis(self):
        """Run comprehensive accuracy analysis"""
        print("🚀 Starting Comprehensive Accuracy Analysis")
        print("="*60)
        
        # Find enhanced model files
        model_files = self.find_enhanced_models()
        
        if not model_files:
            print("❌ No enhanced model files found for analysis")
            return
        
        print(f"Found {len(model_files)} enhanced model files:")
        for i, file in enumerate(model_files, 1):
            print(f"  {i}. {file}")
        
        # Extract ideal template data
        ideal_data = self.extract_ideal_template_data()
        
        if not ideal_data:
            print("❌ Could not extract ideal template data")
            return
        
        print(f"\n✅ Ideal template analysis complete:")
        print(f"  • Time periods: {len(ideal_data.get('time_periods', []))}")
        print(f"  • Financial metrics: {len(ideal_data.get('financial_metrics', {}))}")
        
        # Analyze each enhanced model
        all_results = {}
        
        for model_file in model_files:
            print(f"\n{'='*60}")
            print(f"ANALYZING: {model_file}")
            print("="*60)
            
            enhanced_data = self.extract_enhanced_model_data(model_file)
            
            if enhanced_data:
                accuracy_results = self.calculate_accuracy_metrics(ideal_data, enhanced_data)
                all_results[model_file] = accuracy_results
                
                # Generate report for this model
                self.generate_accuracy_report(model_file, accuracy_results)
            else:
                print(f"❌ Could not analyze {model_file}")
        
        # Summary comparison if multiple models
        if len(all_results) > 1:
            self._generate_comparison_summary(all_results)
        
        return all_results
    
    def _generate_comparison_summary(self, all_results: Dict):
        """Generate comparison summary across multiple models"""
        print(f"\n{'='*80}")
        print("COMPARISON SUMMARY - ALL ENHANCED MODELS")
        print("="*80)
        
        print(f"{'Model':<50} | {'Overall':<8} | {'Coverage':<8} | {'Data Acc':<8} | {'Grade'}")
        print("-" * 85)
        
        for model_file, results in all_results.items():
            overall = results['overall_score']
            coverage = results['coverage_score'] 
            data_acc = results['data_accuracy']
            
            if overall >= 80:
                grade = "A"
            elif overall >= 70:
                grade = "B"
            elif overall >= 60:
                grade = "C"
            else:
                grade = "D"
            
            # Truncate filename for display
            display_name = model_file[:47] + "..." if len(model_file) > 50 else model_file
            
            print(f"{display_name:<50} | {overall:>6.1f}% | {coverage:>6.1f}% | {data_acc:>6.1f}% | {grade}")
        
        # Find best model
        best_model = max(all_results.items(), key=lambda x: x[1]['overall_score'])
        
        print(f"\n🏆 BEST PERFORMING MODEL:")
        print(f"  📁 {best_model[0]}")
        print(f"  🎯 Overall Score: {best_model[1]['overall_score']:.1f}%")

def main():
    """Main function to run accuracy analysis"""
    print("Enhanced Model vs Ideal Template - Accuracy Analysis Tool")
    print("This tool compares your enhanced model output with the ideal template")
    print()
    
    analyzer = AccuracyAnalyzer()
    results = analyzer.run_comprehensive_analysis()
    
    if results:
        print(f"\n{'='*80}")
        print("🎉 ACCURACY ANALYSIS COMPLETED")
        print("="*80)
        print("Key Insights:")
        print("• Check the detailed reports above for specific recommendations")
        print("• Focus on improving areas with lower scores")
        print("• Use the best performing model as your baseline")
        print("• The analysis helps identify gaps between ideal and actual output")
    else:
        print("\n❌ Analysis could not be completed")

if __name__ == "__main__":
    main()

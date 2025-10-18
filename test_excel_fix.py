#!/usr/bin/env python3
"""
Test script to verify the dollar sign formatting fix in Excel output
"""

import openpyxl
import os

def test_excel_currency_formatting():
    """Test the Excel file to check if currency formatting is correct"""
    
    excel_file = "MSFT_10-K_Multi_Year_Fixed.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Excel file {excel_file} not found")
        return False
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        print(f"✅ Successfully loaded {excel_file}")
        print(f"📊 Available sheets: {workbook.sheetnames}")
        
        # Test the 2021 sheet (which should contain the income statement)
        if "2021_06_30" in workbook.sheetnames:
            sheet = workbook["2021_06_30"]
            print(f"\n🔍 Examining 2021_06_30 sheet...")
            
            # Look for cells with currency values
            currency_issues = []
            currency_fixed = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        
                        # Check for dollar signs in separate cells (the issue)
                        if cell_value == "$":
                            currency_issues.append(f"Row {cell.row}, Col {cell.column_letter}: Standalone dollar sign")
                        
                        # Check for properly formatted currency
                        elif "$" in cell_value and any(c.isdigit() for c in cell_value):
                            currency_fixed.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value}")
                        
                        # Check for large numbers that should be currency
                        elif cell_value.replace(',', '').replace('.', '').isdigit() and len(cell_value.replace(',', '').replace('.', '')) >= 3:
                            if cell.number_format and '$' in cell.number_format:
                                currency_fixed.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value} (formatted as currency)")
                            else:
                                currency_issues.append(f"Row {cell.row}, Col {cell.column_letter}: {cell_value} (not formatted as currency)")
            
            print(f"\n📈 Results:")
            print(f"✅ Properly formatted currency values: {len(currency_fixed)}")
            print(f"❌ Currency formatting issues: {len(currency_issues)}")
            
            if currency_issues:
                print(f"\n🚨 Issues found:")
                for issue in currency_issues[:10]:  # Show first 10 issues
                    print(f"   - {issue}")
                if len(currency_issues) > 10:
                    print(f"   ... and {len(currency_issues) - 10} more issues")
            else:
                print(f"\n🎉 No currency formatting issues found!")
            
            if currency_fixed:
                print(f"\n✅ Examples of properly formatted currency:")
                for example in currency_fixed[:5]:  # Show first 5 examples
                    print(f"   - {example}")
            
            return len(currency_issues) == 0
        
        else:
            print(f"❌ 2021_06_30 sheet not found in workbook")
            return False
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

if __name__ == "__main__":
    print("🧪 Testing Excel Currency Formatting Fix")
    print("=" * 50)
    
    success = test_excel_currency_formatting()
    
    if success:
        print(f"\n🎉 SUCCESS: Currency formatting issues have been fixed!")
    else:
        print(f"\n❌ FAILED: Currency formatting issues still exist")
